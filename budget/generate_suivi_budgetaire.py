"""
Génère le fichier Excel de suivi budgétaire du Monique Festival.
9 onglets : Synthèse, Crowdfunding, Billetterie, F&B, Ateliers, Artistes, Logistique, Frais divers, Trésorerie.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# === Styles ===
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12, color="2F5496")
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
RECETTE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
DEPENSE_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
EUR_FORMAT = '#,##0 €'
DATE_FORMAT = 'DD/MM/YYYY'

# Dictionnaire pour stocker les lignes de totaux (pour la Synthèse)
totals = {}


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    if fmt == "eur":
        cell.number_format = EUR_FORMAT
    elif fmt == "date":
        cell.number_format = DATE_FORMAT
    return cell


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)


def section_title(ws, row, title, max_col):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    return row + 1


def total_row(ws, row, label, col_ranges, max_col):
    ws.cell(row=row, column=1, value=label)
    for col, (sr, er) in col_ranges.items():
        cl = get_column_letter(col)
        ws.cell(row=row, column=col).value = f"=SUM({cl}{sr}:{cl}{er})"
        ws.cell(row=row, column=col).number_format = EUR_FORMAT
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).font = TOTAL_FONT
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
    return row


def data_block(ws, r, headers_list, data_rows, empty_count, col_count, eur_cols, ecart_formula=None, date_col=None):
    """Écrit un bloc de données : en-têtes + données + lignes vides. Retourne (start_data, end_data, r)."""
    for c, h in enumerate(headers_list, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, col_count)
    r += 1
    start = r

    for row_data in data_rows:
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val if val != "" else None)
        for ec in eur_cols:
            style_data_cell(ws, r, ec, "eur")
        if date_col:
            style_data_cell(ws, r, date_col, "date")
        if ecart_formula:
            ws.cell(row=r, column=ecart_formula[0]).value = ecart_formula[1].format(r=r)
            style_data_cell(ws, r, ecart_formula[0], "eur")
        for c in range(1, col_count + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    for _ in range(empty_count):
        if date_col:
            style_data_cell(ws, r, date_col, "date")
        for ec in eur_cols:
            style_data_cell(ws, r, ec, "eur")
        if ecart_formula:
            ws.cell(row=r, column=ecart_formula[0]).value = ecart_formula[1].format(r=r)
            style_data_cell(ws, r, ecart_formula[0], "eur")
        for c in range(1, col_count + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    end = r - 1
    return start, end, r


# ============================================================
# CROWDFUNDING
# ============================================================
ws = wb.active
ws.title = "Crowdfunding"
NC = 5

r = section_title(ws, 1, "CROWDFUNDING — Recettes HelloAsso", NC)
r += 1
data = [
    ("", "Objectif crowdfunding (palier 3)", 8000, None, None),
]
s, e, r = data_block(ws, r,
    ["Date", "Description", "Prévu", "Réel", "Écart"],
    data, 10, NC, [3, 4, 5], ecart_formula=(5, "=D{r}-C{r}"), date_col=1)
r += 1
tr = total_row(ws, r, "TOTAL CROWDFUNDING", {3: (s, e), 4: (s, e), 5: (s, e)}, NC)
totals["crowd_rec"] = (tr, "C", "D")  # row, col_prevu, col_reel
auto_width(ws)


# ============================================================
# BILLETTERIE
# ============================================================
ws = wb.create_sheet("Billetterie")
NC = 7

r = section_title(ws, 1, "BILLETTERIE — Recettes", NC)
r += 1

headers = ["Type", "Tarif", "Nb prévu", "Recettes prévues", "Nb réel", "Recettes réelles", "Écart"]
for c, h in enumerate(headers, 1):
    ws.cell(row=r, column=c, value=h)
style_header_row(ws, r, NC)
r += 1

billets = [
    ("Early bird pass 3 jours", 20, 120),
    ("Pass 3 jours", 25, 480),
    ("Vendredi", 10, 150),
    ("Samedi", 15, 300),
    ("Dimanche", 15, 150),
]
s = r
for nom, tarif, nb in billets:
    ws.cell(row=r, column=1, value=nom)
    style_data_cell(ws, r, 2, "eur").value = tarif
    ws.cell(row=r, column=3, value=nb)
    ws.cell(row=r, column=4).value = f"=B{r}*C{r}"
    style_data_cell(ws, r, 4, "eur")
    style_data_cell(ws, r, 5)
    ws.cell(row=r, column=6).value = f"=B{r}*E{r}"
    style_data_cell(ws, r, 6, "eur")
    ws.cell(row=r, column=7).value = f"=F{r}-D{r}"
    style_data_cell(ws, r, 7, "eur")
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1
e = r - 1
r += 1
tr = total_row(ws, r, "TOTAL BILLETTERIE", {3: (s, e), 4: (s, e), 5: (s, e), 6: (s, e), 7: (s, e)}, NC)
totals["bill_rec"] = (tr, "D", "F")
auto_width(ws)


# ============================================================
# F&B (Buvettes + Restauration)
# ============================================================
ws = wb.create_sheet("F&B")
NC = 6
headers = ["Date", "Description", "Type", "Prévu", "Réel", "Écart"]

r = section_title(ws, 1, "F&B — Buvettes + Restauration", NC)
r += 1

# Recettes
r = section_title(ws, r, "Recettes", NC)
r += 1
fb_rec = [
    ("", "Buvette — Vendredi", "Recette", 4238, None, None),
    ("", "Buvette — Samedi", "Recette", 12244, None, None),
    ("", "Buvette — Dimanche", "Recette", 4238, None, None),
    ("", "Restauration — Vendredi", "Recette", 2398, None, None),
    ("", "Restauration — Samedi", "Recette", 6926, None, None),
    ("", "Restauration — Dimanche", "Recette", 2398, None, None),
]
s_rec, e_rec, r = data_block(ws, r, headers, fb_rec, 0, NC, [4, 5, 6], ecart_formula=(6, "=E{r}-D{r}"))
r += 1
tr_rec = total_row(ws, r, "TOTAL RECETTES F&B", {4: (s_rec, e_rec), 5: (s_rec, e_rec), 6: (s_rec, e_rec)}, NC)
totals["fb_rec"] = (tr_rec, "D", "E")

r += 2
# Dépenses
r = section_title(ws, r, "Dépenses", NC)
r += 1
fb_dep = [
    ("", "Achats buvette (bières, softs, vins…)", "Dépense", 4440, None, None),
    ("", "Achats restauration (matières premières)", "Dépense", 2930, None, None),
    ("", "Restauration artistes (prix coûtant)", "Dépense", 555, None, None),
]
s_dep, e_dep, r = data_block(ws, r, headers, fb_dep, 5, NC, [4, 5, 6], ecart_formula=(6, "=E{r}-D{r}"), date_col=1)
r += 1
tr_dep = total_row(ws, r, "TOTAL DÉPENSES F&B", {4: (s_dep, e_dep), 5: (s_dep, e_dep), 6: (s_dep, e_dep)}, NC)
totals["fb_dep"] = (tr_dep, "D", "E")
auto_width(ws)


# ============================================================
# ATELIERS
# ============================================================
ws = wb.create_sheet("Ateliers")
NC = 6
headers = ["Date", "Description", "Type", "Prévu", "Réel", "Écart"]

r = section_title(ws, 1, "ATELIERS — Recettes & Dépenses", NC)
r += 1

# Recettes (cotisations participants)
r = section_title(ws, r, "Recettes (cotisations participants)", NC)
r += 1
at_rec = [
    ("", "Masterclass chant lyrique (4 élèves × 60 €)", "Recette", 240, None, None),
    ("", "Atelier écriture (15 pers × 50 € + 150 € lycée)", "Recette", 900, None, None),
    ("", "Atelier musique & théâtre (30 pers × 50 €)", "Recette", 1500, None, None),
    ("", "Atelier DJ (15 pers × 50 €)", "Recette", 750, None, None),
]
s_rec, e_rec, r = data_block(ws, r, headers, at_rec, 3, NC, [4, 5, 6], ecart_formula=(6, "=E{r}-D{r}"), date_col=1)
r += 1
tr_rec = total_row(ws, r, "TOTAL RECETTES ATELIERS", {4: (s_rec, e_rec), 5: (s_rec, e_rec), 6: (s_rec, e_rec)}, NC)
totals["at_rec"] = (tr_rec, "D", "E")

r += 2
# Dépenses (intervenants)
r = section_title(ws, r, "Dépenses (intervenants ateliers)", NC)
r += 1
at_dep = [
    ("", "Rép. Fabulo (4 art. x 2 jours x 100€)", "Dépense", 800, None, None),
    ("", "Rép. Masterclass chant — Romane (incl. Fabulo)", "Dépense", 0, None, None),
    ("", "Rép. Atelier écriture — Sylvain (2 jours x 100€)", "Dépense", 200, None, None),
    ("", "Rép. Atelier DJ — Jeff (2 jours x 100€)", "Dépense", 200, None, None),
]
s_dep, e_dep, r = data_block(ws, r, headers, at_dep, 3, NC, [4, 5, 6], ecart_formula=(6, "=E{r}-D{r}"), date_col=1)
r += 1
tr_dep = total_row(ws, r, "TOTAL DÉPENSES ATELIERS", {4: (s_dep, e_dep), 5: (s_dep, e_dep), 6: (s_dep, e_dep)}, NC)
totals["at_dep"] = (tr_dep, "D", "E")
auto_width(ws)


# ============================================================
# ARTISTES
# ============================================================
ws = wb.create_sheet("Artistes")
NC = 14
headers = ["Nom", "Catégorie", "Atelier", "Nb cachets", "Nb rép.",
           "Coût cachets prévu", "Coût rép. prévu", "Défraiement prévu", "Total prévu",
           "Paiement réel", "Défraiement réel", "Total réel", "Date paiement", "Écart"]

r = section_title(ws, 1, "ARTISTES — Dépenses (cachet 250€, rép. 100€, défr. 200€ max)", NC)
r += 1
for c, h in enumerate(headers, 1):
    ws.cell(row=r, column=c, value=h)
style_header_row(ws, r, NC)
r += 1

# (nom, catégorie, atelier, nb_cachets, nb_rep, cout_cachets, cout_rep, défraiement)
# Cachet spectacle = 250€ chargé, Répétition = 100€ chargé, Défr = 200€ max
artistes = [
    # MUSIQUE
    ("George Cabaret", "Musique", "", 1, 0, 250, 0, 200),
    ("Loretta", "Musique", "", 1, 0, 250, 0, 200),
    ("? (Ven 19h30)", "Musique", "", 1, 0, 250, 0, 200),
    ("Manu (Wambo)", "Musique", "", 1, 0, 250, 0, 200),
    ("Jeff The Fool", "Musique", "Atelier DJ", 0, 2, 0, 200, 200),
    ("Wet Enough (5 art.)", "Musique — Coll", "", 5, 0, 1250, 0, 1000),
    ("Osmosis (3 art.)", "Musique — Coll", "", 3, 0, 750, 0, 600),
    ("Léopold (famille)", "Musique", "Atelier DJ", 0, 2, 0, 0, 0),
    ("Luther (famille)", "Musique", "", 0, 0, 0, 0, 0),
    # FABULO
    ("Marius (Fabulo)", "Fabulo — Coll", "Musique & théâtre", 2, 2, 500, 200, 200),
    ("Maya (Fabulo)", "Fabulo — Coll", "Musique & théâtre", 2, 2, 500, 200, 200),
    ("Raphaëlle (Fabulo)", "Fabulo — Coll", "Musique & théâtre", 2, 2, 500, 200, 200),
    ("Romane Cabaret", "Fabulo — Coll", "Masterclass chant", 4, 2, 1000, 200, 200),
    ("Lorraine (famille)", "Fabulo", "Musique & théâtre", 0, 0, 0, 0, 0),
    # CLASSIQUE
    ("Violoncelliste 1 (Marc T.)", "Classique", "", 2, 0, 500, 0, 200),
    ("Violoncelliste 2", "Classique", "", 2, 0, 500, 0, 200),
    ("Violoncelliste 3", "Classique", "", 2, 0, 500, 0, 200),
    ("Violoncelliste 4", "Classique", "", 2, 0, 500, 0, 200),
    ("Volodia Van Keulen", "Classique", "", 2, 0, 500, 0, 200),
    ("Pianiste (Augustin)", "Classique", "", 1, 0, 250, 0, 200),
    # THÉÂTRE — confirmé
    ("Sylvain Septours", "Théâtre", "Atelier écriture", 0, 2, 0, 200, 200),
    ("Comédien 1 (Treize Clique)", "Théâtre — Coll", "", 1, 0, 250, 0, 200),
    ("Comédien 2 (Treize Clique)", "Théâtre — Coll", "", 1, 0, 250, 0, 200),
    ("Comédien 3 (Treize Clique)", "Théâtre — Coll", "", 1, 0, 250, 0, 200),
    ("Lewis (famille)", "Théâtre", "Éloquence", 0, 0, 0, 0, 0),
    # THÉÂTRE — appel à projet (locaux, pas de défraiement)
    ("Cie théâtre 1 (4 art.)", "Théâtre — AP", "", 4, 0, 1000, 0, 0),
    ("Cie théâtre 2 (4 art.)", "Théâtre — AP", "", 4, 0, 1000, 0, 0),
    ("Cie théâtre 3 (4 art.)", "Théâtre — AP", "", 4, 0, 1000, 0, 0),
    ("Cie théâtre 4 (4 art.)", "Théâtre — AP", "", 4, 0, 1000, 0, 0),
    # NOUVEAUX (Line Up)
    ("Fenêtres des heures (2 art.)", "Nouveau", "", 2, 0, 500, 0, 400),
    ("Fannelie", "Nouveau", "", 1, 0, 250, 0, 200),
    ("Karina Ziegler (3 art.)", "Nouveau — Coll", "", 3, 0, 750, 0, 600),
]

s = r
for nom, cat, atelier, nb_cachets, nb_rep, cout_cachets, cout_rep, defrai in artistes:
    ws.cell(row=r, column=1, value=nom)
    ws.cell(row=r, column=2, value=cat)
    ws.cell(row=r, column=3, value=atelier if atelier else None)
    ws.cell(row=r, column=4, value=nb_cachets)
    ws.cell(row=r, column=5, value=nb_rep)
    style_data_cell(ws, r, 6, "eur").value = cout_cachets
    style_data_cell(ws, r, 7, "eur").value = cout_rep
    style_data_cell(ws, r, 8, "eur").value = defrai
    ws.cell(row=r, column=9).value = f"=F{r}+G{r}+H{r}"
    style_data_cell(ws, r, 9, "eur")
    style_data_cell(ws, r, 10, "eur")
    style_data_cell(ws, r, 11, "eur")
    ws.cell(row=r, column=12).value = f"=J{r}+K{r}"
    style_data_cell(ws, r, 12, "eur")
    style_data_cell(ws, r, 13, "date")
    ws.cell(row=r, column=14).value = f"=L{r}-I{r}"
    style_data_cell(ws, r, 14, "eur")
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1
e = r - 1
r += 1
tr = total_row(ws, r, "TOTAL ARTISTES", {
    6: (s, e), 7: (s, e), 8: (s, e), 9: (s, e), 10: (s, e), 11: (s, e), 12: (s, e), 14: (s, e)
}, NC)
totals["art_dep"] = (tr, "I", "L")  # Total prévu col I, Total réel col L
auto_width(ws)


# ============================================================
# LOGISTIQUE
# ============================================================
ws = wb.create_sheet("Logistique")
NC = 5

r = section_title(ws, 1, "LOGISTIQUE — Dépenses", NC)
r += 1
logistique = [
    ("", "Location salle / terrain", 3000, None, None),
    ("", "Sonorisation", 3000, None, None),
    ("", "Éclairage / lumière", 2000, None, None),
    ("", "Scène / podium", 2000, None, None),
    ("", "Toilettes / sanitaires", 1500, None, None),
    ("", "Électricité / groupe électrogène", 1500, None, None),
    ("", "Barrières / sécurité", 1000, None, None),
    ("", "Mobilier (tables, chaises, tentes)", 1000, None, None),
]
s, e, r = data_block(ws, r,
    ["Date", "Description", "Prévu", "Réel", "Écart"],
    logistique, 5, NC, [3, 4, 5], ecart_formula=(5, "=D{r}-C{r}"), date_col=1)
r += 1
tr = total_row(ws, r, "TOTAL LOGISTIQUE", {3: (s, e), 4: (s, e), 5: (s, e)}, NC)
totals["log_dep"] = (tr, "C", "D")
auto_width(ws)


# ============================================================
# FRAIS DIVERS
# ============================================================
ws = wb.create_sheet("Frais divers")
NC = 6
headers = ["Date", "Description", "Type", "Prévu", "Réel", "Écart"]

r = section_title(ws, 1, "FRAIS DIVERS — Recettes & Dépenses", NC)
r += 1

# Recettes
r = section_title(ws, r, "Recettes (subventions, dons, sponsors)", NC)
r += 1
s_rec, e_rec, r = data_block(ws, r, headers, [], 5, NC, [4, 5, 6], ecart_formula=(6, "=E{r}-D{r}"), date_col=1)
r += 1
tr_rec = total_row(ws, r, "TOTAL RECETTES DIVERS", {4: (s_rec, e_rec), 5: (s_rec, e_rec), 6: (s_rec, e_rec)}, NC)
totals["div_rec"] = (tr_rec, "D", "E")

r += 2
# Dépenses
r = section_title(ws, r, "Dépenses (SACEM, communication, assurance, admin, imprévus)", NC)
r += 1
frais = [
    ("", "SACEM Billetterie (8,8%)", "Dépense", 1993, None, None),
    ("", "SACEM Restauration/Buvettes (4,4%)", "Dépense", 1427, None, None),
    ("", "Assurance événement", "Dépense", 1000, None, None),
    ("", "Communication (affiches, flyers, vidéo)", "Dépense", 1000, None, None),
    ("", "Frais administratifs", "Dépense", 500, None, None),
    ("", "Imprévus", "Dépense", 500, None, None),
]
s_dep, e_dep, r = data_block(ws, r, headers, frais, 5, NC, [4, 5, 6], ecart_formula=(6, "=E{r}-D{r}"), date_col=1)
r += 1
tr_dep = total_row(ws, r, "TOTAL DÉPENSES DIVERS", {4: (s_dep, e_dep), 5: (s_dep, e_dep), 6: (s_dep, e_dep)}, NC)
totals["div_dep"] = (tr_dep, "D", "E")
auto_width(ws)


# ============================================================
# TRÉSORERIE
# ============================================================
ws = wb.create_sheet("Trésorerie")
NC = 6

r = section_title(ws, 1, "TRÉSORERIE — Flux par date", NC)
r += 1

# Solde initial
ws.cell(row=r, column=1, value="Solde initial")
style_data_cell(ws, r, 4, "eur").value = 0
style_data_cell(ws, r, 6, "eur").value = 0
for c in range(1, NC + 1):
    ws.cell(row=r, column=c).font = TOTAL_FONT
    ws.cell(row=r, column=c).fill = TOTAL_FILL
    ws.cell(row=r, column=c).border = THIN_BORDER
solde_row = r
r += 1

headers = ["Date", "Catégorie", "Description", "Entrée", "Sortie", "Solde cumulé"]
for c, h in enumerate(headers, 1):
    ws.cell(row=r, column=c, value=h)
style_header_row(ws, r, NC)
r += 1

s = r
for i in range(30):
    style_data_cell(ws, r, 1, "date")
    style_data_cell(ws, r, 4, "eur")
    style_data_cell(ws, r, 5, "eur")
    prev = solde_row if i == 0 else r - 1
    ws.cell(row=r, column=6).value = f"=F{prev}+D{r}-E{r}"
    style_data_cell(ws, r, 6, "eur")
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 1
e = r - 1

r += 1
total_row(ws, r, "TOTAUX", {4: (s, e), 5: (s, e)}, NC)
ws.cell(row=r, column=6).value = f"=F{e}"
style_data_cell(ws, r, 6, "eur")

dv = DataValidation(type="list", formula1='"Crowdfunding,Billetterie,F&B,Ateliers,Artistes,Logistique,Frais divers"')
dv.error = "Choisir une catégorie"
ws.add_data_validation(dv)
dv.add(f"B{s}:B{e}")
auto_width(ws)


# ============================================================
# SYNTHÈSE (premier onglet, créé en dernier pour avoir les bonnes refs)
# ============================================================
ws = wb.create_sheet("Synthèse", 0)
NC = 10

r = section_title(ws, 1, "SYNTHÈSE BUDGÉTAIRE — Monique Festival", NC)
r += 1
headers = ["Catégorie", "Recettes prévues", "Recettes réelles", "Écart recettes",
           "Dépenses prévues", "Dépenses réelles", "Écart dépenses",
           "Marge prévue", "Marge réelle", "Écart marge"]
for c, h in enumerate(headers, 1):
    ws.cell(row=r, column=c, value=h)
style_header_row(ws, r, NC)
r += 1

# Construire les formules à partir des totals
synth = [
    ("Crowdfunding",
     f"=Crowdfunding!{totals['crowd_rec'][1]}{totals['crowd_rec'][0]}",
     f"=Crowdfunding!{totals['crowd_rec'][2]}{totals['crowd_rec'][0]}",
     None, None),
    ("Billetterie + Artistes",
     f"=Billetterie!{totals['bill_rec'][1]}{totals['bill_rec'][0]}",
     f"=Billetterie!{totals['bill_rec'][2]}{totals['bill_rec'][0]}",
     f"=Artistes!F{totals['art_dep'][0]}+Artistes!H{totals['art_dep'][0]}",
     f"=Artistes!J{totals['art_dep'][0]}+Artistes!K{totals['art_dep'][0]}"),
    ("F&B (Buvettes + Restauration)",
     f"='F&B'!{totals['fb_rec'][1]}{totals['fb_rec'][0]}",
     f"='F&B'!{totals['fb_rec'][2]}{totals['fb_rec'][0]}",
     f"='F&B'!{totals['fb_dep'][1]}{totals['fb_dep'][0]}",
     f"='F&B'!{totals['fb_dep'][2]}{totals['fb_dep'][0]}"),
    ("Ateliers (recettes + coûts ateliers)",
     f"=Ateliers!{totals['at_rec'][1]}{totals['at_rec'][0]}",
     f"=Ateliers!{totals['at_rec'][2]}{totals['at_rec'][0]}",
     f"=Ateliers!{totals['at_dep'][1]}{totals['at_dep'][0]}",
     f"=Ateliers!{totals['at_dep'][2]}{totals['at_dep'][0]}"),
    ("Logistique",
     None, None,
     f"=Logistique!{totals['log_dep'][1]}{totals['log_dep'][0]}",
     f"=Logistique!{totals['log_dep'][2]}{totals['log_dep'][0]}"),
    ("Frais divers",
     f"='Frais divers'!{totals['div_rec'][1]}{totals['div_rec'][0]}",
     f"='Frais divers'!{totals['div_rec'][2]}{totals['div_rec'][0]}",
     f"='Frais divers'!{totals['div_dep'][1]}{totals['div_dep'][0]}",
     f"='Frais divers'!{totals['div_dep'][2]}{totals['div_dep'][0]}"),
]

s = r
for cat, rec_p, rec_r, dep_p, dep_r in synth:
    ws.cell(row=r, column=1, value=cat)
    ws.cell(row=r, column=2).value = rec_p if rec_p else 0
    ws.cell(row=r, column=3).value = rec_r if rec_r else 0
    ws.cell(row=r, column=4).value = f"=C{r}-B{r}"
    ws.cell(row=r, column=5).value = dep_p if dep_p else 0
    ws.cell(row=r, column=6).value = dep_r if dep_r else 0
    ws.cell(row=r, column=7).value = f"=F{r}-E{r}"
    ws.cell(row=r, column=8).value = f"=B{r}-E{r}"
    ws.cell(row=r, column=9).value = f"=C{r}-F{r}"
    ws.cell(row=r, column=10).value = f"=I{r}-H{r}"
    for c in range(1, NC + 1):
        ws.cell(row=r, column=c).border = THIN_BORDER
        if c >= 2:
            ws.cell(row=r, column=c).number_format = EUR_FORMAT
    for c in [2, 3, 4]:
        ws.cell(row=r, column=c).fill = RECETTE_FILL
    for c in [5, 6, 7]:
        ws.cell(row=r, column=c).fill = DEPENSE_FILL
    r += 1
e = r - 1

# Ajustement défraiement -25% (prise en charge réelle 75%)
r += 1
ws.cell(row=r, column=1, value="Ajustement défraiement artistes (-25%)")
ws.cell(row=r, column=5).value = -1650
ws.cell(row=r, column=5).number_format = EUR_FORMAT
for c in range(1, NC + 1):
    ws.cell(row=r, column=c).border = THIN_BORDER
    if c >= 2:
        ws.cell(row=r, column=c).number_format = EUR_FORMAT

e = r

r += 1
total_row(ws, r, "TOTAL", {c: (s, e) for c in range(2, NC + 1)}, NC)
for c in range(2, NC + 1):
    ws.cell(row=r, column=c).number_format = EUR_FORMAT
total_synth_row = r

# ============================================================
# SCÉNARIOS JAUGE -20% / +20%
# ============================================================
r += 3
r = section_title(ws, r, "SCÉNARIOS PAR JAUGE", 6)
r += 1
sc_headers = ["Scénario", "Recettes", "Coûts", "Marge", "Jauge", "Nb festivaliers"]
for c, h in enumerate(sc_headers, 1):
    ws.cell(row=r, column=c, value=h)
style_header_row(ws, r, 6)
r += 1

# Données scénarios (recettes variables, coûts artistes/logistique/frais divers fixes)
scenarios = [
    ("Jauge -20%", 46785, 49845, -3060, "960", 960),
    ("Jauge standard", 58482, 49845, 8637, "1 200", 1200),
    ("Jauge +20%", 66554, 49845, 16709, "1 440", 1440),
]

# Recalculer les coûts variables pour -20% et +20%
# Coûts fixes: artistes(21055-1650+555=19960), logistique(15000), frais div fixes(2000)
# Coûts variables: SACEM bill(8.8%), coût resto(25%), coût buv(21%), SACEM resto/buv(4.4%)
import math

def calc_scenario(factor):
    rec_bill = 22650 * factor
    rec_buv = 20720 * factor
    rec_resto = 11722 * factor
    rec_atel = 3390  # fixe
    rec_crowd = 8000  # fixe
    total_rec = rec_bill + rec_buv + rec_resto + rec_atel

    sacem_bill = rec_bill * 0.088
    sacem_fb = (rec_buv + rec_resto) * 0.044
    cout_resto = rec_resto * 0.25
    cout_buv = rec_buv * 0.21
    artistes = 15700 + 6600 - 1650 + 555  # cachets+rép + défr - ajust + resto
    logistique = 15000
    frais_div = 3000 - sacem_bill - sacem_fb  # SACEM est dans frais divers
    # En fait SACEM est séparé, frais div = 3000 fixe

    total_dep = sacem_bill + sacem_fb + cout_resto + cout_buv + artistes + logistique + 3000
    marge = total_rec - total_dep
    return round(total_rec), round(total_dep), round(marge)

for label, factor, jauge_label, nb in [
    ("Jauge -20%", 0.8, "960", 960),
    ("Jauge standard", 1.0, "1 200", 1200),
    ("Jauge +20%", 1.2, "1 440", 1440),
]:
    rec, dep, marge = calc_scenario(factor)
    ws.cell(row=r, column=1, value=label)
    style_data_cell(ws, r, 2, "eur").value = rec
    style_data_cell(ws, r, 3, "eur").value = dep
    style_data_cell(ws, r, 4, "eur").value = marge
    ws.cell(row=r, column=5, value=jauge_label)
    ws.cell(row=r, column=6, value=nb)

    # Coloration marge
    marge_cell = ws.cell(row=r, column=4)
    if marge < 0:
        marge_cell.fill = DEPENSE_FILL
    else:
        marge_cell.fill = RECETTE_FILL

    for c in range(1, 7):
        ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=c).font = Font(bold=(label == "Jauge standard"), size=11)
    r += 1

auto_width(ws)


# ============================================================
# SAUVEGARDE
# ============================================================
output = "c:/Users/charl/Monique-Festival/budget/Suivi_budgetaire_Monique_Festival.xlsx"
wb.save(output)
print(f"Fichier créé : {output}")
print(f"Onglets : {wb.sheetnames}")
print(f"Références totaux : {totals}")
