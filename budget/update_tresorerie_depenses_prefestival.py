"""
Met à jour l'onglet Trésorerie avec le plan détaillé des dépenses
PRÉ-FESTIVAL (mai → août 2026, avant le 28/08).

Source : analyse de tous les onglets dépenses (Artistes, Logistique,
Frais divers, Ateliers, F&B) + Calibrage_buvette_2026.md.

Hypothèses d'échéance :
- Logistique fournisseurs : acompte 30 % à la commande, solde au montage
- Artistes (cachets) : 30 % acompte à la signature contrat, 70 % au festival
- Artistes (répétitions) : payées le mois où elles ont lieu
- Communication : payée à la commande
- Assurance : prime annuelle à la souscription
- F&B équipement loué : acompte 30 % réservation + solde livraison
- F&B consommables/écocups/jetons : paiement à la commande
- F&B matières fraîches : paiement à la livraison J-2

Usage : python budget/update_tresorerie_depenses_prefestival.py
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

REPO = Path(__file__).resolve().parent.parent
FILE = REPO / "budget" / "Suivi_budgetaire_Monique_Festival.xlsx"

# Styles
THIN = Side(border_style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H1 = Font(bold=True, size=13, color="FFFFFF")
H2 = Font(bold=True, size=11, color="FFFFFF")
HEADER = Font(bold=True, size=10)
TOTAL = Font(bold=True, size=11)
FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_HISTO = PatternFill("solid", fgColor="2E7D32")
FILL_MAI = PatternFill("solid", fgColor="6A1B9A")
FILL_JUIN = PatternFill("solid", fgColor="EF6C00")
FILL_JUIL = PatternFill("solid", fgColor="C62828")
FILL_AOUT = PatternFill("solid", fgColor="0277BD")
FILL_HEADER = PatternFill("solid", fgColor="EEEEEE")
FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")

# === DONNÉES ===

# Flux historiques (déjà passés)
HISTORIQUE = [
    ("23/04/2026", "Crowdfunding", "Cumul 14 dons HelloAsso au 23/04", 845, 0),
    ("20/04/2026", "Don Charles", "Don manuel par virement Qonto", 1000, 0),
    ("Avr-Mai 2026", "Frais divers", "Petites dépenses opérationnelles cumul (cartes de visite, frais admin)", 0, 255),
    ("01/05/2026", "Crowdfunding", "Cumul dons supplémentaires (23/04 → 01/05)", 360, 0),
]

# Dépenses prévisionnelles pré-festival (par mois)
# Format : (Pôle, Description, Montant_prévu, Note)

DEPENSES_MAI = [
    ("Admin", "Souscription assurance RC pro festival (prime annuelle estimative)", 800, "Cible AIAC ou MAIF — 600-1000 €"),
    ("Admin", "Cartes de visite Pixartprinting (commande passée 30/04)", 50, "Pour collecte mécénat"),
    ("Admin", "Frais postaux LRAR rescrit DDFIP + courriers partenaires", 30, "Recommandé +AR"),
]

DEPENSES_JUIN = [
    # Communication
    ("Communication", "1ère vague affiches + flyers (impression Pixartprinting)", 500, "Lancement com programmation"),
    # Logistique - acomptes 30%
    ("Logistique", "Acompte 30% Location salle La Grange Huguenet", 900, "Sur 3000 € total"),
    ("Logistique", "Acompte 30% Sonorisation", 900, "Sur 3000 €"),
    ("Logistique", "Acompte 30% Éclairage / lumière", 600, "Sur 2000 €"),
    ("Logistique", "Acompte 30% Scène / podium", 600, "Sur 2000 €"),
    # F&B
    ("F&B", "Commande écocups Atomic 1000 unités", 600, "Délai livraison ~3 semaines"),
    ("F&B", "Commande jetons flousy ~5000 unités", 400, "Stock {1×2000, 2×1000, 5×1500, 10×500}"),
    ("F&B", "Acompte 30% Machine pop-corn (location)", 80, "Sur 252 € total"),
    # Note : tous les cachets artistes payés POST-festival (décision Charles 03/05/2026)
]

DEPENSES_JUILLET = [
    # Communication
    ("Communication", "2ème vague communication + vidéo promo", 500, "Final push avant festival"),
    # Logistique - acomptes 30%
    ("Logistique", "Acompte 30% Toilettes / sanitaires", 450, "Sur 1500 €"),
    ("Logistique", "Acompte 30% Électricité / groupe électrogène", 450, "Sur 1500 €"),
    ("Logistique", "Acompte 30% Barrières / sécurité", 300, "Sur 1000 €"),
    ("Logistique", "Acompte 30% Mobilier (tables, chaises, tentes)", 300, "Sur 1000 €"),
    # Artistes - répétitions
    ("Artistes", "Répétitions Fabulo (4 art × 2 jours × 100 €)", 800, "Préparation collectif Fabulo"),
    ("Artistes", "Répétitions Atelier écriture — Sylvain (2j × 100 €)", 200, ""),
    ("Artistes", "Répétitions Atelier DJ — Jeff (2j × 100 €)", 200, ""),
    # Artistes - défraiement étrangers (cachets payés post-festival)
    ("Artistes", "Avance défraiement artistes étrangers (50% billets early)", 600, "Pour booking train/avion avant 15/06 deadline"),
    # F&B - matériel non périssable
    ("F&B", "Vaisselle bio jetable (gobelets, assiettes, serviettes)", 319, "Stock pour J-2"),
    ("F&B", "HACCP (tablier, gants, sondes T°) + caisse + ustensiles préparation", 540, "Préparation stand"),
]

DEPENSES_AOUT_AVANT_FESTIVAL = [
    # Logistique - soldes 70%
    ("Logistique", "Solde 70% Location salle La Grange Huguenet", 2100, "Au montage"),
    ("Logistique", "Solde 70% Sonorisation", 2100, "Au montage J-2"),
    ("Logistique", "Solde 70% Éclairage / lumière", 1400, "Au montage J-2"),
    ("Logistique", "Solde 70% Scène / podium", 1400, "Au montage J-2"),
    ("Logistique", "Solde 70% Toilettes / sanitaires", 1050, "Au montage J-1"),
    ("Logistique", "Solde 70% Électricité / groupe électrogène", 1050, "Au montage J-2"),
    ("Logistique", "Solde 70% Barrières / sécurité", 700, "Au montage J-2"),
    ("Logistique", "Solde 70% Mobilier (tables, chaises, tentes)", 700, "Au montage J-1"),
    # F&B équipement loué
    ("F&B", "Solde Machine pop-corn (location)", 172, "Au montage"),
    ("F&B", "Percolateurs 4×10L ABC LOCATION", 345, "Réservation + livraison J-1"),
    ("F&B", "Friteuses ×2 ABC LOCATION", 172, "Réservation + livraison J-1"),
    ("F&B", "Eau gratuite — fontaines + carafes", 400, "Location J-1"),
    ("F&B", "Trousse premiers secours buvette + affichage", 60, "Préparation stand"),
    ("F&B", "Sacs poubelles tri + lavage écocups + récup huile", 120, "Stand + préparation"),
    # F&B boissons (commandes août)
    ("F&B", "Bière Le Pintadier — 10 fûts initial", 800, "Commande J-3, livraison J-1"),
    ("F&B", "Vin Cubi Jura — 29 cubis (Hyper Boisson)", 1450, "Commande J-7"),
    ("F&B", "Soft (sirop + jus + Mortuacienne + eaux 1L verre)", 775, "Commande J-7"),
    ("F&B", "Café moulu équitable BFC + sucre + lait + mélangeurs + gobelets", 110, "Commande J-7"),
    ("F&B", "Sucre bar additionnel + glace pilée", 60, "Achat J-1"),
    # F&B matières fraîches restauration (J-2)
    ("F&B", "Assiettes : fromage Hôpitaux Vieux + Comté/Morbier (~70 kg)", 1100, "Commande J-3, livraison J-2"),
    ("F&B", "Assiettes : charcuterie Ferme Ligny (~77 kg)", 1200, "Commande J-3, livraison J-2"),
    ("F&B", "Œufs Saveur de la Ferme (3 plateaux 360 + extras)", 200, "Commande J-3, livraison J-2"),
    ("F&B", "Pain La Ronde des Pains (~110 pains 350g)", 245, "Commande J-3, livraison J-1"),
    ("F&B", "Tomates cerise BRUNO BFC (~75 kg, restauration + snack)", 90, "Commande J-3, livraison J-2"),
    ("F&B", "Pommes de terre BFC (~50 kg pour frites snack + grandes assiettes)", 60, "Commande J-3, livraison J-2"),
    ("F&B", "Sachets sous vide Ferme Ligny + autres consommables resto", 50, "Avec commande charcuterie"),
    # F&B snacks matières
    ("F&B", "Pop-corn (maïs 10kg + huile + sel + gobelets bio 72cL)", 63, "Commande J-7"),
    ("F&B", "Frites snack (huile friture + sel + barquettes kraft + sticks sauces)", 88, "Commande J-7 (pdt comptées séparément)"),
    ("F&B", "Glaces ERHARD (43 bacs 2,5L)", 101, "Commande J-3, livraison J-1"),
    ("F&B", "Cornets gaufrette comestibles (560 unités) + back-up", 70, "Commande J-7"),
    # Artistes - défraiement étrangers solde uniquement (cachets payés post-festival)
    ("Artistes", "Défraiement étrangers solde (sur réception justificatifs)", 1200, "Solde après réception tickets, payable J ou J+7"),
    # Frais divers
    ("Admin", "Frais administratifs divers (encres, fournitures bureau)", 100, "Préparation festival"),
    ("Admin", "Imprévus de dernière minute (provision)", 500, "Aléas"),
]

# Dépenses post-festival (juste pour référence — non comptabilisées dans pré-festival)
DEPENSES_POST_FESTIVAL = [
    ("Artistes", "Cachets artistes (TOUS payés post-festival, J+7)", 15700, "Septembre 2026 — décision Charles 03/05/2026"),
    ("Artistes", "Défraiement artistes (étrangers - solde, locaux/régionaux)", 4800, "Septembre 2026 — sur justificatifs"),
    ("Admin", "SACEM Billetterie 8.8% (déclaration post-event)", 1993, "Septembre 2026 — sur recettes billetterie réelles"),
]

def style_row(ws, row, fill=None, font=None, border=True, align=None, ncols=6):
    for c in range(1, ncols+1):
        cell = ws.cell(row, c)
        if fill: cell.fill = fill
        if font: cell.font = font
        if border: cell.border = BORDER
        if align: cell.alignment = align


def main():
    wb = load_workbook(FILE)
    if "Trésorerie" not in wb.sheetnames:
        sys.exit("ERREUR : onglet Trésorerie absent")

    # Supprimer et recréer
    idx = wb.sheetnames.index("Trésorerie")
    del wb["Trésorerie"]
    ws = wb.create_sheet("Trésorerie", idx)

    # Largeurs
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 14

    # Titre
    ws['A1'] = "TRÉSORERIE — Plan détaillé des dépenses pré-festival (mai → août 2026)"
    ws.merge_cells('A1:F1')
    ws['A1'].font = H1
    ws['A1'].fill = FILL_TITLE
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    r = 3
    # === Section HISTORIQUE ===
    ws.cell(r, 1, "FLUX HISTORIQUES (déjà passés au 03/05/2026)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1).font = H2
    ws.cell(r, 1).fill = FILL_HISTO
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 2

    headers_histo = ["Date", "Catégorie", "Description", "Entrée", "Sortie", "Solde"]
    for c, h in enumerate(headers_histo, 1):
        ws.cell(r, c, h).font = HEADER
        ws.cell(r, c).fill = FILL_HEADER
        ws.cell(r, c).border = BORDER
        ws.cell(r, c).alignment = Alignment(horizontal="center")
    r += 1

    histo_start = r
    solde = 0
    for date, cat, desc, ent, sor in HISTORIQUE:
        ws.cell(r, 1, date)
        ws.cell(r, 2, cat)
        ws.cell(r, 3, desc)
        if ent: ws.cell(r, 4, ent)
        if sor: ws.cell(r, 5, sor)
        solde += (ent or 0) - (sor or 0)
        ws.cell(r, 6, solde)
        style_row(ws, r)
        r += 1

    # Solde Qonto au 03/05
    ws.cell(r, 2, "SOLDE Qonto estimé au 03/05/2026")
    ws.cell(r, 6, solde)
    style_row(ws, r, fill=FILL_TOTAL, font=TOTAL)
    solde_initial = solde
    r += 2

    # === Sections DÉPENSES par mois ===
    months = [
        ("MAI 2026 — Dépenses prévues", DEPENSES_MAI, FILL_MAI),
        ("JUIN 2026 — Dépenses prévues", DEPENSES_JUIN, FILL_JUIN),
        ("JUILLET 2026 — Dépenses prévues", DEPENSES_JUILLET, FILL_JUIL),
        ("AOÛT 2026 (avant festival 28/08) — Dépenses prévues", DEPENSES_AOUT_AVANT_FESTIVAL, FILL_AOUT),
    ]

    grand_total = 0
    headers_dep = ["Échéance", "Pôle", "Description", "Montant prévu", "Statut", "Note"]

    for title, items, fill in months:
        ws.cell(r, 1, title)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1).font = H2
        ws.cell(r, 1).fill = fill
        ws.cell(r, 1).alignment = Alignment(horizontal="center")
        r += 2

        for c, h in enumerate(headers_dep, 1):
            ws.cell(r, c, h).font = HEADER
            ws.cell(r, c).fill = FILL_HEADER
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = Alignment(horizontal="center")
        r += 1

        section_start = r
        section_total = 0
        for pole, desc, montant, note in items:
            ws.cell(r, 2, pole)
            ws.cell(r, 3, desc)
            ws.cell(r, 4, montant)
            ws.cell(r, 5, "Prévu")
            ws.cell(r, 6, note)
            section_total += montant
            style_row(ws, r)
            r += 1

        # Sous-total mois
        ws.cell(r, 3, f"Sous-total {title.split('—')[0].strip()}")
        ws.cell(r, 4, f"=SUM(D{section_start}:D{r-1})")
        style_row(ws, r, fill=FILL_TOTAL, font=TOTAL)
        grand_total += section_total
        r += 2

    # === Total cumulé pré-festival ===
    ws.cell(r, 1, "TOTAL DÉPENSES PRÉ-FESTIVAL (mai-août avant 28/08)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(r, 4, grand_total)
    ws.cell(r, 1).font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(r, 4).font = Font(bold=True, color="FFFFFF", size=12)
    for c in range(1, 7):
        ws.cell(r, c).fill = PatternFill("solid", fgColor="C62828")
        ws.cell(r, c).border = BORDER
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 2

    # === Notes ===
    ws.cell(r, 1, "NOTES & HYPOTHÈSES")
    ws.cell(r, 1).font = HEADER
    r += 1
    notes = [
        f"• Solde Qonto au 03/05/2026 : {solde_initial} € — Besoin de trésorerie cumulé jusqu'au 27/08 : {grand_total} €",
        f"• Gap à financer : {grand_total - solde_initial} € via crowdfunding + mécénat + subventions + early bird billetterie",
        "• Hypothèses fournisseurs : acompte 30% à la commande / solde 70% au montage (à confirmer avec chaque devis)",
        "• Cachets artistes : 30% acompte signature contrat, 50% post-prestation J+7 (à confirmer convention)",
        "• Défraiement étrangers : avance possible 50% pour booking précoce, solde sur justificatifs",
        "• Matières fraîches F&B : commandées J-3 (hors période 'pré-festival' stricte mais cash sortant avant recettes billetterie)",
        "• SACEM Billetterie 1 993 € : payée septembre 2026 (post-festival, sur recettes réelles)",
        "• SACEM Buvette 1 427 € : exclue (à confirmer non due)",
        "• Repas artistes 555 € : à imputer budget Artistes (non doublonné ici)",
    ]
    for n in notes:
        ws.cell(r, 1, n)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1).font = Font(size=9, italic=True)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 18
        r += 1

    # === Référence post-festival (info) ===
    r += 1
    ws.cell(r, 1, "POUR INFO — Dépenses POST-festival (non comptées ci-dessus)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1).font = Font(size=10, italic=True, color="666666")
    ws.cell(r, 1).fill = PatternFill("solid", fgColor="EEEEEE")
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 1

    for pole, desc, montant, note in DEPENSES_POST_FESTIVAL:
        ws.cell(r, 2, pole)
        ws.cell(r, 3, desc)
        ws.cell(r, 4, montant)
        ws.cell(r, 6, note)
        for c in range(1, 7):
            ws.cell(r, c).font = Font(size=9, italic=True, color="666666")
            ws.cell(r, c).border = BORDER
        r += 1

    wb.save(FILE)

    # Récap console
    total_mai = sum(m for _, _, m, _ in DEPENSES_MAI)
    total_juin = sum(m for _, _, m, _ in DEPENSES_JUIN)
    total_juil = sum(m for _, _, m, _ in DEPENSES_JUILLET)
    total_aout = sum(m for _, _, m, _ in DEPENSES_AOUT_AVANT_FESTIVAL)

    print("OK Trésorerie mise à jour avec plan dépenses pré-festival")
    print(f"  Solde Qonto au 03/05/2026 : {solde_initial:>7,} €".replace(",", " "))
    print(f"  Dépenses Mai 2026         : {total_mai:>7,} €".replace(",", " "))
    print(f"  Dépenses Juin 2026        : {total_juin:>7,} €".replace(",", " "))
    print(f"  Dépenses Juillet 2026     : {total_juil:>7,} €".replace(",", " "))
    print(f"  Dépenses Août pré-fest    : {total_aout:>7,} €".replace(",", " "))
    print(f"  TOTAL pré-festival        : {grand_total:>7,} €".replace(",", " "))
    print(f"  Gap à financer            : {grand_total - solde_initial:>7,} €".replace(",", " "))


if __name__ == "__main__":
    main()
