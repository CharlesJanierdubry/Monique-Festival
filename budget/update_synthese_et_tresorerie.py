"""
Met à jour 2 onglets du Suivi_budgetaire :
1. Synthèse — corrige les formules Artistes (ligne TOTAL passée de 39 à 36,
   inclut maintenant cachets F + répétitions G + défraiement H)
2. Trésorerie — inscrit les flux réels au 1er mai 2026 (état Qonto)

Usage : python budget/update_synthese_et_tresorerie.py
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

REPO = Path(__file__).resolve().parent.parent
FILE = REPO / "budget" / "Suivi_budgetaire_Monique_Festival.xlsx"

# Trésorerie Qonto réelle au 1er mai 2026 (uniquement flux cash sur Qonto)
# NB : l'acompte 500 € du 10/01 et son abandon de créance du 25/04 sont
# des opérations comptables (registre des dons) sans flux Qonto associé.
TRESORERIE = [
    # (Date, Catégorie, Description, Entrée, Sortie)
    ("23/04/2026", "Crowdfunding HelloAsso", "Cumul dons collectés au 23/04 (14 dons)", 845, 0),
    ("20/04/2026", "Don manuel Charles", "Don manuel par virement Qonto", 1000, 0),
    ("Avr-Mai 2026", "Frais divers", "Petites dépenses opérationnelles (cumul estimation)", 0, 255),
    ("01/05/2026", "Crowdfunding HelloAsso", "Cumul dons supplémentaires (23/04 → 01/05)", 360, 0),
]
# Note hors trésorerie : l'abandon de créance Charles 500 € (25/04) est inscrit
# au registre des dons mais ne génère pas de flux Qonto. Total don Charles
# comptablement = 1500 € (1000 € cash + 500 € abandon créance).

# Solde initial (ligne 2) : 0 €

def main():
    wb = load_workbook(FILE)

    # ========================================
    # 1. CORRECTION FORMULES SYNTHÈSE
    # ========================================
    if "Synthèse" not in wb.sheetnames:
        sys.exit("ERREUR : onglet 'Synthèse' absent")
    ws = wb["Synthèse"]

    # E5 = Dépenses prévues Billetterie+Artistes — ancienne formule pointait vers F39+H39
    # Maintenant : ligne TOTAL = 36, et on inclut F (cachets) + G (répétitions) + H (défraiement)
    ws['E5'] = "=Artistes!F36+Artistes!G36+Artistes!H36"
    # F5 = Dépenses réelles — colonnes J (paiement réel) + K (défraiement réel)
    ws['F5'] = "=Artistes!J36+Artistes!K36"
    print("OK Synthèse : formules E5 + F5 mises à jour (TOTAL Artistes ligne 36, inclut répétitions)")

    # Supprimer la ligne d'ajustement défraiement -1650 (plus pertinente, le coefficient
    # 75% est désormais intégré dans les estimations 150€ par artiste)
    ws['A11'] = "(ajustement défraiement supprimé : 150€ moyen intégré directement)"
    ws['B11'] = None
    ws['E11'] = 0
    print("OK Synthèse : ligne 11 d'ajustement neutralisée (E11=0)")

    # ========================================
    # 2. ONGLET TRÉSORERIE
    # ========================================
    if "Trésorerie" not in wb.sheetnames:
        sys.exit("ERREUR : onglet 'Trésorerie' absent")
    ws_t = wb["Trésorerie"]

    # Effacer toutes les lignes de données (on garde le titre L1 + Solde initial L2 + en-têtes L3)
    for row in range(ws_t.max_row, 3, -1):
        ws_t.delete_rows(row)

    # Solde initial à 0
    # Ligne 2 = "Solde initial | | | 0 |"
    # Ligne 3 = en-têtes
    # Lignes 4+ = mouvements

    THIN = Side(border_style="thin", color="999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    row = 4
    solde = 0
    for date, categorie, description, entree, sortie in TRESORERIE:
        ws_t.cell(row, 1, date)
        ws_t.cell(row, 2, categorie)
        ws_t.cell(row, 3, description)
        ws_t.cell(row, 4, entree if entree else "")
        ws_t.cell(row, 5, sortie if sortie else "")
        # Solde cumulé en formule
        if row == 4:
            ws_t.cell(row, 6, f"=D2+D{row}-E{row}")  # D2 = solde initial
        else:
            ws_t.cell(row, 6, f"=F{row-1}+D{row}-E{row}")
        for c in range(1, 7):
            ws_t.cell(row, c).border = BORDER
        solde += (entree or 0) - (sortie or 0)
        row += 1

    # Ligne TOTAUX
    ws_t.cell(row, 1, "TOTAUX")
    ws_t.cell(row, 4, f"=SUM(D4:D{row-1})")
    ws_t.cell(row, 5, f"=SUM(E4:E{row-1})")
    ws_t.cell(row, 6, f"=F{row-1}")
    for c in range(1, 7):
        ws_t.cell(row, c).font = Font(bold=True)
        ws_t.cell(row, c).fill = PatternFill("solid", fgColor="FCE4D6")
        ws_t.cell(row, c).border = BORDER

    # Ligne SOLDE ACTUEL pour info
    row_solde = row + 2
    ws_t.cell(row_solde, 1, "SOLDE Qonto au 01/05/2026 (estimation)")
    ws_t.cell(row_solde, 1).font = Font(bold=True, size=12, color="1F4E78")
    ws_t.cell(row_solde, 6, f"=F{row-1}")
    ws_t.cell(row_solde, 6).font = Font(bold=True, size=14)
    ws_t.cell(row_solde, 6).fill = PatternFill("solid", fgColor="D9E1F2")

    print(f"OK Trésorerie : {len(TRESORERIE)} mouvements inscrits, solde calculé = {solde} €")

    # Sauvegarde
    wb.save(FILE)
    print(f"\nSauvegarde : {FILE.relative_to(REPO)}")


if __name__ == "__main__":
    main()
