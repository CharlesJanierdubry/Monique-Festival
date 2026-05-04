"""
Réécrit l'onglet F&B avec maximum de détail :
- RECETTES : 1 ligne par (jour × produit) — colonne Jour explicite
- DÉPENSES : 1 ligne par achat (achat unique, pas par jour)
  avec colonnes Mois / Type (Fixe/Invest/Variable) / Catégorie

Source : Calibrage_buvette_2026.md, Tarifs_buvette_restauration.md, Scénario B validé.

Usage : python budget/update_fb_detaille.py
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

REPO = Path(__file__).resolve().parent.parent
FILE = REPO / "budget" / "Suivi_budgetaire_Monique_Festival.xlsx"

# ============= STYLES =============
THIN = Side(border_style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H1 = Font(bold=True, size=13, color="FFFFFF")
H2 = Font(bold=True, size=11, color="FFFFFF")
HEADER = Font(bold=True, size=10)
TOTAL = Font(bold=True, size=11)
FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_REC = PatternFill("solid", fgColor="2E7D32")
FILL_DEP = PatternFill("solid", fgColor="C62828")
FILL_SYNTH = PatternFill("solid", fgColor="0277BD")
FILL_HEADER = PatternFill("solid", fgColor="EEEEEE")
FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")
FILL_VEN = PatternFill("solid", fgColor="E8F5E9")
FILL_SAM = PatternFill("solid", fgColor="FFF8E1")
FILL_DIM = PatternFill("solid", fgColor="E3F2FD")

# ============= DONNÉES RECETTES =============
# Format : (Jour, Catégorie, Produit, Format, Quantité, Prix flousy, Prix €, Recette €)
# Volumes calculés depuis Calibrage_buvette_2026.md (modulation par jour)

RECETTES = [
    # ===== VENDREDI (270 fest) =====
    ("Vendredi", "Buvette", "Bière blonde/blanche Le Pintadier", "Verre 25 cL", 300, 3, 4.50, 1350),
    ("Vendredi", "Buvette", "Vin Cubi Jura tradition", "Verre 25 cL", 108, 3, 4.50, 486),
    ("Vendredi", "Buvette", "Sirop dilué", "Verre 25 cL", 146, 2, 3.00, 438),
    ("Vendredi", "Buvette", "Jus de pomme Saveur de la Ferme", "Verre 25 cL", 73, 2, 3.00, 219),
    ("Vendredi", "Buvette", "Mortuacienne (limonade)", "Verre 25 cL", 29, 2, 3.00, 87),
    ("Vendredi", "Buvette", "Eau pétillante Bisontine", "Btl verre 1 L", 7, 2, 3.00, 21),
    ("Vendredi", "Buvette", "Eau plate (BFC)", "Btl verre 1 L", 7, 2, 3.00, 21),
    ("Vendredi", "Buvette", "Café moulu équitable BFC", "Verre 25 cL", 40, 1, 1.50, 60),
    ("Vendredi", "Snacks", "Pop-corn", "Gobelet 72 cL", 40, 1, 1.50, 60),
    ("Vendredi", "Snacks", "Frites (snack à part)", "Barquette kraft 150 g", 40, 2, 3.00, 120),
    ("Vendredi", "Snacks", "Glace 1 boule", "Cornet gaufrette", 31, 1, 1.50, 47),
    ("Vendredi", "Snacks", "Glace 2 boules", "Cornet gaufrette", 21, 2, 3.00, 63),
    ("Vendredi", "Snacks", "Tomates cerise BRUNO", "Gobelet 25 cL (150 g)", 100, 1, 1.50, 150),
    ("Vendredi", "Restauration", "Petite assiette franc-comtoise", "Plateau", 47, 6, 9.00, 423),
    ("Vendredi", "Restauration", "Petite assiette végétarienne", "Plateau", 21, 6, 9.00, 189),
    ("Vendredi", "Restauration", "Grande assiette franc-comtoise", "Plateau + frites", 110, 8, 12.00, 1320),
    ("Vendredi", "Restauration", "Grande assiette végétarienne", "Plateau + frites", 47, 8, 12.00, 564),
    ("Vendredi", "Consommables", "Consigne écocup non remboursée", "Jeton", 41, 1, 1.50, 62),
    # ===== SAMEDI (780 fest) =====
    ("Samedi", "Buvette", "Bière blonde/blanche Le Pintadier", "Verre 25 cL", 1150, 3, 4.50, 5175),
    ("Samedi", "Buvette", "Vin Cubi Jura tradition", "Verre 25 cL", 390, 3, 4.50, 1755),
    ("Samedi", "Buvette", "Sirop dilué", "Verre 25 cL", 634, 2, 3.00, 1902),
    ("Samedi", "Buvette", "Jus de pomme Saveur de la Ferme", "Verre 25 cL", 317, 2, 3.00, 951),
    ("Samedi", "Buvette", "Mortuacienne (limonade)", "Verre 25 cL", 127, 2, 3.00, 381),
    ("Samedi", "Buvette", "Eau pétillante Bisontine", "Btl verre 1 L", 32, 2, 3.00, 96),
    ("Samedi", "Buvette", "Eau plate (BFC)", "Btl verre 1 L", 32, 2, 3.00, 96),
    ("Samedi", "Buvette", "Café moulu équitable BFC", "Verre 25 cL", 240, 1, 1.50, 360),
    ("Samedi", "Snacks", "Pop-corn", "Gobelet 72 cL", 285, 1, 1.50, 428),
    ("Samedi", "Snacks", "Frites (snack à part)", "Barquette kraft 150 g", 240, 2, 3.00, 720),
    ("Samedi", "Snacks", "Glace 1 boule", "Cornet gaufrette", 243, 1, 1.50, 365),
    ("Samedi", "Snacks", "Glace 2 boules", "Cornet gaufrette", 162, 2, 3.00, 486),
    ("Samedi", "Snacks", "Tomates cerise BRUNO", "Gobelet 25 cL (150 g)", 290, 1, 1.50, 435),
    ("Samedi", "Restauration", "Petite assiette franc-comtoise", "Plateau", 137, 6, 9.00, 1233),
    ("Samedi", "Restauration", "Petite assiette végétarienne", "Plateau", 57, 6, 9.00, 513),
    ("Samedi", "Restauration", "Grande assiette franc-comtoise", "Plateau + frites", 319, 8, 12.00, 3828),
    ("Samedi", "Restauration", "Grande assiette végétarienne", "Plateau + frites", 137, 8, 12.00, 1644),
    ("Samedi", "Consommables", "Consigne écocup non remboursée", "Jeton", 118, 1, 1.50, 177),
    # ===== DIMANCHE (270 fest) =====
    ("Dimanche", "Buvette", "Bière blonde/blanche Le Pintadier", "Verre 25 cL", 150, 3, 4.50, 675),
    ("Dimanche", "Buvette", "Vin Cubi Jura tradition", "Verre 25 cL", 68, 3, 4.50, 306),
    ("Dimanche", "Buvette", "Sirop dilué", "Verre 25 cL", 219, 2, 3.00, 657),
    ("Dimanche", "Buvette", "Jus de pomme Saveur de la Ferme", "Verre 25 cL", 110, 2, 3.00, 330),
    ("Dimanche", "Buvette", "Mortuacienne (limonade)", "Verre 25 cL", 44, 2, 3.00, 132),
    ("Dimanche", "Buvette", "Eau pétillante Bisontine", "Btl verre 1 L", 11, 2, 3.00, 33),
    ("Dimanche", "Buvette", "Eau plate (BFC)", "Btl verre 1 L", 11, 2, 3.00, 33),
    ("Dimanche", "Buvette", "Café moulu équitable BFC", "Verre 25 cL", 80, 1, 1.50, 120),
    ("Dimanche", "Snacks", "Pop-corn", "Gobelet 72 cL", 80, 1, 1.50, 120),
    ("Dimanche", "Snacks", "Frites (snack à part)", "Barquette kraft 150 g", 80, 2, 3.00, 240),
    ("Dimanche", "Snacks", "Glace 1 boule", "Cornet gaufrette", 62, 1, 1.50, 93),
    ("Dimanche", "Snacks", "Glace 2 boules", "Cornet gaufrette", 41, 2, 3.00, 123),
    ("Dimanche", "Snacks", "Tomates cerise BRUNO", "Gobelet 25 cL (150 g)", 110, 1, 1.50, 165),
    ("Dimanche", "Restauration", "Petite assiette franc-comtoise", "Plateau", 47, 6, 9.00, 423),
    ("Dimanche", "Restauration", "Petite assiette végétarienne", "Plateau", 21, 6, 9.00, 189),
    ("Dimanche", "Restauration", "Grande assiette franc-comtoise", "Plateau + frites", 110, 8, 12.00, 1320),
    ("Dimanche", "Restauration", "Grande assiette végétarienne", "Plateau + frites", 47, 8, 12.00, 564),
    ("Dimanche", "Consommables", "Consigne écocup non remboursée", "Jeton", 41, 1, 1.50, 62),
]

# ============= DONNÉES DÉPENSES =============
# Format : (Mois, Type, Catégorie, Produit/Achat, Quantité, Unité, Prix unit €, Coût total €, Fournisseur, Note)

DEPENSES = [
    # ===== JUIN 2026 =====
    ("Juin", "Invest", "Vaisselle réutilisable", "Écocups réutilisables PP 25 cL", 1000, "unités", 0.60, 600, "Atomic", "Délai livraison ~3 sem · réutilisable édition 2027+"),
    ("Juin", "Invest", "Caisse / paiement", "Jetons flousy ~5000 unités (mix 1/2/5/10)", 5000, "unités", 0.08, 400, "jeton-monnaie.fr ou gravoplaque.com", "{1×2000, 2×1000, 5×1500, 10×500}"),
    ("Juin", "Fixe", "Équipement loué", "Acompte 30% machine pop-corn (location)", 1, "forfait", 80, 80, "À finaliser", "Solde 172 € en août"),

    # ===== JUILLET 2026 =====
    ("Juillet", "Fixe", "HACCP", "Tablier + charlotte + gants jetables + sonde T° + désinfectant", 1, "kit", 80, 80, "Metro / fournisseur HoReCa", "Obligation cuisson frites"),
    ("Juillet", "Fixe", "Caisse / Admin", "Caisse + fond de caisse 300 € + ustensiles préparation", 1, "kit", 280, 280, "Divers", "Espèces + planche + couteaux + plateaux"),
    ("Juillet", "Fixe", "HACCP", "Trousse premiers secours buvette", 1, "kit", 30, 30, "Pharmacie", "Obligation"),
    ("Juillet", "Fixe", "Caisse / Admin", "Affichage prix (ardoises grand format + marqueurs)", 1, "lot", 30, 30, "Cultura / Bureau Vallée", "Bar + Restauration + allergènes"),
    ("Juillet", "Fixe", "HACCP", "Sacs poubelles tri (verre/biodéchets/OMR) + bidon huile usagée", 1, "lot", 70, 70, "Metro", "Tri sélectif charte éthique"),
    ("Juillet", "Fixe", "Vaisselle bio jetable", "Vaisselle bio jetable globale (gobelets + assiettes + cuillères + serviettes + barquettes)", 1, "lot", 319, 319, "Cdiscount Pro / Vegware", "Détail composition voir Calibrage_buvette_2026.md"),

    # ===== AOÛT 2026 =====
    # --- Équipement loué ---
    ("Août", "Fixe", "Équipement loué", "Solde 70% machine pop-corn", 1, "forfait", 172, 172, "À finaliser", "Au montage J-1"),
    ("Août", "Fixe", "Équipement loué", "Percolateurs 4 × 10 L", 4, "unités", 86.20, 345, "ABC LOCATION", "Capacité 320 cafés/tournée"),
    ("Août", "Fixe", "Équipement loué", "Friteuses pro × 2", 2, "unités", 86.20, 172, "ABC LOCATION", "~2-3 tournées/heure samedi"),
    ("Août", "Fixe", "Équipement loué", "Eau gratuite : 2-3 fontaines + carafes/pichets", 1, "lot", 400, 400, "Loca-Buffet / Loxam / Kiloutou", "Charte éthique + obligation bénévoles"),
    ("Août", "Fixe", "HACCP", "Lavage écocups (produit vaisselle pro + brosses + éponges)", 1, "lot", 50, 50, "Metro / fournisseur HoReCa", "Stand rinçage rapide près du bar"),

    # --- Boissons (variables) ---
    ("Août", "Variable", "Boisson - Bière", "Bière blonde+blanche fûts 40 L (commande initiale ferme)", 10, "fûts", 80, 800, "Le Pintadier", "1600 verres · réassort 5 fûts décision Ven soir"),
    ("Août", "Variable", "Boisson - Vin", "Vin Cubi Jura tradition 5 L (prix négocié)", 29, "cubis", 50, 1450, "Hyper Boisson", "580 verres · prix à confirmer"),
    ("Août", "Variable", "Boisson - Soft", "Sirop dilué (parfums grenadine/menthe/citron à préciser)", 50, "L (estim.)", 6.50, 300, "Hyper Boisson", "1000 verres dilués"),
    ("Août", "Variable", "Boisson - Soft", "Jus de pomme local poche 3 L", 21, "poches", 8.40, 176, "Saveur de la Ferme", "Pour 500 verres (à recouper)"),
    ("Août", "Variable", "Boisson - Soft", "Mortuacienne btl verre 1 L (vendue au verre)", 50, "btl", 1.16, 58, "Hyper Boisson", "200 verres"),
    ("Août", "Variable", "Boisson - Soft", "Eau pétillante Bisontine btl verre 1 L", 50, "btl", 0.70, 35, "Hyper Boisson", "Vendue entière + consigne 1 fl"),
    ("Août", "Variable", "Boisson - Soft", "Eau plate btl verre 1 L (à sourcer)", 50, "btl", 0.65, 32, "Velleminfroy / Saint-Antonin (à confirmer)", "Vendue entière + consigne 1 fl"),
    ("Août", "Variable", "Boisson - Café", "Café moulu équitable BFC", 3, "kg", 15, 45, "Torréfacteur BFC à sourcer", "7 g/tasse, 360 cafés"),
    ("Août", "Variable", "Boisson - Café", "Sucre sticks", 360, "sticks", 0.03, 11, "Metro", ""),
    ("Août", "Variable", "Boisson - Café", "Lait UHT mini-portions", 250, "portions", 0.10, 25, "Metro", ""),
    ("Août", "Variable", "Boisson - Café", "Mélangeurs bois", 360, "unités", 0.02, 7, "Metro", ""),
    ("Août", "Variable", "Boisson - Bar", "Sucre bar additionnel + glace pilée bar", 1, "lot", 60, 60, "Metro", ""),

    # --- Snacks (variables) ---
    ("Août", "Variable", "Snack - Pop-corn", "Maïs pop-corn", 10, "kg", 1.50, 15, "Metro / grossiste", "405 portions"),
    ("Août", "Variable", "Snack - Pop-corn", "Huile + sel pop-corn", 1, "lot", 8, 8, "Metro", ""),
    ("Août", "Variable", "Snack - Frites", "Pommes de terre BFC (frites snack 50 kg + grandes assiettes 110 kg)", 160, "kg", 1.20, 192, "Producteur BFC à sourcer (Bruno ?)", "Total snack + grandes assiettes"),
    ("Août", "Variable", "Snack - Frites", "Huile friture", 18, "L", 1.80, 32, "Metro / grossiste", "Cuisson + renouvellement"),
    ("Août", "Variable", "Snack - Frites", "Sel friture", 1, "lot", 2, 2, "Metro", ""),
    ("Août", "Variable", "Snack - Frites", "Barquettes kraft 150 g (snack à part uniquement)", 360, "unités", 0.05, 18, "Cdiscount Pro / Metro", ""),
    ("Août", "Variable", "Snack - Frites", "Sticks ketchup + mayo", 720, "sticks", 0.05, 36, "Metro / fournisseur HoReCa", "2 sticks/portion snack"),
    ("Août", "Variable", "Snack - Glaces", "Glaces vanille/chocolat/fraise bacs 2,5 L", 39, "bacs", 2.35, 92, "ERHARD", "780 boules · à confirmer prix"),
    ("Août", "Variable", "Snack - Glaces", "Cornets gaufrette comestibles", 560, "unités", 0.12, 67, "Nicolas Glaces / Metro / HoReCa", "Zéro déchet"),
    ("Août", "Variable", "Snack - Glaces", "Gobelets bio 25 cL back-up allergiques/enfants", 50, "unités", 0.06, 3, "Cdiscount Pro", "+ mini-cuillères bois"),
    ("Août", "Variable", "Snack - Tomates", "Tomates cerise BFC", 75, "kg", 1.20, 90, "BRUNO", "500 portions × 150 g"),

    # --- Restauration repas (variables) ---
    ("Août", "Variable", "Restauration - Fromage", "Comté + Comté doux + Morbier (mix 70 kg)", 70, "kg", 14.50, 1015, "Hôpitaux Vieux", "Mix selon assiettes FC + VG"),
    ("Août", "Variable", "Restauration - Charcuterie", "Saucisse paysanne + lard fumé + jambon blanc (~77 kg)", 77, "kg", 15.50, 1194, "Ferme Ligny", "Pour 770 grandes + 231 petites FC"),
    ("Août", "Variable", "Restauration - Œufs", "Œufs plateau 360", 4, "plateaux", 66, 264, "Saveur de la Ferme", "1100 œufs nécessaires"),
    ("Août", "Variable", "Restauration - Pain", "Pain 350 g (10 tranches/pain)", 110, "pains", 2.20, 242, "La Ronde des Pains", "1100 tranches"),
    ("Août", "Variable", "Restauration - Conditionnement", "Sachets sous vide (par 3 kg)", 30, "lots de 3kg", 1.00, 30, "Ferme Ligny", "Pour conserver fromage + charcut J-2"),

    # --- Ajustements ---
    ("Août", "Variable", "Boisson - Bière", "Réassort éventuel 5 fûts (décision Vendredi soir)", 5, "fûts", 80, 0, "Le Pintadier", "Activation conditionnelle - 0 € par défaut"),
]

# ============= NOTES (post-festival, hors scope F&B) =============
NOTES_HORS_FB = """
NOTES IMPORTANTES :
• SACEM Buvette 4,4 % (~1 427 €) : EXCLUE du F&B - à confirmer non due (déjà 8,8 % billetterie)
• SACEM Billetterie 8,8 % (~1 993 €) : payée septembre 2026 (post-festival, dans Frais divers)
• Repas artistes (~555 €) : EXCLUS du F&B - à imputer au budget Artistes
• Coût gobelets bio (café, soft, pop-corn, tomates) inclus dans ligne globale "Vaisselle bio jetable" 319 € (Juillet)
• Réassort bière 5 fûts : montant 0 € par défaut, à activer si forte demande (+400 € coût, +3 200 € recettes)
• Volumes par jour estimés selon modulation Calibrage_buvette_2026.md
"""


def style_row(ws, row, fill=None, font=None, ncols=10):
    for c in range(1, ncols+1):
        cell = ws.cell(row, c)
        if fill: cell.fill = fill
        if font: cell.font = font
        cell.border = BORDER


def main():
    wb = load_workbook(FILE)

    # Supprimer et recréer onglet F&B
    if "F&B" in wb.sheetnames:
        idx = wb.sheetnames.index("F&B")
        del wb["F&B"]
    else:
        idx = 3
    ws = wb.create_sheet("F&B", idx)

    # ----- Largeurs colonnes -----
    widths = [11, 22, 50, 24, 9, 9, 9, 11, 24, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # ----- L1 : Titre -----
    ws['A1'] = "F&B — Buvettes + Restauration (détail max — Scénario B + Flousy)"
    ws.merge_cells('A1:J1')
    ws['A1'].font = H1
    ws['A1'].fill = FILL_TITLE
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    r = 3

    # ============= SECTION RECETTES =============
    ws.cell(r, 1, "RECETTES — Détail par jour et par produit")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(r, 1).font = H2
    ws.cell(r, 1).fill = FILL_REC
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 2

    # Headers recettes
    headers_rec = ["Jour", "Catégorie", "Produit", "Format", "Quantité", "Prix flousy", "Prix €", "Recette €", "", ""]
    for c, h in enumerate(headers_rec, 1):
        if h:
            ws.cell(r, c, h).font = HEADER
            ws.cell(r, c).fill = FILL_HEADER
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = Alignment(horizontal="center")
    r += 1

    rec_start = r
    current_jour = None
    jour_start = r
    fills_jour = {"Vendredi": FILL_VEN, "Samedi": FILL_SAM, "Dimanche": FILL_DIM}
    sous_totaux_jour = {}

    for jour, cat, prod, fmt, qte, fl, prix, rec in RECETTES:
        # Sous-total quand changement de jour
        if current_jour and current_jour != jour:
            ws.cell(r, 1, f"Sous-total {current_jour}")
            ws.cell(r, 8, f"=SUM(H{jour_start}:H{r-1})")
            for c in range(1, 9):
                ws.cell(r, c).font = TOTAL
                ws.cell(r, c).fill = FILL_TOTAL
                ws.cell(r, c).border = BORDER
            sous_totaux_jour[current_jour] = sum(rr[7] for rr in RECETTES if rr[0] == current_jour)
            r += 1
            jour_start = r

        ws.cell(r, 1, jour)
        ws.cell(r, 2, cat)
        ws.cell(r, 3, prod)
        ws.cell(r, 4, fmt)
        ws.cell(r, 5, qte)
        ws.cell(r, 6, fl)
        ws.cell(r, 7, prix)
        ws.cell(r, 8, rec)
        for c in range(1, 9):
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).fill = fills_jour.get(jour, FILL_HEADER)
        current_jour = jour
        r += 1

    # Sous-total dernier jour
    if current_jour:
        ws.cell(r, 1, f"Sous-total {current_jour}")
        ws.cell(r, 8, f"=SUM(H{jour_start}:H{r-1})")
        for c in range(1, 9):
            ws.cell(r, c).font = TOTAL
            ws.cell(r, c).fill = FILL_TOTAL
            ws.cell(r, c).border = BORDER
        sous_totaux_jour[current_jour] = sum(rr[7] for rr in RECETTES if rr[0] == current_jour)
        r += 1

    # TOTAL RECETTES F&B
    total_rec_row = r
    ws.cell(r, 1, "TOTAL RECETTES F&B")
    ws.cell(r, 8, f"=SUM(H{rec_start}:H{r-1})/2")  # /2 car les sous-totaux comptent double
    for c in range(1, 9):
        ws.cell(r, c).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(r, c).fill = FILL_REC
        ws.cell(r, c).border = BORDER
    r += 2

    # ============= SECTION DÉPENSES =============
    ws.cell(r, 1, "DÉPENSES — Détail par achat (achat unique, distinction Type / Catégorie / Mois)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(r, 1).font = H2
    ws.cell(r, 1).fill = FILL_DEP
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 2

    # Headers dépenses
    headers_dep = ["Mois", "Type", "Catégorie", "Produit / Achat", "Quantité", "Unité", "Prix unit €", "Coût total €", "Fournisseur", "Note"]
    for c, h in enumerate(headers_dep, 1):
        ws.cell(r, c, h).font = HEADER
        ws.cell(r, c).fill = FILL_HEADER
        ws.cell(r, c).border = BORDER
        ws.cell(r, c).alignment = Alignment(horizontal="center")
    r += 1

    dep_start = r
    type_colors = {
        "Fixe": PatternFill("solid", fgColor="F3E5F5"),
        "Invest": PatternFill("solid", fgColor="FFE0B2"),
        "Variable": PatternFill("solid", fgColor="FFCDD2"),
    }

    for mois, typ, cat, prod, qte, unit, prix, total, fourn, note in DEPENSES:
        ws.cell(r, 1, mois)
        ws.cell(r, 2, typ)
        ws.cell(r, 3, cat)
        ws.cell(r, 4, prod)
        ws.cell(r, 5, qte)
        ws.cell(r, 6, unit)
        ws.cell(r, 7, prix)
        ws.cell(r, 8, total)
        ws.cell(r, 9, fourn)
        ws.cell(r, 10, note)
        for c in range(1, 11):
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).fill = type_colors.get(typ, FILL_HEADER)
        r += 1

    # Sous-totaux par Type
    r += 1
    types_to_sum = ["Fixe", "Invest", "Variable"]
    for typ in types_to_sum:
        ws.cell(r, 2, f"Sous-total {typ}")
        # Formula sumif
        ws.cell(r, 8, f"=SUMIF(B{dep_start}:B{r-2},\"{typ}\",H{dep_start}:H{r-2})")
        for c in range(1, 11):
            ws.cell(r, c).font = TOTAL
            ws.cell(r, c).fill = type_colors.get(typ, FILL_TOTAL)
            ws.cell(r, c).border = BORDER
        r += 1

    # TOTAL DÉPENSES F&B
    total_dep_row = r
    ws.cell(r, 1, "TOTAL DÉPENSES F&B")
    ws.cell(r, 8, f"=SUM(H{dep_start}:H{dep_start + len(DEPENSES) - 1})")
    for c in range(1, 11):
        ws.cell(r, c).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(r, c).fill = FILL_DEP
        ws.cell(r, c).border = BORDER
    r += 2

    # ============= SECTION SYNTHÈSE =============
    ws.cell(r, 1, "SYNTHÈSE F&B")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(r, 1).font = H2
    ws.cell(r, 1).fill = FILL_SYNTH
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 2

    ws.cell(r, 1, "TOTAL RECETTES F&B")
    ws.cell(r, 8, f"=H{total_rec_row}")
    for c in range(1, 11):
        ws.cell(r, c).font = TOTAL
        ws.cell(r, c).fill = FILL_TOTAL
        ws.cell(r, c).border = BORDER
    rec_synth_row = r
    r += 1

    ws.cell(r, 1, "TOTAL DÉPENSES F&B")
    ws.cell(r, 8, f"=H{total_dep_row}")
    for c in range(1, 11):
        ws.cell(r, c).font = TOTAL
        ws.cell(r, c).fill = FILL_TOTAL
        ws.cell(r, c).border = BORDER
    dep_synth_row = r
    r += 1

    ws.cell(r, 1, "MARGE F&B (sans réassort bière, hors SACEM, hors repas artistes)")
    ws.cell(r, 8, f"=H{rec_synth_row}-H{dep_synth_row}")
    marge_synth_row = r
    for c in range(1, 11):
        ws.cell(r, c).font = Font(bold=True, color="FFFFFF", size=11)
        ws.cell(r, c).fill = FILL_SYNTH
        ws.cell(r, c).border = BORDER
    r += 2

    # ============= NOTES =============
    for line in NOTES_HORS_FB.strip().split("\n"):
        ws.cell(r, 1, line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.cell(r, 1).font = Font(size=9, italic=True)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 16
        r += 1

    # === Mise à jour Synthèse pour pointer sur les nouvelles cellules ===
    ws_synth = wb["Synthèse"]
    ws_synth['B6'] = f"='F&B'!H{rec_synth_row}"
    ws_synth['C6'] = f"='F&B'!I{rec_synth_row}"  # placeholder réelles, à customiser plus tard
    ws_synth['E6'] = f"='F&B'!H{dep_synth_row}"
    ws_synth['F6'] = f"='F&B'!I{dep_synth_row}"

    # Sauvegarde
    wb.save(FILE)

    # Console récap
    total_rec = sum(r[7] for r in RECETTES)
    total_dep = sum(d[7] for d in DEPENSES)
    fixe = sum(d[7] for d in DEPENSES if d[1] == "Fixe")
    invest = sum(d[7] for d in DEPENSES if d[1] == "Invest")
    variable = sum(d[7] for d in DEPENSES if d[1] == "Variable")
    juin = sum(d[7] for d in DEPENSES if d[0] == "Juin")
    juillet = sum(d[7] for d in DEPENSES if d[0] == "Juillet")
    aout = sum(d[7] for d in DEPENSES if d[0] == "Août")

    print("OK F&B sheet refaite avec maximum de detail")
    print(f"  Lignes recettes : {len(RECETTES)} (18 produits x 3 jours)")
    print(f"  Lignes depenses : {len(DEPENSES)}")
    print()
    print(f"  TOTAL RECETTES F&B : {total_rec:>7,} EUR".replace(",", " "))
    for jour in ["Vendredi", "Samedi", "Dimanche"]:
        print(f"    - {jour:>10} : {sum(r[7] for r in RECETTES if r[0]==jour):>7,} EUR".replace(",", " "))
    print()
    print(f"  TOTAL DEPENSES F&B : {total_dep:>7,} EUR".replace(",", " "))
    print(f"    par Type :")
    print(f"    - Fixe     : {fixe:>7,} EUR".replace(",", " "))
    print(f"    - Invest   : {invest:>7,} EUR".replace(",", " "))
    print(f"    - Variable : {variable:>7,} EUR".replace(",", " "))
    print(f"    par Mois :")
    print(f"    - Juin     : {juin:>7,} EUR".replace(",", " "))
    print(f"    - Juillet  : {juillet:>7,} EUR".replace(",", " "))
    print(f"    - Aout     : {aout:>7,} EUR".replace(",", " "))
    print()
    print(f"  MARGE F&B : {total_rec - total_dep:>7,} EUR".replace(",", " "))


if __name__ == "__main__":
    main()
