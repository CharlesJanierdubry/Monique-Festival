"""
Met à jour l'onglet Logistique avec :
- Devis Régis Régis DE2026002 du 26/02/2026 : 5 280 € TTC (4 400 € HT)
- Annotation des lignes "DEMANDE VILLE" pour les items demandés à la Ville de Besançon
- Ajout colonne Source/Devis et Statut

Source devis : Devis/Devis_RegisRegis_DE2026002_2026-02-26.pdf
Source demande Ville : compte-rendu/CR_2026-01-18_Regis-Regis.md (matériel demandé) + CR_2026-03-02.md (statut Ville)

Usage : python budget/update_logistique_avec_regis.py
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

REPO = Path(__file__).resolve().parent.parent
FILE = REPO / "budget" / "Suivi_budgetaire_Monique_Festival.xlsx"

THIN = Side(border_style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H1 = Font(bold=True, size=13, color="FFFFFF")
HEADER = Font(bold=True, size=10)
TOTAL = Font(bold=True, size=11)
FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_HEADER = PatternFill("solid", fgColor="EEEEEE")
FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")
FILL_VILLE = PatternFill("solid", fgColor="FFE0B2")  # Orange clair pour items demandés Ville
FILL_DEVIS = PatternFill("solid", fgColor="C8E6C9")  # Vert clair pour items avec devis confirmé

# Format : (Description, Prévu €, Statut, Source/Devis, Note)
LIGNES_LOGISTIQUE = [
    # === RÉGIS RÉGIS (Devis confirmé) ===
    ("Régis Régis — Rédaction dossier sécurité (3j × 400 € HT, TTC)", 1440, "Devis reçu",
     "Devis DE2026002 du 26/02/2026", "Vigipirate + sanitaire + sûreté + plans Vectorworks"),
    ("Régis Régis — Préparation régie site (2j × 400 € HT, TTC)", 960, "Devis reçu",
     "Devis DE2026002 du 26/02/2026", "Compilation besoins + analyse faisabilité + coordination"),
    ("Régis Régis — Régisseur Montage 2j + Exploitation 3j + Démontage 1j (TTC)", 2880, "Devis reçu",
     "Devis DE2026002 du 26/02/2026", "Régisseur site présent du 26/08 au 31/08 - repas à fournir asso"),

    # === LOCATION LIEU ===
    ("Location salle/terrain La Grange Huguenet", 3000, "À négocier",
     "Convention propriétaire à signer mai 2026", "Cf. Calendrier_demarches_administratives.md"),

    # === FOURNISSEURS À DEVIS ===
    ("Sonorisation pro (scène principale + petites scènes)", 3000, "À devis",
     "Devis non reçu", "À demander - prestataire local Besançon"),
    ("Éclairage / lumière scène + ambiance", 2000, "À devis",
     "Devis non reçu", "À demander en cohérence avec sonorisation"),
    ("Scène / podium principal", 2000, "À devis OU Demande Ville",
     "Devis non reçu", "DEMANDE VILLE : podium roulant rouge et blanc (CR 18/01)"),
    ("Toilettes / sanitaires Trivial Compost (cabine PMR + urinoir + sanitaires compost)", 1500, "À devis",
     "Trivial Compost - devis à demander", "Charte éthique : toilettes sèches"),
    ("Électricité / groupe électrogène", 1500, "À devis",
     "Devis non reçu", "Vérifier ampérage avec Grange + percolateurs/friteuses F&B"),

    # === DEMANDES VILLE BESANÇON (potentiellement gratuites si accordées) ===
    ("Barrières Heras 3×35 + Vauban 300 unités", 1000, "DEMANDE VILLE",
     "Demande visite faite 02/03/2026", "0 € si Ville accorde - 1000 € sinon (achat/location)"),
    ("Mobilier : 25 tables + 50 bancs + 200 chaises", 1000, "DEMANDE VILLE",
     "Demande visite faite 02/03/2026", "0 € si Ville accorde - 1000 € sinon (location)"),
    ("Poubelles double flux × 5", 0, "DEMANDE VILLE",
     "Demande visite faite 02/03/2026", "Mention CR 18/01 - probablement 0 € si Ville accorde"),
    ("Vitabris (tentes pliables avec murs et lestes)", 500, "DEMANDE VILLE",
     "Demande visite faite 02/03/2026", "0 € si Ville accorde - 500 € sinon (location)"),

    # === DOSSIER ARBRES (gros risque budget) ===
    ("État sanitaire des arbres", 0, "DEMANDE VILLE - aide partenariat",
     "CR 02/03 - en attente réponse Ville", "Coût plein 5-15 K€ si à charge asso. Négocier prise en charge Ville/Grange/CFA"),

    # === PETIT MATÉRIEL ===
    ("Talkies-walkies × 4", 100, "À acheter",
     "Détail dans Billetterie.md L309", "Ou prêt si possible"),
]


def main():
    wb = load_workbook(FILE)
    if "Logistique" not in wb.sheetnames:
        sys.exit("ERREUR : onglet Logistique absent")

    idx = wb.sheetnames.index("Logistique")
    del wb["Logistique"]
    ws = wb.create_sheet("Logistique", idx)

    # Largeurs
    widths = [55, 10, 22, 35, 50, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # Titre
    ws['A1'] = "LOGISTIQUE — Dépenses (incl. Régis Régis + demandes Ville Besançon)"
    ws.merge_cells('A1:F1')
    ws['A1'].font = H1
    ws['A1'].fill = FILL_TITLE
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header
    headers = ["Description", "Prévu €", "Statut", "Source / Devis", "Note", "Réel €"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h).font = HEADER
        ws.cell(3, c).fill = FILL_HEADER
        ws.cell(3, c).border = BORDER
        ws.cell(3, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Lignes
    r = 4
    for desc, prev, statut, source, note in LIGNES_LOGISTIQUE:
        ws.cell(r, 1, desc)
        ws.cell(r, 2, prev)
        ws.cell(r, 3, statut)
        ws.cell(r, 4, source)
        ws.cell(r, 5, note)
        # Coloration selon statut
        if "DEMANDE VILLE" in statut:
            fill = FILL_VILLE
        elif "Devis reçu" in statut:
            fill = FILL_DEVIS
        else:
            fill = None
        for c in range(1, 7):
            ws.cell(r, c).border = BORDER
            if fill:
                ws.cell(r, c).fill = fill
            ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)
        r += 1

    # TOTAL
    total_row = r
    ws.cell(r, 1, "TOTAL LOGISTIQUE (hypothèse haute - Ville n'accorde rien)")
    ws.cell(r, 2, f"=SUM(B4:B{r-1})")
    ws.cell(r, 6, f"=SUM(F4:F{r-1})")
    for c in range(1, 7):
        ws.cell(r, c).font = TOTAL
        ws.cell(r, c).fill = FILL_TOTAL
        ws.cell(r, c).border = BORDER
    r += 1

    # TOTAL si Ville accorde matériel mobilier+barrières+vitabris
    ws.cell(r, 1, "TOTAL LOGISTIQUE (hypothèse basse - Ville accorde barrières + mobilier + vitabris)")
    ws.cell(r, 2, f"=B{total_row}-1000-1000-500")
    for c in range(1, 7):
        ws.cell(r, c).font = TOTAL
        ws.cell(r, c).fill = FILL_VILLE
        ws.cell(r, c).border = BORDER
    r += 2

    # Notes
    notes = [
        "LÉGENDE :",
        "  Vert : Devis reçu et confirmé (Régis Régis 5 280 € TTC)",
        "  Orange : Demande à la Ville de Besançon (en attente confirmation - 0 € si accordé, montant indiqué sinon)",
        "  Sans couleur : À demander en devis ou négocier",
        "",
        "DÉTAIL DEVIS RÉGIS RÉGIS DE2026002 (échéance 05/03/2026 - À RECONFIRMER) :",
        "  - 4 400 € HT + TVA 20% = 5 280 € TTC",
        "  - Date d'échéance dépassée → demander devis actualisé ou confirmation validité",
        "  - À PRÉVOIR par asso : repas du régisseur pendant montage/exploitation/démontage (6 jours)",
        "",
        "DEMANDES VILLE BESANÇON (CR 18/01 + 02/03) - statut au 03/05/2026 :",
        "  - Visite du lieu pour matériel et sécurité demandée le 02/03/2026 (en attente réponse)",
        "  - Voir fiche détaillée : officiel/partenaires/Demandes_Ville_Besancon.md",
        "  - Économie potentielle si toutes demandes accordées : -2 500 € (barrières + mobilier + vitabris)",
        "",
        "HORS PÉRIMÈTRE LOGISTIQUE (gérés ailleurs) :",
        "  - Percolateurs ABC LOCATION (devis D0022382 = 265 € TTC) → onglet F&B",
        "  - Friteuses ABC LOCATION → onglet F&B (à confirmer dans devis séparé)",
        "  - Iphone billetterie + jetons billetterie + gobelets billetterie → cf. Billetterie.md (1 831 €)",
    ]
    for n in notes:
        ws.cell(r, 1, n)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1).font = Font(size=9, italic=True)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 16
        r += 1

    wb.save(FILE)

    # Console
    total = sum(l[1] for l in LIGNES_LOGISTIQUE)
    ville_demandes = sum(l[1] for l in LIGNES_LOGISTIQUE if "DEMANDE VILLE" in l[2])
    devis_recus = sum(l[1] for l in LIGNES_LOGISTIQUE if "Devis reçu" in l[2])
    a_devis = sum(l[1] for l in LIGNES_LOGISTIQUE if "À devis" in l[2] or "À acheter" in l[2] or "À négocier" in l[2])
    print("OK Logistique sheet refaite avec Régis Régis + demandes Ville")
    print(f"  Total prévu (hyp haute) : {total:>6,} EUR".replace(",", " "))
    print(f"  Devis confirmes (Regis) : {devis_recus:>6,} EUR".replace(",", " "))
    print(f"  Demandes Ville          : {ville_demandes:>6,} EUR (0 si accorde)".replace(",", " "))
    print(f"  A devis / a acheter     : {a_devis:>6,} EUR".replace(",", " "))


if __name__ == "__main__":
    main()
