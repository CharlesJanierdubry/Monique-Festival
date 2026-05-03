"""
Met à jour l'onglet "Artistes" de budget/Suivi_budgetaire_Monique_Festival.xlsx
avec la programmation actualisée (sources Drive du 28/04/2026).

Changements appliqués :
- Renommage des compagnies théâtre (ex appel à projet) avec leurs vrais noms
- Renommage Karina Ziegler → Kaïnza (catégorie Musique au lieu de Classique)
- Renommage Fannelie → Fanélie Nava
- Noms réels des violoncellistes Europe Cellists
- Ajout de Vladimir Torres Trio (Constantia Prod, Jazz)
- Ajout de Cie L'entre (V.I.E.N.S samedi - à confirmer)
- Suppression de "? (Ven 19h30)" (créneau attribué à Fanélie)
- Ajustement du nombre d'artistes des Cies théâtre (réels au lieu de 4 par défaut)

Les autres onglets ne sont pas modifiés.
Usage : python budget/update_artistes_suivi_budgetaire.py
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

sys.stdout.reconfigure(encoding='utf-8')

REPO = Path(__file__).resolve().parent.parent
FILE = REPO / "budget" / "Suivi_budgetaire_Monique_Festival.xlsx"

# Tarifs unitaires standards
CACHET = 250
DEFRAI_PAR_PERS = 200
REP_TARIF = 100  # par jour de répétition (cas Fabulo, ateliers)

# Source : programmation officielle confirmée 28/04/2026
# Format : (Nom, Catégorie, Atelier, Nb cachets, Nb rép., Coût cachets, Coût rép., Défraiement, Total)
ARTISTES = [
    # Musique actuelle - solos / duos
    ("Georges Cabaret (Sète, 34) - France hors BFC", "Musique", "", 1, 0, 250, 0, 150, 400),
    ("Loretta (locale)", "Musique", "", 1, 0, 250, 0, 0, 250),
    ("Fanélie Nava", "Musique", "", 1, 0, 250, 0, 150, 400),  # ex "? (Ven 19h30)" — défraiement estimé 150€ (plafond 200€)
    ("Manu - Wambo (local)", "Musique", "", 1, 0, 250, 0, 0, 250),
    ("Antoine Vermot - Jeff The Fool (local, parents BFC)", "Musique", "Atelier DJ", 0, 2, 0, 200, 0, 200),
    # Musique actuelle - collectifs
    ("Wet Enough?! (5 art.)", "Musique — Collectif", "", 5, 0, 1250, 0, 750, 2000),
    ("Osmosis (3 art.)", "Musique — Collectif", "", 3, 0, 750, 0, 450, 1200),
    ("Vladimir Torres Trio (3 art.) - France hors BFC", "Musique — Collectif", "", 3, 0, 750, 0, 450, 1200),  # Jazz, Constantia Prod (Paris probable, 150€/pers)
    ("Kaïnza (3 art.)", "Musique — Collectif", "", 3, 0, 750, 0, 450, 1200),  # ex Karina Ziegler
    # Musique - famille
    ("Léopold (famille)", "Musique", "Atelier DJ", 0, 2, 0, 0, 0, 0),
    ("Luther (famille)", "Musique", "", 0, 0, 0, 0, 0, 0),
    # Fabulo
    ("Marius (Fabulo)", "Fabulo — Collectif", "Musique & théâtre", 2, 2, 500, 200, 150, 850),
    ("Maya (Fabulo)", "Fabulo — Collectif", "Musique & théâtre", 2, 2, 500, 200, 150, 850),
    ("Raphaëlle (Fabulo)", "Fabulo — Collectif", "Musique & théâtre", 2, 2, 500, 200, 150, 850),
    ("Romane Cabaret (vient avec Lorraine)", "Fabulo — Collectif", "Masterclass chant", 4, 2, 1000, 200, 0, 1200),
    ("Lorraine (famille)", "Fabulo", "Musique & théâtre", 0, 0, 0, 0, 0, 0),
    # Classique - Europe Cellists (noms réels)
    ("Marc Trembovelski", "Classique", "", 2, 0, 500, 0, 200, 700),  # UK — étranger 200€
    ("Sophie Ehling", "Classique", "", 2, 0, 500, 0, 200, 700),       # NL — étranger 200€
    ("Jakub Wycislik", "Classique", "", 2, 0, 500, 0, 200, 700),      # PL — étranger 200€
    ("Beatriz Correia", "Classique", "", 2, 0, 500, 0, 200, 700),     # PT — étranger 200€
    ("Volodia Van Keulen", "Classique", "", 2, 0, 500, 0, 150, 650),  # France 150€
    ("Augustin Lemonnier (pianiste)", "Classique", "", 1, 0, 250, 0, 150, 400),  # France 150€
    ("Nils Van Keulen (Masterclass)", "Classique", "Masterclass chant", 0, 1, 0, 100, 0, 100),  # NOUVEAU
    # Théâtre
    ("Sylvain Septours (vient avec Lewis)", "Théâtre", "Atelier écriture", 0, 2, 0, 200, 0, 200),
    ("Marius Ponnelle (Treize Clique)", "Théâtre — Collectif", "", 1, 0, 250, 0, 150, 400),
    ("Comédien 2 (Treize Clique)", "Théâtre — Collectif", "", 1, 0, 250, 0, 150, 400),
    ("Comédien 3 (Treize Clique)", "Théâtre — Collectif", "", 1, 0, 250, 0, 150, 400),
    ("Lewis (famille)", "Théâtre", "Éloquence", 0, 0, 0, 0, 0, 0),
    # Théâtre - cies retenues par appel à projet (compagnies locales)
    # Pas de défraiement (compagnies locales — décision Bureau)
    ("Cie Nenni ma foi - Patatra (4 art.)", "Théâtre — Coll AP", "", 4, 0, 1000, 0, 0, 1000),
    ("Cie Vous faites un feu? (2 art.)", "Théâtre — Coll AP", "", 2, 0, 500, 0, 0, 500),
    ("Cie Derrière le mur - Trafic (2 art.)", "Théâtre — Coll AP", "", 2, 0, 500, 0, 0, 500),
    # Cie L'Entre = 1 seul spectacle V.I.E.N.S (Vers Ici Entre Nos rienS), dimanche 19h30-20h
    # Référente : Juliette Jeanmougin — Pelousey 25170 (BFC)
    # Toujours Primevère candidature AAP non retenue
    ("Cie L'Entre - V.I.E.N.S dimanche (3 art.)", "Théâtre — Coll AP", "", 3, 0, 750, 0, 0, 750),
    # Classique - autres
    ("À travers la fenêtre des heures (2 art.) - Belgique", "Classique — Coll", "", 2, 0, 500, 0, 400, 900),  # 200€ × 2 (étrangers)
]


def main():
    if not FILE.exists():
        sys.exit(f"ERREUR : {FILE} introuvable")

    wb = load_workbook(FILE)
    if "Artistes" not in wb.sheetnames:
        sys.exit("ERREUR : onglet 'Artistes' absent")

    ws = wb["Artistes"]

    # Sauver styles d'en-tête (ligne 1) et tableau (ligne 2)
    title_style = {
        'font': copy(ws.cell(1, 1).font),
        'fill': copy(ws.cell(1, 1).fill),
        'alignment': copy(ws.cell(1, 1).alignment),
    }
    header_styles = []
    for c in range(1, 15):
        cell = ws.cell(2, c)
        header_styles.append({
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'alignment': copy(cell.alignment),
            'border': copy(cell.border),
        })
    body_border = copy(ws.cell(3, 1).border)
    total_styles = []
    last_row = ws.max_row
    for c in range(1, 15):
        cell = ws.cell(last_row, c)
        total_styles.append({
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
        })

    # Effacer toutes les lignes de données (à partir ligne 3)
    for row in range(ws.max_row, 2, -1):
        ws.delete_rows(row)

    # Réécrire les artistes
    row = 3
    for nom, cat, atelier, nb_c, nb_r, cout_c, cout_r, defrai, total in ARTISTES:
        ws.cell(row, 1, nom)
        ws.cell(row, 2, cat)
        ws.cell(row, 3, atelier)
        ws.cell(row, 4, nb_c)
        ws.cell(row, 5, nb_r)
        ws.cell(row, 6, cout_c)
        ws.cell(row, 7, cout_r)
        ws.cell(row, 8, defrai)
        ws.cell(row, 9, total)
        ws.cell(row, 10, "")  # Paiement réel
        ws.cell(row, 11, "")  # Défraiement réel
        ws.cell(row, 12, 0)   # Total réel
        ws.cell(row, 13, "")  # Date paiement
        ws.cell(row, 14, -total)  # Écart (= -prévu si réel = 0)
        for c in range(1, 15):
            ws.cell(row, c).border = body_border
        row += 1

    # Ligne TOTAL ARTISTES avec formules SUM
    ws.cell(row, 1, "TOTAL ARTISTES")
    ws.cell(row, 4, f"=SUM(D3:D{row-1})")
    ws.cell(row, 5, f"=SUM(E3:E{row-1})")
    ws.cell(row, 6, f"=SUM(F3:F{row-1})")
    ws.cell(row, 7, f"=SUM(G3:G{row-1})")
    ws.cell(row, 8, f"=SUM(H3:H{row-1})")
    ws.cell(row, 9, f"=SUM(I3:I{row-1})")
    ws.cell(row, 10, f"=SUM(J3:J{row-1})")
    ws.cell(row, 11, f"=SUM(K3:K{row-1})")
    ws.cell(row, 12, f"=SUM(L3:L{row-1})")
    ws.cell(row, 14, f"=SUM(N3:N{row-1})")
    for c in range(1, 15):
        cell = ws.cell(row, c)
        s = total_styles[c-1]
        cell.font = s['font']
        cell.fill = s['fill']
        cell.border = s['border']
        cell.alignment = s['alignment']

    # Sauvegarde
    wb.save(FILE)
    print(f"OK : {FILE.relative_to(REPO)}")
    print(f"  - {len(ARTISTES)} artistes/groupes inscrits")

    # Calcul des totaux pour info
    tot_cachets = sum(a[5] for a in ARTISTES)
    tot_rep = sum(a[6] for a in ARTISTES)
    tot_defrai = sum(a[7] for a in ARTISTES)
    tot = sum(a[8] for a in ARTISTES)
    print(f"  - Total cachets prévus : {tot_cachets} €")
    print(f"  - Total répétitions prévues : {tot_rep} €")
    print(f"  - Total défraiements prévus : {tot_defrai} €")
    print(f"  - TOTAL ARTISTES PRÉVU : {tot} €")


if __name__ == "__main__":
    main()
