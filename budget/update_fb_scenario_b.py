"""
Réécrit l'onglet F&B du Suivi_budgetaire selon le Scénario B validé le 03/05/2026 :
- Sans SACEM (à imputer au budget Droits d'auteur, pas au F&B)
- Sans repas artistes (à imputer au budget Artistes)
- Distinction explicite Coûts FIXES / INVEST écocups / Coûts VARIABLES
- Recettes par jour : Buvette / Restauration repas / Snacks / Consignes

Usage : python budget/update_fb_scenario_b.py
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

REPO = Path(__file__).resolve().parent.parent
FILE = REPO / "budget" / "Suivi_budgetaire_Monique_Festival.xlsx"

# Styles
THIN = Side(border_style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H1 = Font(bold=True, size=13, color="FFFFFF")
H2 = Font(bold=True, size=11, color="FFFFFF")
HEADER = Font(bold=True, size=10)
TOTAL = Font(bold=True, size=11)
FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_RECETTES = PatternFill("solid", fgColor="2E7D32")
FILL_FIXES = PatternFill("solid", fgColor="6A1B9A")
FILL_INVEST = PatternFill("solid", fgColor="EF6C00")
FILL_VAR = PatternFill("solid", fgColor="C62828")
FILL_SYNTH = PatternFill("solid", fgColor="0277BD")
FILL_HEADER = PatternFill("solid", fgColor="EEEEEE")
FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")

# ===========================================================
# DONNÉES — Scénario B (assiettes 9 € sans frites / 12 € avec)
# Mix : 70 % avec frites, 70 % FC / 30 % VG, 1 100 menus
# Modulation jour : Ven 270 fest / Sam 780 / Dim 270
# ===========================================================

RECETTES = [
    # (description, type, montant)
    ("Buvette Vendredi : bière 1200 + vin 540 + soft 793 + café 81", "Recette", 2614),
    ("Buvette Samedi : bière 4600 + vin 1950 + soft 3439 + café 477 (jour fort)", "Recette", 10466),
    ("Buvette Dimanche : bière 600 + vin 340 + soft 1188 + café 162", "Recette", 2290),
    ("Restauration repas Vendredi : 225 assiettes (70% Grande 12€ / 30% Petite 9€, FC ou végé)", "Recette", 2498),
    ("Restauration repas Samedi : 650 assiettes (Petite 9€ / Grande 12€, FC ou végé)", "Recette", 7215),
    ("Restauration repas Dimanche : 225 assiettes (Petite 9€ / Grande 12€, FC ou végé)", "Recette", 2498),
    ("Snacks Vendredi : pop-corn 60 (40p × 1,50€) + frites 120 + glaces 109 + tomates 150 (100p × 1,50€)", "Recette", 439),
    ("Snacks Samedi : pop-corn 428 (285p × 1,50€) + frites 705 + glaces 850 + tomates 435 (290p × 1,50€)", "Recette", 2418),
    ("Snacks Dimanche : pop-corn 120 (80p × 1,50€) + frites 240 + glaces 218 + tomates 165 (110p × 1,50€)", "Recette", 743),
    ("Consignes écocup non remboursées (1 000 écocups × 20 % non-retour × 1 €)", "Recette", 200),
]
# Total recettes attendu : 31 822 €

COUTS_FIXES = [
    ("Percolateurs 4 × 10 L (ABC LOCATION)", "Dépense", 345),
    ("Friteuses × 2 (ABC LOCATION)", "Dépense", 172),
    ("Machine pop-corn (location)", "Dépense", 252),
    ("Eau gratuite : fontaines + carafes (charge éthique)", "Dépense", 400),
    ("HACCP : tablier + charlotte + gants jetables + sonde T° + désinfectant", "Dépense", 80),
    ("Lavage écocups + sacs poubelles tri + récup huile usagée", "Dépense", 120),
    ("Trousse premiers secours buvette", "Dépense", 30),
    ("Affichage prix (ardoises + marqueurs) + caisse + ustensiles préparation", "Dépense", 310),
]
# Total fixes : 1 709 €

INVEST = [
    ("Écocups réutilisables Atomic — 1 000 pièces 25 cL (réutilisables éditions futures)", "Invest", 600),
]
# Total invest : 600 € (consignes 200 € comptabilisées en recettes ligne 12)

COUTS_VAR = [
    ("Bière Le Pintadier — 10 fûts initial (commande ferme)", "Dépense", 800),
    ("Bière Le Pintadier — 5 fûts réassort éventuel (à activer vendredi soir si forte demande — +3200 € recettes)", "Dépense", 0),
    ("Vin Cubi Jura 5 L (Hyper Boisson, prix négocié 50 €/cubi) × 29 cubis", "Dépense", 1450),
    ("Soft : sirop 300 + jus 350 + Mortuacienne 58 + eau pét. 35 + eau plate 32", "Dépense", 775),
    ("Café : poudre BFC + sucre + lait + mélangeurs + gobelets bio", "Dépense", 110),
    ("Sucre bar additionnel + glace pilée bar", "Dépense", 60),
    ("Pop-corn (1 flousy=1,50€) : maïs 10kg (15) + huile/sel (8) + gobelets bio 72cL (40) — 405 portions", "Dépense", 63),
    ("Frites snack : pdt 50kg BFC (60) + huile (32) + sel (2) + barquettes (18) + sauces (36) — 360 portions", "Dépense", 148),
    ("Glaces ERHARD (1 boule=1,50€/2 boules=3€) : 39 bacs 2,5L (94) + cornets gaufrette 560 (67) + back-up (3) — 560 portions, mix 60/40", "Dépense", 164),
    ("Tomates cerise BRUNO (1 flousy=1,50€) : 75kg × 1,20€/kg (90) + gobelets bio 25cL (32) — 500 portions", "Dépense", 122),
    ("Assiettes : 770 grandes (3,07 €/u) + 330 petites (2,48 €/u) — mix 70% FC / 30% végé", "Dépense", 3182),
    ("Frites incluses dans grandes assiettes : 770 × 0,42 €", "Dépense", 323),
    ("Vaisselle bio jetable : gobelets soft+café+pop-corn + serviettes + back-up bière 300v", "Dépense", 319),
]
# Total variables sans réassort : 7 431 €
# Total variables avec réassort : 7 831 €
# Note : SACEM 4,4 % CA F&B exclue (à imputer au budget Droits d'auteur)
# Note : Repas artistes 555 € exclus (à imputer au budget Artistes)


def style_row(ws, row, fill=None, font=None, border=True, align=None):
    for c in range(1, 7):
        cell = ws.cell(row, c)
        if fill: cell.fill = fill
        if font: cell.font = font
        if border: cell.border = BORDER
        if align: cell.alignment = align


def main():
    wb = load_workbook(FILE)
    if "F&B" not in wb.sheetnames:
        sys.exit("ERREUR : onglet F&B absent")

    # Supprimer entièrement l'onglet F&B et le recréer pour repartir propre
    idx = wb.sheetnames.index("F&B")
    del wb["F&B"]
    ws = wb.create_sheet("F&B", idx)

    # ----- Largeurs colonnes -----
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # ----- L1 : titre -----
    ws['A1'] = "F&B — Buvettes + Restauration (Scénario B — validé 03/05/2026)"
    ws.merge_cells('A1:F1')
    ws['A1'].font = H1
    ws['A1'].fill = FILL_TITLE
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ----- L3 : section RECETTES -----
    ws['A3'] = "RECETTES par jour"
    ws.merge_cells('A3:F3')
    ws['A3'].font = H2
    ws['A3'].fill = FILL_RECETTES
    ws['A3'].alignment = Alignment(horizontal="center")

    # ----- L5 : header -----
    headers = ["Date", "Description", "Type", "Prévu", "Réel", "Écart"]
    for c, h in enumerate(headers, 1):
        ws.cell(5, c, h).font = HEADER
        ws.cell(5, c).fill = FILL_HEADER
        ws.cell(5, c).border = BORDER
        ws.cell(5, c).alignment = Alignment(horizontal="center")

    # ----- L6+ : Recettes -----
    r = 6
    for desc, typ, montant in RECETTES:
        ws.cell(r, 2, desc)
        ws.cell(r, 3, typ)
        ws.cell(r, 4, montant)
        ws.cell(r, 6, f"=E{r}-D{r}")
        style_row(ws, r)
        r += 1

    # ----- TOTAL RECETTES -----
    total_rec_row = r
    ws.cell(r, 1, "TOTAL RECETTES F&B")
    ws.cell(r, 4, f"=SUM(D6:D{r-1})")
    ws.cell(r, 5, f"=SUM(E6:E{r-1})")
    ws.cell(r, 6, f"=E{r}-D{r}")
    style_row(ws, r, fill=FILL_TOTAL, font=TOTAL)
    r += 2

    # ----- Section COÛTS FIXES -----
    ws.cell(r, 1, "COÛTS FIXES (incompressibles — indépendants de la fréquentation)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1).font = H2
    ws.cell(r, 1).fill = FILL_FIXES
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 2

    for c, h in enumerate(headers, 1):
        ws.cell(r, c, h).font = HEADER
        ws.cell(r, c).fill = FILL_HEADER
        ws.cell(r, c).border = BORDER
        ws.cell(r, c).alignment = Alignment(horizontal="center")
    r += 1

    fixes_start = r
    for desc, typ, montant in COUTS_FIXES:
        ws.cell(r, 2, desc)
        ws.cell(r, 3, typ)
        ws.cell(r, 4, montant)
        ws.cell(r, 6, f"=E{r}-D{r}")
        style_row(ws, r)
        r += 1
    fixes_end = r - 1

    total_fixes_row = r
    ws.cell(r, 1, "TOTAL FIXES")
    ws.cell(r, 4, f"=SUM(D{fixes_start}:D{fixes_end})")
    ws.cell(r, 5, f"=SUM(E{fixes_start}:E{fixes_end})")
    ws.cell(r, 6, f"=E{r}-D{r}")
    style_row(ws, r, fill=FILL_TOTAL, font=TOTAL)
    r += 2

    # ----- Section INVEST -----
    ws.cell(r, 1, "INVESTISSEMENT semi-fixe (amortissable sur éditions futures)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1).font = H2
    ws.cell(r, 1).fill = FILL_INVEST
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 2

    for c, h in enumerate(headers, 1):
        ws.cell(r, c, h).font = HEADER
        ws.cell(r, c).fill = FILL_HEADER
        ws.cell(r, c).border = BORDER
        ws.cell(r, c).alignment = Alignment(horizontal="center")
    r += 1

    invest_start = r
    for desc, typ, montant in INVEST:
        ws.cell(r, 2, desc)
        ws.cell(r, 3, typ)
        ws.cell(r, 4, montant)
        ws.cell(r, 6, f"=E{r}-D{r}")
        style_row(ws, r)
        r += 1
    invest_end = r - 1

    total_invest_row = r
    ws.cell(r, 1, "TOTAL INVEST")
    ws.cell(r, 4, f"=SUM(D{invest_start}:D{invest_end})")
    ws.cell(r, 5, f"=SUM(E{invest_start}:E{invest_end})")
    ws.cell(r, 6, f"=E{r}-D{r}")
    style_row(ws, r, fill=FILL_TOTAL, font=TOTAL)
    r += 2

    # ----- Section COÛTS VARIABLES -----
    ws.cell(r, 1, "COÛTS VARIABLES (proportionnels à la fréquentation, ajustables)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1).font = H2
    ws.cell(r, 1).fill = FILL_VAR
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 2

    for c, h in enumerate(headers, 1):
        ws.cell(r, c, h).font = HEADER
        ws.cell(r, c).fill = FILL_HEADER
        ws.cell(r, c).border = BORDER
        ws.cell(r, c).alignment = Alignment(horizontal="center")
    r += 1

    var_start = r
    for desc, typ, montant in COUTS_VAR:
        ws.cell(r, 2, desc)
        ws.cell(r, 3, typ)
        ws.cell(r, 4, montant)
        ws.cell(r, 6, f"=E{r}-D{r}")
        style_row(ws, r)
        r += 1
    var_end = r - 1

    total_var_row = r
    ws.cell(r, 1, "TOTAL VARIABLES")
    ws.cell(r, 4, f"=SUM(D{var_start}:D{var_end})")
    ws.cell(r, 5, f"=SUM(E{var_start}:E{var_end})")
    ws.cell(r, 6, f"=E{r}-D{r}")
    style_row(ws, r, fill=FILL_TOTAL, font=TOTAL)
    r += 2

    # ----- SYNTHÈSE -----
    ws.cell(r, 1, "SYNTHÈSE F&B")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1).font = H2
    ws.cell(r, 1).fill = FILL_SYNTH
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 2

    ws.cell(r, 2, "TOTAL COÛTS F&B (Fixes + Invest + Variables)")
    ws.cell(r, 4, f"=D{total_fixes_row}+D{total_invest_row}+D{total_var_row}")
    ws.cell(r, 5, f"=E{total_fixes_row}+E{total_invest_row}+E{total_var_row}")
    ws.cell(r, 6, f"=E{r}-D{r}")
    style_row(ws, r, fill=FILL_TOTAL, font=TOTAL)
    r += 1

    ws.cell(r, 2, "TOTAL RECETTES F&B")
    ws.cell(r, 4, f"=D{total_rec_row}")
    ws.cell(r, 5, f"=E{total_rec_row}")
    ws.cell(r, 6, f"=E{r}-D{r}")
    style_row(ws, r, fill=FILL_TOTAL, font=TOTAL)
    r += 1

    marge_row = r
    ws.cell(r, 2, "MARGE F&B (sans réassort bière) — hors SACEM (Droits d'auteur) et hors repas artistes (Artistes)")
    ws.cell(r, 4, f"=D{r-1}-D{r-2}")
    ws.cell(r, 5, f"=E{r-1}-E{r-2}")
    ws.cell(r, 6, f"=E{r}-D{r}")
    style_row(ws, r, fill=FILL_SYNTH, font=Font(bold=True, color="FFFFFF", size=11))
    r += 2

    # ----- Notes -----
    ws.cell(r, 1, "Notes")
    ws.cell(r, 1).font = HEADER
    r += 1
    notes = [
        "• Réassort bière (5 fûts +400 € coûts ; +3 200 € recettes) : décision vendredi soir selon consommation observée",
        "• Marge avec réassort = ~24 882 € (vs 22 082 € sans). Activable à coût marginal nul.",
        "• SACEM (4,4 % buvette) EXCLUE : à confirmer auprès de SACEM BFC qu'elle n'est pas due (déjà 8,8 % billetterie)",
        "• Repas artistes (~555 €) EXCLUS : à imputer au budget Artistes (cohérence avec convention défraiement)",
        "• Charges fixes 1 709 € + invest écocups 600 € = seuil de couverture ~127 fest cumulés (largement atteint)",
        "• Marge variable par festivalier (sans réassort) = ~18,30 €/fest",
    ]
    for n in notes:
        ws.cell(r, 1, n)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1).font = Font(size=9, italic=True)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 16
        r += 1

    # Sauvegarde
    wb.save(FILE)

    # Récap console
    total_rec = sum(m for _, _, m in RECETTES)
    total_fixes = sum(m for _, _, m in COUTS_FIXES)
    total_invest = sum(m for _, _, m in INVEST)
    total_var = sum(m for _, _, m in COUTS_VAR)
    total_couts = total_fixes + total_invest + total_var
    marge = total_rec - total_couts

    print("OK F&B sheet rewritten - Scénario B")
    print(f"  Total RECETTES   : {total_rec:>8,} €".replace(",", " "))
    print(f"  Total FIXES      : {total_fixes:>8,} €".replace(",", " "))
    print(f"  Total INVEST     : {total_invest:>8,} €".replace(",", " "))
    print(f"  Total VARIABLES  : {total_var:>8,} €".replace(",", " "))
    print(f"  Total COÛTS      : {total_couts:>8,} €".replace(",", " "))
    print(f"  MARGE F&B        : {marge:>8,} €".replace(",", " "))
    print(f"  Avec réassort 5 fûts (+400 coûts, +3200 recettes) : marge {marge + 3200 - 400:>5,} €".replace(",", " "))


if __name__ == "__main__":
    main()
