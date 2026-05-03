"""
Génère, envoie par email et archive sur Drive les 14 reçus fiscaux rétroactifs
pour les dons Hello Asso du 20-23 avril 2026.

Pour chaque donateur :
 1. Génération PDF (ReportLab) avec logo + signature
 2. Upload du PDF sur Drive (dossier 'Administratif/Reçus fiscaux dons crowdfunding 2026-04')
 3. Envoi par email via l'API Gmail (vouvoiement)
 4. Log dans Registre_emission_14_retros.csv

Usage :
    python scripts/send_14_recus_retroactifs.py

Prérequis :
    - gmail_auth.py lancé (scope gmail.send)
    - APIs Gmail + Drive activées côté Google Cloud
    - Logo_Monique_Festival.png et Signature_Judith_.png à la racine
"""
import sys
import os
import csv
import json
import base64
import time
from pathlib import Path
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders

sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, HRFlowable
)

# ================== CONFIG ==================
TOKEN_FILE = os.path.expanduser('~/.config/mcp-gdrive/token.json')
REPO_ROOT = Path(__file__).resolve().parent.parent
LOGO_PATH = REPO_ROOT / "Logo_Monique_Festival.png"
SIGNATURE_PATH = REPO_ROOT / "Signature_Judith_.png"
EXPORT_XLSX = REPO_ROOT / "export-le-monique-festival-1-jd-production-20_04_2026-23_04_2026.xlsx"
OUTPUT_DIR = REPO_ROOT / "officiel/dons/recus_emis"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
REGISTRE_CSV = OUTPUT_DIR / "Registre_emission_14_retros.csv"

# Raison sociale choisie pour les 14 reçus rétroactifs
# "MONIQUE FESTIVAL" = validé lors du test v2
# "JD PRODUCTION"    = cohérent avec le reçu HelloAsso auto #15
RAISON_SOCIALE_CHOISIE = "MONIQUE FESTIVAL"

ASSO = {
    "raison_sociale": RAISON_SOCIALE_CHOISIE,
    "ancien_nom": "anciennement JD Production" if RAISON_SOCIALE_CHOISIE == "MONIQUE FESTIVAL" else "",
    "objet": (
        "Organisation annuelle d'un festival pluridisciplinaire à Besançon et en Bourgogne-Franche-Comté, "
        "mêlant musique classique, musiques actuelles, théâtre et arts vivants ; soutien à la création "
        "artistique émergente ; actions d'éducation artistique et culturelle, de médiation et de transmission, "
        "notamment vers les publics éloignés de la pratique ; valorisation du territoire ; promotion de la "
        "diffusion de la culture, à titre non lucratif et d'intérêt général."
    ),
    "rna": "W251011013",
    "siren": "991 055 450",
    "siret": "991 055 450 00019",
    "adresse": "54 chemin de Valentin",
    "cp": "25000",
    "ville": "Besançon",
    "pays": "France",
    "signataire": "Judith LAITHIER",
    "fonction": "Présidente",
    "ville_sig": "Besançon",
}

# Numérotation des reçus rétroactifs : distincte des numéros HelloAsso auto
DATE_EMISSION_FMT = datetime.now().strftime("%d%m%Y")  # "24042026"
DATE_EMISSION_AFFICHEE = datetime.now().strftime("%d/%m/%Y")
NUMERO_PREFIX = f"{DATE_EMISSION_FMT}-MF-RETRO"

# Drive : dossier cible
DRIVE_PARENT_NAME = "Administratif"
DRIVE_NEW_FOLDER = f"Reçus fiscaux — Dons crowdfunding 20-23 avril 2026"


# ================== AUTH ==================

def load_creds():
    with open(TOKEN_FILE) as f:
        data = json.load(f)
    required = {'https://www.googleapis.com/auth/gmail.send', 'https://www.googleapis.com/auth/drive'}
    if not required.issubset(set(data.get('scopes', []))):
        sys.exit(
            "ERREUR : scopes manquants dans le token.\n"
            f"Actuels : {data.get('scopes')}\n"
            f"Requis : {required}\n"
            "Relancer : python scripts/gmail_auth.py"
        )
    creds = Credentials(
        token=data['token'],
        refresh_token=data.get('refresh_token'),
        token_uri=data['token_uri'],
        client_id=data['client_id'],
        client_secret=data['client_secret'],
        scopes=data['scopes'],
    )
    if not creds.valid and creds.refresh_token:
        creds.refresh(Request())
        data['token'] = creds.token
        with open(TOKEN_FILE, 'w') as f:
            json.dump(data, f, indent=2)
    return creds


# ================== NOMBRE EN LETTRES (FR, 0-9999) ==================

def nombre_en_lettres(n):
    units = ["", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf",
             "dix", "onze", "douze", "treize", "quatorze", "quinze", "seize",
             "dix-sept", "dix-huit", "dix-neuf"]
    tens = ["", "", "vingt", "trente", "quarante", "cinquante", "soixante",
            "soixante", "quatre-vingt", "quatre-vingt"]

    def below_100(n):
        if n < 20:
            return units[n]
        t = n // 10
        u = n % 10
        if t == 7:
            base = tens[6]
            return base + ("-et-onze" if u == 1 else f"-{units[10 + u]}" if u else f"-{units[10]}")
        if t == 9:
            base = tens[8]
            return base + f"-{units[10 + u]}" if u else base + f"-{units[10]}"
        base = tens[t]
        if u == 0:
            return base + ("s" if t == 8 else "")
        if u == 1 and t not in (8,):
            return base + "-et-un"
        return f"{base}-{units[u]}"

    if n == 0: return "zéro"
    if n < 100: return below_100(n)
    if n < 1000:
        c, r = n // 100, n % 100
        prefix = "cent" if c == 1 else f"{units[c]} cents"
        if r:
            prefix = prefix.rstrip("s") + f" {below_100(r)}"
        return prefix
    if n < 10000:
        m, r = n // 1000, n % 1000
        prefix = "mille" if m == 1 else f"{units[m]} mille"
        if r:
            if r < 100:
                prefix += f" {below_100(r)}"
            else:
                c, rr = r // 100, r % 100
                cpart = "cent" if c == 1 else f"{units[c]} cents"
                if rr:
                    cpart = cpart.rstrip("s") + f" {below_100(rr)}"
                prefix += f" {cpart}"
        return prefix
    return str(n)


# ================== PDF ==================

def build_pdf(don, output_path: Path):
    doc = SimpleDocTemplate(
        str(output_path), pagesize=A4,
        topMargin=1.8*cm, bottomMargin=1.8*cm, leftMargin=2.2*cm, rightMargin=2.2*cm,
        title=f"Reçu fiscal {don['numero_recu']} — {ASSO['raison_sociale']}",
        author=ASSO['raison_sociale'],
    )
    styles = getSampleStyleSheet()
    brand = colors.HexColor("#d12d2d")
    dark = colors.HexColor("#111111")
    grey = colors.HexColor("#555555")
    light = colors.HexColor("#f7f7f7")
    story = []

    # ======= Header : logo + identite =======
    logo_img = Image(str(LOGO_PATH), width=2.2*cm, height=2.2*cm) if LOGO_PATH.exists() else Paragraph("", styles['Normal'])
    st_title = ParagraphStyle(name="ht", fontSize=18, leading=22, textColor=dark, fontName="Helvetica-Bold")
    st_sub = ParagraphStyle(name="hs", fontSize=8, leading=10, textColor=grey, fontName="Helvetica-Oblique")
    st_line = ParagraphStyle(name="hl", fontSize=8, leading=10, textColor=colors.HexColor("#777777"))
    identite_block = [
        Paragraph(ASSO['raison_sociale'], st_title),
    ]
    if ASSO['ancien_nom']:
        identite_block.append(Paragraph(ASSO['ancien_nom'], st_sub))
    identite_block.extend([
        Paragraph("Association Loi 1901 — Œuvre ou organisme d'intérêt général — Arts et culture",
                  ParagraphStyle(name="hss", fontSize=8, leading=10, textColor=grey)),
        Paragraph(f"RNA {ASSO['rna']} · SIREN {ASSO['siren']} · SIRET {ASSO['siret']}", st_line),
        Paragraph(f"{ASSO['adresse']} — {ASSO['cp']} {ASSO['ville']} — {ASSO['pays']}", st_line),
    ])
    header_table = Table([[logo_img, identite_block]], colWidths=[2.6*cm, None])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=1.5, color=dark))
    story.append(Spacer(1, 14))

    # Titre
    story.append(Paragraph('<para alignment="center" fontSize="16"><b>REÇU AU TITRE DES DONS</b></para>', styles['Normal']))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f'<para alignment="center" fontSize="10" textColor="#555555">Article {don["article_cgi"]} du Code Général des Impôts</para>',
        styles['Normal']))
    story.append(Spacer(1, 12))

    # Numero
    numero_style = ParagraphStyle(name="num", alignment=TA_CENTER, fontSize=11, textColor=dark,
                                   borderColor=dark, borderWidth=1, borderPadding=6, leading=13)
    story.append(Paragraph(f"<b>N° d'ordre : {don['numero_recu']}</b>", numero_style))
    story.append(Spacer(1, 14))

    # Donateur + Benef
    style_block = ParagraphStyle(name="blk", fontSize=10, leading=14)
    style_block_title = ParagraphStyle(name="blktit", fontSize=9, textColor=grey, spaceAfter=4, fontName="Helvetica-Bold")
    col_don = [
        Paragraph("DONATEUR", style_block_title),
        Paragraph(
            f"<b>{don['donateur_prenom']} {don['donateur_nom']}</b><br/>"
            f"{don['donateur_adresse']}<br/>"
            f"{don['donateur_cp']} {don['donateur_ville']}<br/>"
            f"{don['donateur_pays']}",
            style_block),
    ]
    col_benef = [
        Paragraph("BÉNÉFICIAIRE", style_block_title),
        Paragraph(
            f"<b>{ASSO['raison_sociale']}</b><br/>"
            f"{ASSO['adresse']}<br/>"
            f"{ASSO['cp']} {ASSO['ville']}<br/>"
            f"{ASSO['pays']}",
            style_block),
    ]
    blocks_table = Table([[col_don, col_benef]], colWidths=[None, None])
    blocks_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (0, 0), 0), ('RIGHTPADDING', (1, 0), (1, 0), 0),
        ('LEFTPADDING', (1, 0), (1, 0), 12),
    ]))
    story.append(blocks_table)
    story.append(Spacer(1, 14))

    # Objet
    story.append(Paragraph("OBJET DE L'ASSOCIATION", style_block_title))
    objet_style = ParagraphStyle(name="objet", fontSize=9.5, leading=13, leftIndent=10, rightIndent=10)
    objet_box = Table([[Paragraph(ASSO['objet'], objet_style)]], colWidths=[None])
    objet_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), light),
        ('LINEBEFORE', (0, 0), (0, -1), 3, brand),
        ('LEFTPADDING', (0, 0), (-1, -1), 12), ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 10), ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(objet_box)
    story.append(Spacer(1, 14))

    # Montant
    montant_content = [
        Paragraph(f"<b>{don['montant']} € ({don['montant_lettres']})</b>",
                  ParagraphStyle(name="mnt", alignment=TA_CENTER, fontSize=14, textColor=dark, leading=18)),
        Spacer(1, 4),
        Paragraph(
            f"Date du don : <b>{don['date_don']}</b> &nbsp;·&nbsp; Forme : Don manuel &nbsp;·&nbsp; "
            f"Nature : Numéraire &nbsp;·&nbsp; Moyen de paiement : {don['moyen_paiement']}",
            ParagraphStyle(name="mntsub", alignment=TA_CENTER, fontSize=9.5, textColor=grey, leading=12)),
    ]
    montant_box = Table([[montant_content]], colWidths=[None])
    montant_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#fff4f4")),
        ('BOX', (0, 0), (-1, -1), 1, brand),
        ('LEFTPADDING', (0, 0), (-1, -1), 12), ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12), ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    story.append(montant_box)
    story.append(Spacer(1, 14))

    # Certification
    cert_style = ParagraphStyle(name="cert", fontSize=9, textColor="#444444", leading=12, fontName="Helvetica-Oblique",
                                leftIndent=4, rightIndent=4)
    cert_text = (
        f"Le bénéficiaire certifie sur l'honneur que les dons et versements qu'il reçoit ouvrent droit à la "
        f"réduction d'impôt prévue à l'article {don['article_cgi']} du Code Général des Impôts.<br/><br/>"
        f"<b>Particulier</b> : vous pouvez bénéficier d'une réduction d'impôt égale à "
        f"<b>{don['taux_reduction']} %</b> du montant de votre don, dans la limite de 20 % de votre revenu imposable."
    )
    cert_box = Table([[Paragraph(cert_text, cert_style)]], colWidths=[None])
    cert_box.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 0.5, colors.HexColor("#cccccc")),
        ('LINEBELOW', (0, -1), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ('LEFTPADDING', (0, 0), (-1, -1), 10), ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(cert_box)
    story.append(Spacer(1, 22))

    # Lieu + signature
    lieudate_style = ParagraphStyle(name="ld", fontSize=10, leading=14)
    lieudate_html = f"Fait à <b>{ASSO['ville_sig']}</b>,<br/>le <b>{don['date_emission']}</b>"
    sig_img = Image(str(SIGNATURE_PATH), width=3.5*cm, height=1.5*cm, kind='proportional') if SIGNATURE_PATH.exists() else Paragraph('<i>(signature manuscrite)</i>', lieudate_style)
    sig_block = [
        sig_img,
        Paragraph(ASSO['signataire'], ParagraphStyle(name="sn", alignment=TA_CENTER, fontSize=10, leading=13, fontName="Helvetica-Bold")),
        Paragraph(f"{ASSO['fonction']} de l'association {ASSO['raison_sociale']}",
                  ParagraphStyle(name="sf", alignment=TA_CENTER, fontSize=9, textColor=grey, leading=12)),
    ]
    sig_table = Table([[Paragraph(lieudate_html, lieudate_style), sig_block]], colWidths=[7*cm, None])
    sig_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
    ]))
    story.append(sig_table)
    story.append(Spacer(1, 18))

    # Footer
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#bbbbbb"), dash=(2, 2)))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "Reçu émis conformément au modèle Cerfa n° 11580*04. À conserver pour votre déclaration d'impôts.",
        ParagraphStyle(name="ftr", fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor("#888888"))))

    doc.build(story)
    return output_path


# ================== EMAIL ==================

EMAIL_HTML = """\
<html>
<body style="font-family: Helvetica, Arial, sans-serif; font-size: 11pt; color: #222; line-height: 1.55;">
<div style="text-align:center; margin-bottom: 18px;">
  <img src="cid:logo" alt="Monique Festival" style="width: 90px; height: auto;">
</div>

<p>Bonjour {prenom},</p>

<p>Un immense merci pour votre soutien au <strong>Monique Festival</strong> !
Vous avez versé un don de <strong>{montant} €</strong> le {date_don} sur notre
campagne de crowdfunding — votre geste nous aide à concrétiser la première
édition qui se tiendra à La Grange Huguenet (Besançon) les
<strong>28, 29 et 30 août 2026</strong>.</p>

<p>Vous trouverez en pièce jointe votre <strong>reçu fiscal</strong> au format
PDF (Cerfa n° 11580*04). Il vous permet de bénéficier d'une réduction d'impôt
sur le revenu de <strong>{taux_reduction} %</strong> du montant de votre don,
à reporter dans votre déclaration 2027 (revenus 2026).</p>

<p><em>Précision administrative : notre association, initialement appelée
JD Production, change de dénomination pour Monique Festival suite à notre
Assemblée Générale du 19 avril 2026. La personne morale et les numéros
d'enregistrement (RNA W251011013, SIREN 991 055 450) sont inchangés.</em></p>

<p>La programmation se dévoile progressivement sur notre compte Instagram
<a href="https://instagram.com/monique.festival">@monique.festival</a> —
n'hésitez pas à nous suivre.</p>

<p>À très bientôt, et encore merci de votre confiance.</p>

<p>
<strong>Judith Laithier</strong><br>
Présidente de l'association Monique Festival<br>
<a href="mailto:info@monique-festival.fr">info@monique-festival.fr</a>
</p>

<p style="font-size: 9pt; color: #888; margin-top: 24px; border-top: 1px dashed #bbb; padding-top: 8px;">
Monique Festival — Festival émergent et pluridisciplinaire à Besançon<br>
28, 29 et 30 août 2026 — La Grange Huguenet<br>
Instagram : <a href="https://instagram.com/monique.festival">@monique.festival</a>
</p>
</body>
</html>
"""


def build_email(destinataire, objet, prenom, montant, date_don, taux_reduction, pdf_path, logo_path):
    message = MIMEMultipart('related')
    message['To'] = destinataire
    message['Subject'] = objet

    alt = MIMEMultipart('alternative')
    message.attach(alt)
    html_body = EMAIL_HTML.format(
        prenom=prenom, montant=montant, date_don=date_don,
        taux_reduction=taux_reduction,
    )
    alt.attach(MIMEText(html_body, 'html', 'utf-8'))

    if logo_path.exists():
        with open(logo_path, 'rb') as f:
            img = MIMEImage(f.read(), _subtype='png')
            img.add_header('Content-ID', '<logo>')
            img.add_header('Content-Disposition', 'inline', filename=logo_path.name)
            message.attach(img)

    with open(pdf_path, 'rb') as f:
        pdf_attach = MIMEBase('application', 'pdf')
        pdf_attach.set_payload(f.read())
    encoders.encode_base64(pdf_attach)
    pdf_attach.add_header('Content-Disposition', f'attachment; filename="{pdf_path.name}"')
    message.attach(pdf_attach)

    raw = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
    return {'raw': raw}


# ================== DRIVE ==================

def drive_find_folder_by_name(service, name, parent_id=None):
    q = [
        "mimeType='application/vnd.google-apps.folder'",
        "trashed=false",
        f"name='{name}'",
    ]
    if parent_id:
        q.append(f"'{parent_id}' in parents")
    res = service.files().list(
        q=" and ".join(q),
        fields="files(id, name, parents)",
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    return res.get('files', [])


def drive_create_folder(service, name, parent_id):
    metadata = {'name': name, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [parent_id]}
    created = service.files().create(body=metadata, fields='id, name', supportsAllDrives=True).execute()
    return created


def drive_upload_pdf(service, pdf_path: Path, folder_id):
    existing_q = f"name='{pdf_path.name}' and '{folder_id}' in parents and trashed=false"
    res = service.files().list(q=existing_q, fields="files(id)",
                                supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
    files = res.get('files', [])
    media = MediaFileUpload(str(pdf_path), mimetype='application/pdf', resumable=False)
    if files:
        return service.files().update(fileId=files[0]['id'], media_body=media, supportsAllDrives=True).execute()
    metadata = {'name': pdf_path.name, 'parents': [folder_id], 'mimeType': 'application/pdf'}
    return service.files().create(body=metadata, media_body=media, fields='id, name', supportsAllDrives=True).execute()


# ================== MAIN ==================

def read_donors(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = None
    rows = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = list(row)
            continue
        if not any(c is not None and str(c).strip() for c in row):
            continue
        rec = dict(zip(headers, row))
        if rec.get("Statut de la commande") != "Validé" or rec.get("Type") != "Don unique":
            continue
        rows.append(rec)
    rows.sort(key=lambda r: r.get("Date de la commande"))
    return rows


def slug(s):
    import re
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def main():
    if not LOGO_PATH.exists():
        sys.exit(f"ERREUR : logo introuvable : {LOGO_PATH}")
    if not SIGNATURE_PATH.exists():
        print(f"ATTENTION : signature introuvable ({SIGNATURE_PATH}) — PDF sans image de signature")
    if not EXPORT_XLSX.exists():
        sys.exit(f"ERREUR : export HelloAsso introuvable : {EXPORT_XLSX}")

    print(f"=== Reçus rétroactifs — 14 dons crowdfunding 20-23 avril 2026 ===\n")
    print(f"Raison sociale choisie : {RAISON_SOCIALE_CHOISIE}")
    print(f"Numérotation : {NUMERO_PREFIX}-001 à -014")
    print(f"Date d'émission : {DATE_EMISSION_AFFICHEE}\n")

    # --- Auth ---
    print("Chargement credentials Google (Gmail + Drive)...")
    creds = load_creds()
    gmail_service = build('gmail', 'v1', credentials=creds)
    drive_service = build('drive', 'v3', credentials=creds)

    # --- Drive : trouver 'Administratif' + créer sous-dossier ---
    print(f"\nRecherche du dossier '{DRIVE_PARENT_NAME}' sur Drive...")
    admin_folders = drive_find_folder_by_name(drive_service, DRIVE_PARENT_NAME)
    if not admin_folders:
        sys.exit(f"ERREUR : dossier '{DRIVE_PARENT_NAME}' introuvable sur Drive")
    if len(admin_folders) > 1:
        print(f"  Plusieurs dossiers '{DRIVE_PARENT_NAME}' trouvés :")
        for f in admin_folders:
            print(f"   - id={f['id']}")
        print(f"  -> utilisation du 1er : {admin_folders[0]['id']}")
    admin_folder = admin_folders[0]
    print(f"  OK : {admin_folder['name']} (id {admin_folder['id']})")

    # Crée ou réutilise le sous-dossier
    existing_sub = drive_find_folder_by_name(drive_service, DRIVE_NEW_FOLDER, admin_folder['id'])
    if existing_sub:
        subfolder = existing_sub[0]
        print(f"  Sous-dossier existant : {subfolder['name']} (id {subfolder['id']})")
    else:
        subfolder = drive_create_folder(drive_service, DRIVE_NEW_FOLDER, admin_folder['id'])
        print(f"  Sous-dossier créé : {subfolder['name']} (id {subfolder['id']})")
    drive_folder_id = subfolder['id']
    drive_folder_link = f"https://drive.google.com/drive/folders/{drive_folder_id}"

    # --- Lecture donateurs ---
    donors = read_donors(EXPORT_XLSX)
    print(f"\n{len(donors)} donateur(s) détecté(s) dans l'export.\n")

    registre_rows = [["N° reçu", "Date don", "Donateur", "Email", "Montant", "PDF local", "Drive ID", "Gmail ID", "Statut"]]

    for i, don_raw in enumerate(donors, start=1):
        numero = f"{NUMERO_PREFIX}-{i:03d}"
        date_don_raw = don_raw.get("Date de la commande")
        date_don = date_don_raw.strftime("%d/%m/%Y") if isinstance(date_don_raw, datetime) else str(date_don_raw).split(" ")[0]

        montant = int(don_raw.get("Montant") or 0)
        prenom = (don_raw.get("Prénom") or "").strip()
        nom = (don_raw.get("Nom") or "").strip()
        email = don_raw.get("Email") or ""

        # Particulier par défaut
        article = "200"
        taux = 66
        if don_raw.get("Raison sociale") or don_raw.get("SIREN"):
            article = "238 bis"
            taux = 60

        don = {
            "numero_recu": numero,
            "donateur_prenom": prenom,
            "donateur_nom": nom,
            "donateur_adresse": don_raw.get("Adresse") or "",
            "donateur_cp": str(don_raw.get("Code postal") or ""),
            "donateur_ville": don_raw.get("Ville") or "",
            "donateur_pays": don_raw.get("Pays") or "",
            "montant": montant,
            "montant_lettres": nombre_en_lettres(montant) + " euros",
            "date_don": date_don,
            "moyen_paiement": don_raw.get("Moyen de paiement") or "",
            "date_emission": DATE_EMISSION_AFFICHEE,
            "article_cgi": article,
            "taux_reduction": str(taux),
        }

        print(f"--- [{i}/{len(donors)}] {prenom} {nom} · {montant} € · {email}")
        pdf_name = f"RF_{numero}_{slug(prenom)}_{slug(nom)}.pdf"
        pdf_path = OUTPUT_DIR / pdf_name
        drive_id = ""
        gmail_id = ""
        statut = "OK"

        try:
            # 1. PDF
            build_pdf(don, pdf_path)
            print(f"    PDF: {pdf_path.name} ({pdf_path.stat().st_size // 1024} KB)")

            # 2. Drive
            drive_file = drive_upload_pdf(drive_service, pdf_path, drive_folder_id)
            drive_id = drive_file['id']
            print(f"    Drive upload OK (id {drive_id})")

            # 3. Gmail
            subject = f"Monique Festival — Votre reçu fiscal pour votre don du {date_don}"
            msg = build_email(email, subject, prenom, montant, date_don, taux, pdf_path, LOGO_PATH)
            sent = gmail_service.users().messages().send(userId='me', body=msg).execute()
            gmail_id = sent['id']
            print(f"    Email envoyé (Gmail ID {gmail_id})")

        except Exception as e:
            statut = f"ERREUR: {e}"
            print(f"    ÉCHEC: {e}")

        registre_rows.append([numero, date_don, f"{prenom} {nom}", email, f"{montant} EUR", pdf_name, drive_id, gmail_id, statut])

        # Petit délai entre envois (éviter rate-limit Gmail : 250 quota/s mais soft limit perso)
        time.sleep(0.3)

    # Registre CSV
    with REGISTRE_CSV.open('w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerows(registre_rows)

    # Bilan
    total = len(donors)
    ok = sum(1 for r in registre_rows[1:] if r[-1] == "OK")
    print(f"\n=== BILAN ===")
    print(f"{ok}/{total} reçus traités avec succès")
    print(f"\nRegistre local : {REGISTRE_CSV.relative_to(REPO_ROOT)}")
    print(f"Dossier Drive  : {drive_folder_link}")


if __name__ == "__main__":
    main()
