"""Génère le fichier Excel du planning jour J avec gestion des conflits."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# Styles
header_font = Font(bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
alert_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
alert_font = Font(bold=True, size=10, color="FFFFFF")
scene_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ok_fill = PatternFill(start_color="D5F5D5", end_color="D5F5D5", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
day_font = Font(bold=True, size=12, color="2F5496")

def add_headers(ws, row):
    headers = ["Créneau", "Événement", "Artistes-orga sur scène", "Remplaçant(s)", "Tâches orga à couvrir", "Alerte"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return row + 1

def add_row(ws, row, data, is_alert=False, is_scene=False):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if is_alert and col >= 3:
            cell.fill = alert_fill
            cell.font = alert_font
        elif is_scene and col == 3:
            cell.fill = scene_fill
    return row + 1

def set_widths(ws):
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 15

# ===== VENDREDI =====
ws = wb.active
ws.title = "Vendredi"
ws.merge_cells("A1:F1")
ws["A1"] = "VENDREDI 26 SEPTEMBRE 2026"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center")

row = add_headers(ws, 3)

vendredi = [
    (["14h-15h30", "Montage / installation", "-", "Tous disponibles", "Installation technique, restauration", ""], False),
    (["15h30", "Ouverture des portes", "-", "Tous disponibles", "Billetterie, accueil, sécurité", ""], False),
    (["15h55", "Discours du festival", "-", "Judith (Présidente)", "Discours officiel", ""], False),
    (["16h-16h30", "Duo Cabaret", "ROMANE", "Loretta couvre éthique", "Safe place, stands éthiques", ""], True),
    (["17h-17h45", "Concert Fabulo", "LORRAINE", "Luther (com), Clotilde (artistes)", "Posts live, accueil artistes suivants", "ATTENTION"], True),
    (["18h15-19h", "Cie Treize Clique", "-", "Tous disponibles", "Repos Lorraine", ""], False),
    (["19h30-20h15", "Artiste à définir", "-", "Tous disponibles", "Prépa technique Luther & Loretta", ""], False),
    (["20h45-21h45", "Luther & Loretta", "LUTHER + LORETTA", "Léopold (tech), Charles (com)", "Régie technique, posts live", "CRITIQUE"], True),
    (["22h15-23h15", "Wet Enough!?", "-", "Tous disponibles", "", ""], False),
    (["23h-00h", "Fermeture", "-", "Tous disponibles", "Rangement, sécurisation", ""], False),
]

for data, is_alert in vendredi:
    row = add_row(ws, row, data, is_alert=is_alert, is_scene=bool(data[2] != "-"))

set_widths(ws)

# ===== SAMEDI =====
ws2 = wb.create_sheet("Samedi")
ws2.merge_cells("A1:F1")
ws2["A1"] = "SAMEDI 27 SEPTEMBRE 2026"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A1"].alignment = Alignment(horizontal="center")

ws2.merge_cells("A2:F2")
ws2["A2"] = "ALERTE : 4 organisateurs sur scène de 10h à 12h"
ws2["A2"].font = Font(bold=True, size=11, color="FF0000")
ws2["A2"].alignment = Alignment(horizontal="center")

row = add_headers(ws2, 4)

samedi = [
    (["9h-10h", "Préparation / balances", "-", "Léopold supervise", "Installation ateliers", ""], False),
    (["10h-12h", "Atelier Fabulo", "LORRAINE", "Clotilde couvre orga", "Accueil, validation artistes", "CRITIQUE"], True),
    (["10h-12h", "Atelier éloquence", "LEWIS", "Clotilde couvre restauration", "Suivi prépa restauration", "CRITIQUE"], True),
    (["10h-12h", "Atelier DJ", "LÉOPOLD", "Charles couvre artistes", "Contact artistes, urgences tech", "CRITIQUE"], True),
    (["10h-12h", "Masterclass chant", "ROMANE", "Loretta couvre éthique", "Stands éthiques", "CRITIQUE"], True),
    (["12h-12h45", "Pause / transition", "-", "Tous disponibles", "Debriefing ateliers", ""], False),
    (["12h45-13h30", "Wambo", "-", "Tous disponibles", "", ""], False),
    (["14h-15h", "Cie théâtre", "-", "Tous disponibles", "", ""], False),
    (["15h30-16h40", "À travers les fenêtres...", "-", "Tous disponibles", "", ""], False),
    (["17h10-18h10", "Fannelie", "-", "Tous disponibles", "", ""], False),
    (["18h40-19h30", "Cie théâtre", "-", "Tous disponibles", "", ""], False),
    (["20h-21h10", "Europe Cellists meet Augustin", "-", "Tous disponibles", "", ""], False),
    (["21h35-22h35", "Osmosis", "-", "Tous disponibles", "", ""], False),
    (["23h", "Fermeture + AFTER", "-", "Léopold gère l'after", "", ""], False),
]

for data, is_alert in samedi:
    row = add_row(ws2, row, data, is_alert=is_alert, is_scene=bool(data[2] != "-"))

# Équipe dispo 10h-12h
row += 1
ws2.merge_cells(f"A{row}:F{row}")
ws2[f"A{row}"] = "Équipe disponible 10h-12h : Luther, Clotilde, Charles, Loretta, Judith, Wadih"
ws2[f"A{row}"].font = Font(bold=True, size=11, color="2F5496")
row += 1

remplacement_headers = ["Rôle à couvrir", "Remplaçant", "Contact urgence"]
for col, h in enumerate(remplacement_headers, 1):
    cell = ws2.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
row += 1

remplacements = [
    ["Technique (Léopold absent)", "Luther", "Léopold entre morceaux"],
    ["Restauration (Lewis absent)", "Clotilde", "Lewis pendant pauses"],
    ["Relations Artistes (Léopold absent)", "Charles", "Léopold par talkie"],
    ["Communication / réseaux", "Luther", "-"],
    ["Accueil public / billetterie", "Wadih", "Judith"],
    ["Coordination générale (tour de contrôle)", "Clotilde", "-"],
]
for r in remplacements:
    for col, val in enumerate(r, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
    row += 1

set_widths(ws2)

# ===== DIMANCHE =====
ws3 = wb.create_sheet("Dimanche")
ws3.merge_cells("A1:F1")
ws3["A1"] = "DIMANCHE 28 SEPTEMBRE 2026"
ws3["A1"].font = Font(bold=True, size=14)
ws3["A1"].alignment = Alignment(horizontal="center")

ws3.merge_cells("A2:F2")
ws3["A2"] = "JOUR LE PLUS CHARGÉ : Lorraine dans 4 performances"
ws3["A2"].font = Font(bold=True, size=11, color="FF0000")
ws3["A2"].alignment = Alignment(horizontal="center")

row = add_headers(ws3, 4)

dimanche = [
    (["10h-12h", "Atelier Fabulo", "LORRAINE", "Clotilde couvre orga", "Idem samedi", "CRITIQUE"], True),
    (["10h-12h", "Atelier éloquence", "LEWIS", "Clotilde couvre restauration", "Idem samedi", "CRITIQUE"], True),
    (["10h-12h", "Atelier DJ", "LÉOPOLD", "Charles couvre artistes", "Idem samedi", "CRITIQUE"], True),
    (["10h-12h", "Masterclass chant", "ROMANE", "Loretta couvre éthique", "Idem samedi", "CRITIQUE"], True),
    (["11h30-12h", "Générale tous ateliers", "LOR+LEW+LÉO+ROM", "Même équipe", "Idem", "CRITIQUE"], True),
    (["12h15-14h", "Suites de Bach", "LORRAINE (violoncelle)", "Luther (com), Clotilde (artistes)", "Posts live, accueil artiste suivant", "ATTENTION"], True),
    (["14h-14h30", "PAUSE LORRAINE", "-", "-", "Repos obligatoire Lorraine", ""], False),
    (["14h30-15h", "Volodia Sitariste", "-", "Tous disponibles", "", ""], False),
    (["15h30-16h30", "Cie théâtre", "-", "Tous disponibles", "", ""], False),
    (["17h-17h45", "Nos Voix", "LORRAINE + ROMANE", "Luther (com), Loretta (éthique)", "Posts live, stands", "ATTENTION"], True),
    (["17h45-18h15", "PAUSE LORRAINE", "-", "-", "Repos obligatoire Lorraine", ""], False),
    (["18h15-19h", "Karina Ziegler", "-", "Tous disponibles", "", ""], False),
    (["19h30-20h15", "Cie théâtre", "-", "Tous disponibles", "Prépa Création Monique", ""], False),
    (["20h30-21h", "Création Monique", "LORRAINE (Fabulo)", "Luther (com), Clotilde (coord)", "Posts live, clôture technique", "ATTENTION"], True),
    (["21h-22h", "Fermeture / clôture", "-", "Tous disponibles", "Rangement, remerciements", ""], False),
]

for data, is_alert in dimanche:
    row = add_row(ws3, row, data, is_alert=is_alert, is_scene=bool(data[2] != "-" and "PAUSE" not in data[1]))

set_widths(ws3)

# ===== SYNTHÈSE =====
ws4 = wb.create_sheet("Synthèse indisponibilités")
ws4.merge_cells("A1:D1")
ws4["A1"] = "Synthèse des indisponibilités par membre"
ws4["A1"].font = Font(bold=True, size=14)
ws4["A1"].alignment = Alignment(horizontal="center")

row = 3
headers = ["Membre", "Vendredi", "Samedi", "Dimanche"]
for col, h in enumerate(headers, 1):
    cell = ws4.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
row += 1

synth = [
    ["Lorraine", "17h-17h45", "10h-12h", "10h-12h, 12h15-14h, 17h-17h45, 20h30-21h"],
    ["Luther", "20h45-21h45", "Disponible", "Disponible"],
    ["Lewis", "Disponible", "10h-12h", "10h-12h"],
    ["Léopold", "Disponible", "10h-12h", "10h-12h"],
    ["Romane", "16h-16h30", "10h-12h", "10h-12h, 17h-17h45"],
    ["Clotilde", "Disponible", "Disponible", "Disponible"],
    ["Charles", "Disponible", "Disponible", "Disponible"],
    ["Judith", "Disponible", "Disponible", "Disponible"],
    ["Wadih", "Disponible", "Disponible", "Disponible"],
]

for s in synth:
    for col, val in enumerate(s, 1):
        cell = ws4.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if val == "Disponible":
            cell.fill = ok_fill
        elif val not in ("", s[0]):
            cell.fill = scene_fill
    row += 1

ws4.column_dimensions["A"].width = 15
ws4.column_dimensions["B"].width = 20
ws4.column_dimensions["C"].width = 20
ws4.column_dimensions["D"].width = 45

output = os.path.join(os.path.dirname(__file__), "Planning_jour_J_Monique_Festival.xlsx")
wb.save(output)
print(f"Fichier créé : {output}")
