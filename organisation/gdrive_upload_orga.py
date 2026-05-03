"""
Cree un xlsx Organisation generale puis l'uploade sur le Drive.

⚠️ SORTIE = DOCUMENT INTERNE. Le xlsx produit contient le nom d'un·e
membre mineur·e (Leontine) : a stocker dans un dossier Drive a acces
restreint (equipe operationnelle uniquement), ne pas diffuser a l'exterieur.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import json, os

TOKEN_FILE = os.path.expanduser('~/.config/mcp-gdrive/token.json')
FOLDER_ID = '16O8yDq2Jvsx36OzzsWWhIiWEzyp245Kv'
OUT_FILE = 'c:/Users/charl/Monique-Festival/organisation/Organisation_generale_Monique_Festival.xlsx'

# ── DATA ──
poles = [
    {
        'name': 'Direction Artistique', 'chef': 'Lewis', 'hex': 'E07B54',
        'avant': [
            ('Programmation', [('Vision artistique globale', 'Lewis'), ('Musique Actuelle', 'Luther'), ('Musique Classique', 'Lorraine'), ('Theatre', 'Lewis')]),
            ('Action culturelle', [('Action culturelle / Ateliers', 'Lorraine'), ('Animations', 'Loretta *'), ('Engagement ethique', 'Loretta + Leontine *')]),
            ('Ambiance & site', [('Ambiance du site (deco, animations, loges)', 'Lewis')]),
        ],
        'pendant': [
            ('Coordination artistique', [('Coordination artistique globale', 'Lewis'), ('Liaison artistes musique actuelle', 'Luther'), ('Liaison artistes classique', 'Lorraine'), ('Liaison compagnies theatre', 'Lewis')]),
            ('Action culturelle', [('Animations jour J', 'Loretta *'), ('Safe place + stands ethiques', 'Loretta + Leontine *')]),
            ('Ambiance & site', [('Animations', 'Lewis')]),
        ],
    },
    {
        'name': 'Communication', 'chef': 'Lorraine', 'hex': 'D4A843',
        'avant': [
            ('Strategie & contenu', [('Strategie de communication', 'Lorraine'), ('Contenu editorial (newsletters, interviews)', 'Lorraine')]),
            ('Digital & visuel', [('Communication visuelle (charte, affiches)', 'Luther'), ('Site internet', 'Luther'), ('Reseaux sociaux', 'Luther'), ('Signaletique / Goodies', 'Luther')]),
            ('Relations exterieures', [('Relations presse', 'Leopold'), ('Relations institutionnelles', 'Lewis')]),
        ],
        'pendant': [
            ('Digital & visuel', [('Communication live / reseaux', 'Luther'), ('Photographe / videastes', 'A definir')]),
            ('Relations exterieures', [('Accueil presse sur site', 'Leopold'), ('Accueil institutionnels / partenaires', 'Lewis')]),
        ],
    },
    {
        'name': 'Production & Administration', 'chef': 'Charles', 'hex': '5B9E6F',
        'avant': [
            ('Finances', [('Budget global', 'Charles'), ('Banque (Qonto, TPE)', 'Charles')]),
            ('Juridique & contrats', [('Contrats & juridique (GUSO, assurances)', 'Charles'), ('Licences & droits (SACEM)', 'Charles')]),
            ('Financement', [('Subventions (DRAC, Region, Pass Culture)', 'Charles'), ('Partenariats prives & mecenat', 'Charles + Clotilde'), ('Financement participatif (HelloAsso)', 'Charles')]),
            ('Administration', [('Administration association', 'Charles'), ('Contact proprietaire / dossier arbres', 'Charles'), ('Presidence (representation legale)', 'Judith'), ('Support coordination', 'Wadih')]),
        ],
        'pendant': [
            ('Finances', [('Billetterie (vente, bracelets, controle)', 'Charles'), ('Caisse / tresorerie jour J', 'Charles')]),
            ('Administration', [('Relation proprietaire sur site', 'Charles'), ('Representation officielle (discours, VIP)', 'Judith'), ('Support coordination', 'Wadih')]),
        ],
    },
    {
        'name': 'Technique', 'chef': 'Leopold', 'hex': '4A8DB7',
        'avant': [
            ('Infrastructure', [('Implantation technique du site', 'Leopold'), ('Infrastructure & amenagement', 'Leopold'), ('Materiel technique', 'Leopold'), ('Toilettes seches (Trivial Compost)', 'Leopold')]),
            ('Scenes & plateau', [('Scene amplifiee', 'Leopold + Luther'), ('Scene acoustique', 'Leopold'), ('Plateau theatre', 'Lewis')]),
            ('Logistique artistes', [('Fiches techniques artistes (riders)', 'Leopold'), ('Logistique artistes (hebergement, transport)', 'Leopold'), ('Planning balances / repetitions', 'Leopold')]),
        ],
        'pendant': [
            ('Regie & scenes', [('Regie generale (changements, timing)', 'Leopold'), ('Son / lumiere scene amplifiee', 'Luther'), ('Plateau theatre', 'Lewis')]),
            ('Infrastructure & artistes', [('Accueil artistes backstage', 'Leopold'), ('Gestion imprevus techniques', 'Leopold'), ('Infrastructure jour J', 'Leopold')]),
        ],
    },
    {
        'name': 'Logistique Site', 'chef': 'Clotilde', 'hex': '9B6BB0',
        'avant': [
            ('Restauration & boissons', [('Restauration (menu, commandes)', 'Clotilde'), ('Buvette / Bar', 'Clotilde'), ('Approvisionnement (circuits courts)', 'Clotilde'), ('Stockage (refrigeration)', 'Clotilde')]),
            ('Ressources humaines', [('Benevoles (recrutement, planning)', 'Francine *')]),
            ('Securite & acces', [('Securite (dossier, mairie, pompiers)', 'Gaetan *'), ('Parking / Circulation', 'A definir')]),
            ('Montage', [('Montage / Demontage', 'Clotilde')]),
        ],
        'pendant': [
            ('Restauration & boissons', [('Service restauration', 'Clotilde'), ('Service bar / buvette', 'Clotilde')]),
            ('Ressources humaines', [('Coordination benevoles jour J', 'Francine *'), ('Entretien / proprete', 'Francine *')]),
            ('Securite & acces', [('Securite jour J', 'Gaetan *'), ('Parking / Circulation', 'A definir')]),
        ],
    },
]

wb = openpyxl.Workbook()

# ══════════ ONGLET 1: Par pole ══════════
ws1 = wb.active
ws1.title = 'Par pole'

thin_border = Border(
    bottom=Side(style='thin', color='E8E4DF')
)
header_font = Font(bold=True, color='FFFFFF', size=11)
sub_header_font = Font(bold=True, size=9, color='555555')
sub_header_fill = PatternFill(start_color='EDEBE8', end_color='EDEBE8', fill_type='solid')

row = 1
for pole in poles:
    fill = PatternFill(start_color=pole['hex'], end_color=pole['hex'], fill_type='solid')

    # Pole header
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws1.cell(row=row, column=1, value=f"{pole['name']}  —  Responsable: {pole['chef']}")
    cell.font = header_font
    cell.fill = fill
    cell.alignment = Alignment(horizontal='left')
    for c in range(1, 5):
        ws1.cell(row=row, column=c).fill = fill
    row += 1

    # Column headers
    for c, label in enumerate(['CATEGORIE', 'SOUS-POLE', 'RESPONSABLE', 'PHASE'], 1):
        cell = ws1.cell(row=row, column=c, value=label)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
    row += 1

    # Data rows
    for phase_name, phase_key in [('Avant', 'avant'), ('Pendant', 'pendant')]:
        for cat_name, items in pole[phase_key]:
            for task, resp in items:
                ws1.cell(row=row, column=1, value=cat_name)
                ws1.cell(row=row, column=2, value=task)
                ws1.cell(row=row, column=3, value=resp)
                ws1.cell(row=row, column=4, value=phase_name)
                for c in range(1, 5):
                    ws1.cell(row=row, column=c).border = thin_border
                row += 1

    row += 1  # blank separator

# Column widths
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 45
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 12

# ══════════ ONGLET 2: Par personne ══════════
ws2 = wb.create_sheet('Par personne')

# Build person index
person_data = {}
person_order = []
for pole in poles:
    for phase_name, phase_key in [('Avant', 'avant'), ('Pendant', 'pendant')]:
        for cat_name, items in pole[phase_key]:
            for task, resp in items:
                if resp == 'A definir':
                    continue
                names = [n.strip() for n in resp.replace(' *', '*').split('+')]
                for nm in names:
                    tent = nm.endswith('*')
                    clean = nm.rstrip('*').strip()
                    if clean not in person_data:
                        person_data[clean] = {'tent': tent, 'chef': [], 'tasks': []}
                        person_order.append(clean)
                    if not tent:
                        person_data[clean]['tent'] = False
                    person_data[clean]['tasks'].append((pole['name'], phase_name, cat_name, task))
    if pole['chef'] in person_data:
        person_data[pole['chef']]['chef'].append(pole['name'])

person_order.sort(key=lambda nm: (0 if person_data[nm]['chef'] else 1, 1 if person_data[nm]['tent'] else 0, nm))

# Header
dark_fill = PatternFill(start_color='2C2C2C', end_color='2C2C2C', fill_type='solid')
dark_font = Font(bold=True, color='FFFFFF', size=10)
headers = ['PRENOM', 'RESPONSABLE DE', 'POLE', 'PHASE', 'CATEGORIE', 'MISSION']
for c, label in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=c, value=label)
    cell.font = dark_font
    cell.fill = dark_fill

# Find pole color map
pole_colors = {p['name']: p['hex'] for p in poles}

row = 2
for nm in person_order:
    p = person_data[nm]
    chef_str = ', '.join(p['chef']) if p['chef'] else ''
    display = nm + (' (a confirmer)' if p['tent'] else '')

    # Person header row with light color
    if p['chef']:
        color_hex = pole_colors.get(p['chef'][0], 'EDEBE8')
    else:
        first_pole = p['tasks'][0][0] if p['tasks'] else ''
        color_hex = pole_colors.get(first_pole, 'EDEBE8')

    # Lighten the color for background
    r_val = int(color_hex[0:2], 16)
    g_val = int(color_hex[2:4], 16)
    b_val = int(color_hex[4:6], 16)
    light_r = min(255, r_val + (255 - r_val) * 7 // 10)
    light_g = min(255, g_val + (255 - g_val) * 7 // 10)
    light_b = min(255, b_val + (255 - b_val) * 7 // 10)
    light_hex = f'{light_r:02X}{light_g:02X}{light_b:02X}'
    person_fill = PatternFill(start_color=light_hex, end_color=light_hex, fill_type='solid')
    person_font = Font(bold=True, size=10)

    first = True
    for pole_name, phase, cat, task in p['tasks']:
        if first:
            ws2.cell(row=row, column=1, value=display).font = person_font
            ws2.cell(row=row, column=2, value=chef_str).font = Font(bold=True, color=color_hex)
            for c in range(1, 7):
                ws2.cell(row=row, column=c).fill = person_fill
            first = False
        else:
            ws2.cell(row=row, column=1, value='')

        ws2.cell(row=row, column=3, value=pole_name)
        ws2.cell(row=row, column=4, value=phase)
        ws2.cell(row=row, column=5, value=cat)
        ws2.cell(row=row, column=6, value=task)
        for c in range(1, 7):
            ws2.cell(row=row, column=c).border = thin_border
        row += 1
    row += 1  # blank separator

# Column widths
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 28
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 22
ws2.column_dimensions['F'].width = 45

# Freeze header
ws2.freeze_panes = 'A2'

wb.save(OUT_FILE)
print(f'Fichier xlsx cree: {OUT_FILE}')

# ══════════ Upload to Drive ══════════
with open(TOKEN_FILE) as f:
    token_data = json.load(f)

creds = Credentials(
    token=token_data['token'],
    refresh_token=token_data['refresh_token'],
    token_uri=token_data['token_uri'],
    client_id=token_data['client_id'],
    client_secret=token_data['client_secret'],
    scopes=token_data['scopes']
)

drive = build('drive', 'v3', credentials=creds)
FILE_NAME = 'Organisation_generale_Monique_Festival.xlsx'

# Check if exists
existing = drive.files().list(
    q=f"name = '{FILE_NAME}' and '{FOLDER_ID}' in parents and trashed = false",
    fields='files(id, name)'
).execute().get('files', [])

media = MediaFileUpload(OUT_FILE, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if existing:
    fid = existing[0]['id']
    drive.files().update(fileId=fid, media_body=media).execute()
    print(f'Mis a jour sur Drive: {FILE_NAME} (ID: {fid})')
else:
    metadata = {'name': FILE_NAME, 'parents': [FOLDER_ID]}
    created = drive.files().create(body=metadata, media_body=media, fields='id').execute()
    print(f'Cree sur Drive: {FILE_NAME} (ID: {created["id"]})')
