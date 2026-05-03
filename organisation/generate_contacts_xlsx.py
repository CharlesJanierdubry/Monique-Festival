"""Génère le fichier Excel des contacts membres JD Production."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

wb = Workbook()
ws = wb.active
ws.title = "Contacts JD Production"

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_font = Font(bold=True, size=12, color="4472C4")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Titre
ws.merge_cells("A1:F1")
ws["A1"] = "Contacts — Membres JD Production"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center")

headers = ["NOM", "PRÉNOM", "FONCTION", "ADRESSE", "EMAIL", "TÉLÉPHONE"]

bureau = [
    ["Laithier", "Judith", "Présidente", "81 rue de Verdun, 51100 Reims", "judithlaithier@gmail.com", "06 20 21 54 66"],
    ["Cormier", "Wadih", "Trésorier", "7 Grande Rue, 49123 Ingrandes-Le Fresne sur Loire", "wadih.cormier007@gmail.com", "07 50 47 51 75"],
]

membres = [
    ["Janier-Dubry", "Léopold", "Relations Publiques, Relations Artistes, Technique", "31 rue Charles Nodier, 25000 Besançon", "leopold.janier.dubry@gmail.com", "06 59 87 75 04"],
    ["Janier-Dubry", "Luther", "Communication & Marketing, Technique", "3 montée Bonafous, 69004 Lyon", "lututersan12@gmail.com", "07 60 78 55 69"],
    ["Janier-Dubry", "Lewis", "Direction Artistique, Restauration, Technique", "48 passage du Bureau, 75011 Paris", "lewis.janierdubry@gmail.com", "06 64 56 94 39"],
    ["Janier-Dubry", "Lorraine", "Communication, Direction Artistique, Financement Participatif", "Rue de Serbie 70, 1180 Ukkel, Belgique", "lorrainejanierdubry@gmail.com", "06 80 93 13 64"],
    ["Janier-Dubry", "Clotilde", "Relations Publiques, Restauration, Direction Artistique", "54 chemin de Valentin, 25000 Besançon", "clotilde@janier-dubry.fr", "06 87 79 34 82"],
    ["Janier-Dubry", "Charles", "Administratif & Logistique, Financement Participatif", "54 chemin de Valentin, 25000 Besançon", "charles@janier-dubry.fr", "07 87 43 87 85"],
]

row = 3

# Section Bureau
ws.merge_cells(f"A{row}:F{row}")
ws[f"A{row}"] = "Bureau"
ws[f"A{row}"].font = section_font
row += 1

# En-têtes Bureau
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
row += 1

for member in bureau:
    for col_idx, value in enumerate(member, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    row += 1

row += 1

# Section Membres
ws.merge_cells(f"A{row}:F{row}")
ws[f"A{row}"] = "Membres"
ws[f"A{row}"].font = section_font
row += 1

# En-têtes Membres
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
row += 1

for member in membres:
    for col_idx, value in enumerate(member, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    row += 1

row += 2

# Infos association
ws.merge_cells(f"A{row}:F{row}")
ws[f"A{row}"] = "Siège de l'association : 54 chemin de Valentin, 25000 Besançon"
ws[f"A{row}"].font = Font(bold=True, size=11)
row += 1
ws.merge_cells(f"A{row}:F{row}")
ws[f"A{row}"] = "Email de l'association : Info@monique-festival.fr"
ws[f"A{row}"].font = Font(bold=True, size=11)

# Largeurs de colonnes
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 45
ws.column_dimensions["E"].width = 35
ws.column_dimensions["F"].width = 18

output_path = os.path.join(os.path.dirname(__file__), "Contacts_membres_JD_Production.xlsx")
wb.save(output_path)
print(f"Fichier créé : {output_path}")
