"""Génère le fichier Excel de la matrice RACI Monique Festival.

⚠️ SORTIE = DOCUMENT INTERNE. Le fichier Excel produit contient le nom d'un·e
membre mineur·e (Léontine) : ne pas diffuser à l'extérieur (presse, mécènes,
institutions, site, réseaux sociaux).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

wb = Workbook()
ws = wb.active
ws.title = "Matrice RACI"

# Styles
header_font = Font(bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
pole_font = Font(bold=True, size=11, color="2F5496")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
r_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
a_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
c_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
i_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
r_font = Font(bold=True, size=10, color="FFFFFF")
a_font = Font(bold=True, size=10, color="FFFFFF")
c_font = Font(bold=True, size=10)
i_font = Font(size=10, color="808080")

def style_cell(cell, value):
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if value == "R":
        cell.fill = r_fill
        cell.font = r_font
    elif value == "A":
        cell.fill = a_fill
        cell.font = a_font
    elif value == "C":
        cell.fill = c_fill
        cell.font = c_font
    elif value == "I":
        cell.fill = i_fill
        cell.font = i_font
    elif value == "-":
        cell.font = Font(color="C0C0C0")

# Titre
ws.merge_cells("A1:L1")
ws["A1"] = "Matrice RACI — Monique Festival 2026"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A2:L2")
ws["A2"] = "R = Responsable (exécute) | A = Approbateur (valide) | C = Consulté | I = Informé"
ws["A2"].font = Font(size=9, italic=True)
ws["A2"].alignment = Alignment(horizontal="center")

# En-têtes
members = ["Pôle", "JU", "WA", "LÉO", "LUT", "LEW", "LOR", "CLO", "CHA", "LORE", "ROM", "LÉON"]
row = 4
for col_idx, header in enumerate(members, 1):
    cell = ws.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Noms complets en ligne 5
names = ["", "Judith", "Wadih", "Léopold", "Luther", "Lewis", "Lorraine", "Clotilde", "Charles", "Loretta", "Romane", "Léontine"]
row = 5
for col_idx, name in enumerate(names, 1):
    cell = ws.cell(row=row, column=col_idx, value=name)
    cell.font = Font(size=8, italic=True)
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Données RACI
data = [
    ["Communication & Marketing", "I", "I", "C", "R", "I", "A", "I", "I", "I", "I", "-"],
    ["Relations Publiques", "I", "I", "A", "I", "R", "C", "C", "I", "I", "I", "-"],
    ["Relations Artistes", "I", "I", "R", "I", "C", "A", "I", "I", "I", "I", "-"],
    ["Relations Bénévoles", "I", "I", "C", "I", "I", "C", "R", "I", "I", "I", "-"],
    ["Direction Artistique", "I", "I", "I", "I", "A", "C", "R", "I", "C", "I", "-"],
    ["Administratif & Logistique", "A", "C", "C", "I", "I", "I", "I", "R", "C", "I", "-"],
    ["Restauration", "I", "I", "I", "I", "R", "I", "A", "I", "I", "I", "-"],
    ["Technique", "I", "I", "R", "C", "A", "I", "I", "I", "I", "I", "-"],
    ["Engagement Éthique", "I", "I", "I", "I", "I", "C", "I", "I", "R", "A", "C"],
    ["Financement Participatif", "I", "C", "C", "C", "I", "A", "I", "R", "I", "I", "-"],
]

row = 6
for pole_data in data:
    cell = ws.cell(row=row, column=1, value=pole_data[0])
    cell.font = Font(bold=True, size=10)
    cell.border = thin_border
    cell.alignment = Alignment(vertical="center")
    for col_idx, val in enumerate(pole_data[1:], 2):
        style_cell(ws.cell(row=row, column=col_idx), val)
    row += 1

# Légende
row += 1
ws.merge_cells(f"A{row}:L{row}")
ws[f"A{row}"] = "Légende"
ws[f"A{row}"].font = Font(bold=True, size=11)
row += 1
legend = [
    ("R", "Responsable — Exécute la tâche", r_fill, r_font),
    ("A", "Approbateur — Valide et décide", a_fill, a_font),
    ("C", "Consulté — Donne son avis avant décision", c_fill, c_font),
    ("I", "Informé — Tenu au courant après décision", i_fill, i_font),
]
for code, desc, fill, font in legend:
    cell = ws.cell(row=row, column=1, value=code)
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
    ws.merge_cells(f"B{row}:L{row}")
    cell2 = ws.cell(row=row, column=2, value=desc)
    cell2.font = Font(size=10)
    cell2.border = thin_border
    row += 1

# Charge par membre
row += 1
ws.merge_cells(f"A{row}:L{row}")
ws[f"A{row}"] = "Charge par membre (nombre de R + A)"
ws[f"A{row}"].font = Font(bold=True, size=11)
row += 1

charge_headers = ["Membre", "R", "A", "Total"]
for col_idx, h in enumerate(charge_headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
row += 1

charge_data = [
    ("Léopold", 3, 2), ("Lewis", 2, 3), ("Clotilde", 2, 2),
    ("Lorraine", 0, 3), ("Charles", 2, 0), ("Luther", 1, 0),
    ("Loretta", 1, 0), ("Romane", 0, 1),
    ("Judith", 0, 3), ("Wadih", 0, 1),
]
for name, r_count, a_count in charge_data:
    ws.cell(row=row, column=1, value=name).border = thin_border
    ws.cell(row=row, column=2, value=r_count).border = thin_border
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3, value=a_count).border = thin_border
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value=r_count + a_count).border = thin_border
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4).font = Font(bold=True)
    row += 1

# Largeurs
ws.column_dimensions["A"].width = 30
for col in range(2, 13):
    ws.column_dimensions[get_column_letter(col)].width = 10

output = os.path.join(os.path.dirname(__file__), "Matrice_RACI_Monique_Festival.xlsx")
wb.save(output)
print(f"Fichier créé : {output}")
