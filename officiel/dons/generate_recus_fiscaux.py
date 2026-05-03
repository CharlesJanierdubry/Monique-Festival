"""
Génère les reçus fiscaux Cerfa 11580 pour les 14 dons antérieurs à l'activation Hello Asso.

Usage :
    python officiel/dons/generate_recus_fiscaux.py

Entrée :
    - Template : officiel/dons/templates/Template_recu_fiscal.md
    - Données donateurs : export-le-monique-festival-1-jd-production-20_04_2026-23_04_2026.xlsx
    - Config : paramètres association en haut de ce script (à mettre à jour post-récépissé)

Sortie :
    - officiel/dons/recus_emis/RF_{numero}_{nom_prenom}.md
    - officiel/dons/recus_emis/RF_{numero}_{nom_prenom}.docx (via pandoc)
    - officiel/dons/recus_emis/Registre_emission.csv (log)
"""
import sys
import csv
import subprocess
import re
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

# =============================================================================
# CONFIG — À METTRE À JOUR DÈS RÉCEPTION DU RÉCÉPISSÉ PRÉFECTORAL
# =============================================================================

ASSOCIATION = {
    # ANCIENNE DÉNOMINATION (à utiliser tant que le récépissé préfectoral n'est pas reçu)
    # "raison_sociale": "JD PRODUCTION",
    # "objet_ancien": "structurer la mise en place d'un festival annuel et encadrer toutes les actions liées à ce projet",

    # NOUVELLE DÉNOMINATION (à activer dès le récépissé — décommenter ci-dessous)
    "raison_sociale": "MONIQUE FESTIVAL",
    "objet": (
        "Organisation annuelle d'un festival pluridisciplinaire à Besançon et en Bourgogne-Franche-Comté, "
        "mêlant musique classique, musiques actuelles, théâtre et arts vivants ; soutien à la création "
        "artistique émergente ; actions d'éducation artistique et culturelle, de médiation et de transmission, "
        "notamment vers les publics éloignés de la pratique ; valorisation du territoire ; promotion de la "
        "diffusion de la culture, à titre non lucratif et d'intérêt général."
    ),
    "rna": "W251011013",
    "siren": "991 055 450",
    "siret": "991 055 450 00019",
    "adresse": "54 chemin de Valentin",
    "code_postal": "25000",
    "ville": "Besançon",
    "pays": "France",
    "signataire_nom": "Judith LAITHIER",
    "signataire_fonction": "Présidente",
    "ville_signature": "Besançon",
    # Chemin optionnel vers la signature PNG (fond transparent)
    # Si absent ou inexistant, laisse un emplacement vide pour signature manuelle
    "signature_png": "officiel/dons/signature_judith.png",
}

# Export Hello Asso des 14 dons à traiter
EXPORT_XLSX = "export-le-monique-festival-1-jd-production-20_04_2026-23_04_2026.xlsx"

# Date d'émission rétroactive (à adapter — date à laquelle on émet les reçus, pas la date du don)
# Recommandation : date post-récépissé préfectoral, quand on valide le lot
DATE_EMISSION = "15/05/2026"  # À ajuster

# Répertoire de sortie
OUTPUT_DIR = Path("officiel/dons/recus_emis")
TEMPLATE_PATH = Path("officiel/dons/templates/Template_recu_fiscal.md")
REGISTRE_CSV = OUTPUT_DIR / "Registre_emission.csv"

# Préfixe de numérotation pour les reçus rétroactifs manuels
# Format : DDMMAAAA-MF-000001 (cohérent avec le format Hello Asso auto)
NUMERO_PREFIXE = "MF"
NUMERO_DEBUT = 1


# =============================================================================
# Helpers
# =============================================================================

def nombre_en_lettres(n):
    """Convertit un entier en lettres françaises (simple, jusqu'à 9 999)."""
    units = ["", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf",
             "dix", "onze", "douze", "treize", "quatorze", "quinze", "seize",
             "dix-sept", "dix-huit", "dix-neuf"]
    tens = ["", "", "vingt", "trente", "quarante", "cinquante", "soixante",
            "soixante", "quatre-vingt", "quatre-vingt"]

    def below_100(n):
        if n < 20:
            return units[n]
        t = n // 10
        u = n % 10
        if t in (7, 9):
            base = tens[t]
            return base + ("-et-onze" if (t == 7 and u == 1) else f"-{units[10 + u]}" if u else f"-{units[10]}")
        base = tens[t]
        if u == 0:
            return base + ("s" if t == 8 else "")
        if u == 1 and t not in (8,):
            return base + "-et-un"
        return f"{base}-{units[u]}"

    if n == 0:
        return "zéro"
    if n < 100:
        return below_100(n)
    if n < 1000:
        c = n // 100
        r = n % 100
        prefix = "cent" if c == 1 else f"{units[c]} cents"
        if r:
            prefix = prefix.rstrip("s") + f" {below_100(r)}"
        return prefix
    if n < 10000:
        m = n // 1000
        r = n % 1000
        prefix = "mille" if m == 1 else f"{units[m]} mille"
        if r:
            # gérer "deux mille cinq" vs "deux mille cent"
            reste = ""
            if r < 100:
                reste = below_100(r)
            else:
                c = r // 100
                rr = r % 100
                cpart = "cent" if c == 1 else f"{units[c]} cents"
                if rr:
                    cpart = cpart.rstrip("s") + f" {below_100(rr)}"
                reste = cpart
            prefix += f" {reste}"
        return prefix
    return str(n)  # fallback


def read_export(path):
    """Lit l'export Hello Asso et retourne la liste des dons validés."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = None
    dons = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = list(row)
            continue
        if not any(cell is not None and str(cell).strip() != '' for cell in row):
            continue
        record = dict(zip(headers, row))
        if record.get("Statut de la commande") != "Validé":
            continue
        if record.get("Type") != "Don unique":
            continue
        dons.append(record)
    return dons


def slug(s):
    """Nettoie une chaine pour nom de fichier."""
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def render_template(template, values):
    """Substitution simple {{KEY}} → value."""
    out = template
    for k, v in values.items():
        out = out.replace("{{" + k + "}}", str(v) if v is not None else "")
    return out


def convert_to_docx(md_path, docx_path):
    """Convertit un .md en .docx via pandoc."""
    try:
        subprocess.run(
            ["pandoc", str(md_path), "-o", str(docx_path)],
            check=True, capture_output=True
        )
        return True
    except (subprocess.CalledProcessError, FileNotFoundError) as e:
        return False


# =============================================================================
# Main
# =============================================================================

def main():
    if not Path(EXPORT_XLSX).exists():
        sys.exit(f"ERREUR : export introuvable : {EXPORT_XLSX}")
    if not TEMPLATE_PATH.exists():
        sys.exit(f"ERREUR : template introuvable : {TEMPLATE_PATH}")

    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    dons = read_export(EXPORT_XLSX)
    print(f"{len(dons)} don(s) valide(s) trouve(s) dans l'export.\n")

    # Tri par date croissante (le plus ancien = n° 000001)
    dons.sort(key=lambda d: d.get("Date de la commande"))

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Registre d'émission
    registre_rows = [["Numéro reçu", "Date du don", "Donateur", "Email", "Montant", "Fichier MD", "Fichier DOCX"]]

    # Signature image si existe
    sig_path = Path(ASSOCIATION.get("signature_png", ""))
    if sig_path.exists():
        sig_md = f"![Signature]({sig_path.as_posix()})"
    else:
        sig_md = "\n\n\n_[signature manuscrite — à apposer après impression]_\n\n\n"

    date_ref = datetime.now().strftime("%d%m%Y")

    for i, don in enumerate(dons, start=NUMERO_DEBUT):
        numero = f"{date_ref}-{NUMERO_PREFIXE}-{i:06d}"
        date_don_raw = don.get("Date de la commande")
        if isinstance(date_don_raw, datetime):
            date_don = date_don_raw.strftime("%d/%m/%Y")
        else:
            date_don = str(date_don_raw).split(" ")[0]

        montant = int(don.get("Montant") or 0)
        prenom = (don.get("Prénom") or "").strip()
        nom = (don.get("Nom") or "").strip()
        email = don.get("Email") or ""

        # Article CGI : 200 pour particulier (pas de SIREN dans l'export = particulier)
        # 238 bis si entreprise (raison sociale + SIREN renseignes)
        if don.get("Raison sociale") or don.get("SIREN"):
            article = "238 bis"
            taux = 60
        else:
            article = "200"
            taux = 66

        values = {
            "ARTICLE_CGI": article,
            "TAUX_REDUCTION": str(taux),
            "RAISON_SOCIALE": ASSOCIATION["raison_sociale"],
            "RNA": ASSOCIATION["rna"],
            "SIREN": ASSOCIATION["siren"],
            "SIRET": ASSOCIATION["siret"],
            "ADRESSE": ASSOCIATION["adresse"],
            "CODE_POSTAL": ASSOCIATION["code_postal"],
            "VILLE": ASSOCIATION["ville"],
            "PAYS": ASSOCIATION["pays"],
            "OBJET": ASSOCIATION["objet"],
            "NUMERO_RECU": numero,
            "DONATEUR_PRENOM": prenom,
            "DONATEUR_NOM": nom,
            "DONATEUR_ADRESSE": don.get("Adresse") or "",
            "DONATEUR_CODE_POSTAL": str(don.get("Code postal") or ""),
            "DONATEUR_VILLE": don.get("Ville") or "",
            "DONATEUR_PAYS": don.get("Pays") or "",
            "MONTANT_CHIFFRES": str(montant),
            "MONTANT_LETTRES": nombre_en_lettres(montant) + " euros",
            "DATE_DON": date_don,
            "MOYEN_PAIEMENT": don.get("Moyen de paiement") or "",
            "VILLE_SIGNATURE": ASSOCIATION["ville_signature"],
            "DATE_EMISSION": DATE_EMISSION,
            "SIGNATAIRE_NOM": ASSOCIATION["signataire_nom"],
            "SIGNATAIRE_FONCTION": ASSOCIATION["signataire_fonction"],
            "SIGNATURE_IMAGE_OR_BLANK": sig_md,
        }

        md_content = render_template(template, values)

        filename_base = f"RF_{numero}_{slug(prenom)}_{slug(nom)}"
        md_path = OUTPUT_DIR / f"{filename_base}.md"
        docx_path = OUTPUT_DIR / f"{filename_base}.docx"

        md_path.write_text(md_content, encoding="utf-8")
        docx_ok = convert_to_docx(md_path, docx_path)

        status = "OK" if docx_ok else "MD seul (pandoc absent)"
        print(f"  {numero} | {prenom} {nom} | {montant} EUR | article {article} | {status}")

        registre_rows.append([
            numero, date_don, f"{prenom} {nom}", email, f"{montant} EUR",
            md_path.name, docx_path.name if docx_ok else "",
        ])

    # Sauvegarde du registre
    with REGISTRE_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(registre_rows)

    print(f"\n{len(dons)} recus generes dans : {OUTPUT_DIR}")
    print(f"Registre d'emission : {REGISTRE_CSV}")


if __name__ == "__main__":
    main()
