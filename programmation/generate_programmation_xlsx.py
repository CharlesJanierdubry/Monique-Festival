"""
Reconstruit le fichier programmation/Programmation_Monique_Festival.xlsx
à partir des sources officielles Drive (Planning des scènes + Check-list artistes)
mises à jour au 28/04/2026.

4 onglets :
- Artistes : tous les intervenants individuels avec contacts
- Compagnies : structures porteuses
- Programmation : planning consolidé des spectacles avec dates corrigées (août 2026)
- Logistique : grille des balances/accessibilité/changements de plateau

Usage : python programmation/generate_programmation_xlsx.py
"""
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "programmation" / "Programmation_Monique_Festival.xlsx"

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
DAY_FILL = PatternFill("solid", fgColor="D9E1F2")
DAY_FONT = Font(bold=True, size=12)
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
TOTAL_FONT = Font(bold=True)
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row_num, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, max_width=50):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# ============================================
# ONGLET 1 : ARTISTES
# ============================================
ARTISTES_HEADERS = ["Statut", "Nom", "Prénom", "Nom de scène / Groupe", "Catégorie",
                    "Référent groupe", "Téléphone", "E-mail", "Mode rémunération",
                    "Famille", "Charte éthique", "Notes"]

ARTISTES = [
    # MUSIQUES ACTUELLES
    "MUSIQUES ACTUELLES",
    ["Complet", "Cabaret", "Romane", "Duo Cabaret", "Musique", "Romane Cabaret", "+32 488 36 17 05", "romanecabaret@gmail.com", "Cachet individuel", "", "À signer", "Aussi Fabulo (Masterclass + Nos Voix)"],
    ["Incomplet", "Cabaret", "Georges", "Duo Cabaret", "Musique", "Romane Cabaret", "", "romanecabaret@gmail.com", "Cachet individuel", "", "À signer", ""],
    ["Complet", "", "Fanélie", "Fanélie Nava", "Musique", "Fanélie", "", "fanelienava@gmail.com", "Cachet individuel", "", "À signer", "Afro-pop / RnB"],
    ["Complet", "Janier-Dubry", "Luther", "Luther et Loretta", "Musique", "Luther", "07 60 78 55 69", "lutheretlorettamusic@gmail.com", "Famille", "Oui", "À signer", "Chanson / Psytrance"],
    ["Incomplet", "", "Loretta", "Luther et Loretta", "Musique", "Luther", "", "lutheretlorettamusic@gmail.com", "Cachet collectif", "", "À signer", ""],
    ["Complet", "", "Laszlo", "Wet Enough?!", "Musique", "Laszlo", "07 67 52 22 69", "contact.wetenough@gmail.com", "Facture collectif", "", "À signer", "Référent — Groove / Funk"],
    ["Incomplet", "", "Marius", "Wet Enough?!", "Musique", "Laszlo", "", "", "Facture collectif", "", "À signer", ""],
    ["Incomplet", "", "Maël", "Wet Enough?!", "Musique", "Laszlo", "", "", "Facture collectif", "", "À signer", "Aussi Osmosis"],
    ["Incomplet", "", "Baptiste", "Wet Enough?!", "Musique", "Laszlo", "", "", "Facture collectif", "", "À signer", "Aussi Osmosis"],
    ["Incomplet", "", "Matthieu", "Wet Enough?!", "Musique", "Laszlo", "", "", "Facture collectif", "", "À signer", ""],
    ["Complet", "", "Manu", "Wambo", "Musique", "Manu", "06 17 50 77 16", "contact@wamboproductions.com", "Cachet individuel", "", "À signer", "Chanson française métissée"],
    ["Complet", "Shiffman", "Martin", "Vladimir Torres Trio", "Musique", "Martin Shiffman (Constantia Prod)", "06 76 33 87 41", "contact@constantiaprod.com", "Facture collectif", "", "À signer", "Référent — Jazz, 3 musiciens"],
    ["Incomplet", "", "Membre 2", "Vladimir Torres Trio", "Musique", "Martin Shiffman", "", "contact@constantiaprod.com", "Facture collectif", "", "À signer", ""],
    ["Incomplet", "", "Membre 3", "Vladimir Torres Trio", "Musique", "Martin Shiffman", "", "contact@constantiaprod.com", "Facture collectif", "", "À signer", ""],
    ["Complet", "", "Maël", "Osmosis", "Musique", "Maël", "06 27 44 67 51", "osmosislab.contact@gmail.com", "Facture collectif", "", "À signer", "Référent — Breakbeat / Drum'n'Bass"],
    ["Incomplet", "", "Paul", "Osmosis", "Musique", "Maël", "", "", "Facture collectif", "", "À signer", ""],
    ["Incomplet", "", "Baptiste", "Osmosis", "Musique", "Maël", "", "", "Facture collectif", "", "À signer", "Aussi Wet Enough?!"],
    ["Complet", "Zeiger", "Carina", "Kaïnza", "Musique", "Carina Zeiger", "07 89 54 07 71", "carina.zeiger@hotmail.de", "Facture collectif", "", "À signer", "Référente — Grooves oriental, 3 pers (ex Karina Ziegler)"],
    ["Incomplet", "", "Membre 2", "Kaïnza", "Musique", "Carina Zeiger", "", "", "Facture collectif", "", "À signer", ""],
    ["Incomplet", "", "Membre 3", "Kaïnza", "Musique", "Carina Zeiger", "", "", "Facture collectif", "", "À signer", ""],
    # CLASSIQUE
    "CLASSIQUE",
    ["Complet", "Janier-Dubry", "Lorraine", "Europe Cellists / Fabulo / Nos Voix", "Classique", "Lorraine", "06 80 93 13 64", "lorrainejanierdubry@gmail.com", "Famille", "Oui", "À signer", "Référente Fabulo + violoncelliste"],
    ["Complet", "Trembovelski", "Marc", "Europe Cellists", "Classique", "Lorraine", "+44 73 94 19 39 21", "m.trembovelski@gmail.com", "Cachet individuel", "", "À signer", "Violoncelliste — Londres UK"],
    ["Complet", "Wycislik", "Jakub", "Europe Cellists", "Classique", "Lorraine", "+48 6 04 07 60 51", "jakub.wycislik1@gmail.com", "Cachet individuel", "", "À signer", "Violoncelliste — Pologne"],
    ["Complet", "Ehling", "Sophie", "Europe Cellists", "Classique", "Lorraine", "+31 6 13 93 39 39", "sophie.ehling@gmail.com", "Cachet individuel", "", "À signer", "Violoncelliste"],
    ["Complet", "Correia", "Beatriz", "Europe Cellists", "Classique", "Lorraine", "+351 9 15 85 56 33", "beamucaco@gmail.com", "Cachet individuel", "", "À signer", "Violoncelliste — Portugal"],
    ["Complet", "Lemonnier", "Augustin", "Augustin Lemonnier", "Classique", "Augustin", "07 78 88 12 73", "augustinlemonnier@yahoo.fr", "Cachet individuel", "", "À signer", "Pianiste — Europe Cellists meet Augustin (sam)"],
    ["Complet", "Van Keulen", "Volodia", "Volodia Van Keulen", "Classique", "Volodia", "06 68 04 04 61", "vankeulenvolodia@hotmail.fr", "Cachet individuel", "", "À signer", "Sithar & Violoncelle — aussi Suites de Bach (dim)"],
    ["Incomplet", "Van Keulen", "Nils", "Nils Van Keulen", "Classique", "Nils", "", "", "Contrat Fabulo", "", "À signer", "Accompagnement Masterclass samedi"],
    ["Complet", "Willems", "Lisa", "À travers la fenêtre des heures", "Classique", "Lisa Willems", "+32 4 78 09 38 27", "willemslisa1@gmail.com", "Facture collectif", "", "À signer", "Référente — Musique de chambre, 2 pers"],
    ["Incomplet", "", "Membre 2", "À travers la fenêtre des heures", "Classique", "Lisa Willems", "", "", "Facture collectif", "", "À signer", ""],
    # FABULO / ACTION CULTURELLE
    "FABULO / ACTION CULTURELLE",
    ["Incomplet", "", "Marius", "Fabulo", "Fabulo", "Lorraine", "", "collectif.fabulo@gmail.com", "Facture collectif", "", "À signer", "Atelier Fabulo"],
    ["Incomplet", "", "Maya", "Fabulo", "Fabulo", "Lorraine", "", "collectif.fabulo@gmail.com", "Facture collectif", "", "À signer", "Atelier Fabulo"],
    ["Incomplet", "", "Raphaëlle", "Fabulo", "Fabulo", "Lorraine", "", "collectif.fabulo@gmail.com", "Facture collectif", "", "À signer", "Atelier Fabulo"],
    ["Complet", "Vermot", "Antoine", "Jeff The Fool / Fabulo", "Fabulo", "Antoine Vermot", "06 78 14 47 10", "jeffthefoolmusic@gmail.com", "Contrat Fabulo", "", "À signer", "Ateliers culturels (sam+dim) + Atelier DJ"],
    ["Complet", "Janier-Dubry", "Léopold", "Léopold", "Musique", "Léopold", "06 59 87 75 04", "leopold.janier.dubry@gmail.com", "Famille", "Oui", "À signer", "Atelier DJ avec Antoine Vermot"],
    # THÉÂTRE
    "THÉÂTRE",
    ["Complet", "Ponnelle", "Marius", "Cie Treize Clique", "Théâtre", "Marius Ponnelle", "06 50 56 79 85", "mariusponnelle@gmail.com", "Facture collectif", "", "À signer", "Référent — Vendredi 16h45"],
    ["Incomplet", "", "Comédien 2", "Cie Treize Clique", "Théâtre", "Marius Ponnelle", "", "mariusponnelle@gmail.com", "Facture collectif", "", "À signer", ""],
    ["Incomplet", "", "Comédien 3", "Cie Treize Clique", "Théâtre", "Marius Ponnelle", "", "mariusponnelle@gmail.com", "Facture collectif", "", "À signer", ""],
    ["Complet", "Delafolie", "Claire", "Cie Nenni ma foi (Patatra)", "Théâtre", "Claire Delafolie", "06 95 30 17 48", "assonennimafoi@gmail.com", "Facture collectif", "", "À signer", "Cirque & Musique — Samedi 14h, déambulatoire"],
    ["Complet", "Jeanmougin", "Juliette", "Cie L'Entre (V.I.E.N.S)", "Théâtre", "Juliette Jeanmougin", "06 49 09 37 54", "compagnielentre@gmail.com", "Facture collectif", "", "À signer", "Référente — Pelousey 25170 (BFC) — Dimanche 19h30, 3 art"],
    ["Complet", "Creuze", "Matthieu", "Cie Derrière le mur (Trafic)", "Théâtre", "Matthieu Creuze", "06 78 98 74 66", "derrierelemurcompagnie@gmail.com", "Facture collectif", "", "À signer", "Théâtre -14 ans — Dimanche 15h45, 2 pers"],
    ["Complet", "Mimoun", "Hannah", "Cie Vous faites un feu?", "Théâtre", "Hannah Mimoun", "07 83 79 53 20", "mimoun.hannah78@gmail.com", "Facture collectif", "", "À signer", "Tu ne pouvais pas être bien... — Samedi 18h15, 2 pers"],
    ["Complet", "Janier-Dubry", "Lewis", "Lewis Janier-Dubry / Cie du 6ème mur", "Théâtre", "Lewis", "06 64 56 94 39", "sixieme.mur@gmail.com", "Famille", "Oui", "À signer", "Atelier éloquence (sam+dim)"],
    ["Incomplet", "Septours", "Sylvain", "Sylvain Septours", "Théâtre", "Sylvain Septours", "", "sixieme.mur@gmail.com", "Cachet individuel", "", "À signer", "Atelier écriture (sam+dim)"],
]


def write_artistes(wb):
    ws = wb.create_sheet("Artistes")
    ws["A1"] = "LISTE DES ARTISTES — Monique Festival 2026"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:L1")

    row = 3
    # En-têtes
    for c, h in enumerate(ARTISTES_HEADERS, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(ARTISTES_HEADERS))
    row += 1

    for item in ARTISTES:
        if isinstance(item, str):
            cell = ws.cell(row=row, column=1, value=item)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            row += 1
        else:
            for c, v in enumerate(item, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = BORDER
                if c == 10 and v == "Oui":
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
            row += 1

    autosize(ws, max_width=35)
    ws.row_dimensions[1].height = 22


# ============================================
# ONGLET 2 : COMPAGNIES
# ============================================
COMPAGNIES_HEADERS = ["Nom de scène", "Type", "Référent", "Téléphone", "E-mail",
                       "Catégorie", "Nb artistes", "Charte éthique", "Notes"]

COMPAGNIES = [
    ["Duo Cabaret", "Duo", "Romane Cabaret", "+32 488 36 17 05", "romanecabaret@gmail.com", "Musique", 2, "À signer", "Chanson cabaret — Vendredi 15h"],
    ["Fanélie Nava", "Solo", "Fanélie", "", "fanelienava@gmail.com", "Musique", 1, "À signer", "Afro-pop/RnB — Vendredi 18h45"],
    ["Luther et Loretta", "Duo", "Luther Janier-Dubry", "07 60 78 55 69", "lutheretlorettamusic@gmail.com", "Musique", 2, "À signer", "Chanson/Psytrance — Vendredi 19h45"],
    ["Wet Enough?!", "Groupe", "Laszlo", "07 67 52 22 69", "contact.wetenough@gmail.com", "Musique", 5, "À signer", "Groove/Funk — Vendredi 21h"],
    ["Wambo", "Solo", "Manu", "06 17 50 77 16", "contact@wamboproductions.com", "Musique", 1, "À signer", "Chanson française métissée — Samedi 12h45"],
    ["Vladimir Torres Trio", "Trio", "Martin Shiffman / Constantia Prod", "06 76 33 87 41", "contact@constantiaprod.com", "Musique", 3, "À signer", "Jazz — Samedi 17h"],
    ["Osmosis", "Groupe", "Maël", "06 27 44 67 51", "osmosislab.contact@gmail.com", "Musique", 3, "À signer", "Breakbeat/Drum'n'Bass — Samedi 21h"],
    ["Kaïnza", "Ensemble", "Carina Zeiger", "07 89 54 07 71", "carina.zeiger@hotmail.de", "Musique", 3, "À signer", "Grooves oriental — Dimanche 18h15 (ex Karina Ziegler)"],
    ["Europe Cellists", "Ensemble", "Lorraine / Marc Trembovelski", "+44 73 94 19 39 21", "m.trembovelski@gmail.com", "Classique", 5, "À signer", "Violoncelles + Augustin pianiste — Sam+Dim"],
    ["Volodia Van Keulen", "Solo", "Volodia Van Keulen", "06 68 04 04 61", "vankeulenvolodia@hotmail.fr", "Classique", 1, "À signer", "Sithar & Violoncelle — Dimanche 14h30"],
    ["À travers la fenêtre des heures", "Duo", "Lisa Willems", "+32 4 78 09 38 27", "willemslisa1@gmail.com", "Classique", 2, "À signer", "Musique de chambre — Samedi 15h15"],
    ["Fabulo", "Collectif / Association", "Lorraine Janier-Dubry", "06 80 93 13 64", "collectif.fabulo@gmail.com", "Classique / Création", 5, "À signer", "Concert + Ateliers + Création Monique"],
    ["Cie Treize Clique", "Compagnie théâtrale", "Marius Ponnelle", "06 50 56 79 85", "mariusponnelle@gmail.com", "Théâtre", 3, "À signer", "Vendredi 16h45"],
    ["Cie Nenni ma foi", "Compagnie", "Claire Delafolie", "06 95 30 17 48", "assonennimafoi@gmail.com", "Théâtre/Cirque", 4, "À signer", "Patatra — Samedi 14h, déambulatoire"],
    ["Cie L'Entre", "Compagnie", "Juliette Jeanmougin", "06 49 09 37 54", "compagnielentre@gmail.com", "Théâtre", 3, "À signer", "V.I.E.N.S (Vers Ici Entre Nos rienS) — Pelousey 25170 (BFC) — Dimanche 19h30"],
    ["Cie Derrière le mur", "Compagnie", "Matthieu Creuze", "06 78 98 74 66", "derrierelemurcompagnie@gmail.com", "Théâtre", 2, "À signer", "Trafic (-14 ans) — Dimanche 15h45"],
    ["Cie Vous faites un feu?", "Compagnie", "Hannah Mimoun", "07 83 79 53 20", "mimoun.hannah78@gmail.com", "Théâtre", 2, "À signer", "Tu ne pouvais pas être bien... — Samedi 18h15"],
    ["Cie du 6ème mur", "Compagnie", "Lewis Janier-Dubry", "06 64 56 94 39", "sixieme.mur@gmail.com", "Théâtre", 2, "À signer", "Atelier éloquence (Lewis) + écriture (Sylvain Septours)"],
]


def write_compagnies(wb):
    ws = wb.create_sheet("Compagnies")
    ws["A1"] = "COMPAGNIES & COLLECTIFS — Monique Festival 2026"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")

    row = 3
    for c, h in enumerate(COMPAGNIES_HEADERS, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(COMPAGNIES_HEADERS))
    row += 1

    for item in COMPAGNIES:
        for c, v in enumerate(item, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = BORDER
        row += 1

    autosize(ws, max_width=40)
    ws.row_dimensions[1].height = 22


# ============================================
# ONGLET 3 : PROGRAMMATION
# ============================================
PROGRAMMATION_HEADERS = ["Date", "Heure début", "Heure fin", "Type", "Spectacle",
                         "Compagnie / Référent", "Scène", "Genre", "Nb artistes", "Notes"]

PROGRAMMATION = [
    "VENDREDI 28 AOÛT 2026",
    ["28/08/2026", "15:00", "15:30", "Spectacle", "Duo Cabaret", "Romane & Georges Cabaret", "Amplifiée", "Chanson cabaret", 2, ""],
    ["28/08/2026", "15:45", "16:30", "Spectacle", "Concert Fabulo", "Fabulo (Lorraine)", "Acoustique", "Musique de chambre", 5, "Lorraine = famille"],
    ["28/08/2026", "16:45", "17:30", "Spectacle", "Comment on en est arrivés là ?", "Cie Treize Clique (Marius Ponnelle)", "Acoustique", "Théâtre", 3, ""],
    ["28/08/2026", "18:45", "19:30", "Spectacle", "Fanélie", "Fanélie Nava", "Amplifiée", "Afro-pop / RnB", 1, ""],
    ["28/08/2026", "19:45", "20:45", "Spectacle", "Luther & Loretta", "Luther et Loretta", "Amplifiée", "Chanson / Psytrance", 2, "Luther = famille"],
    ["28/08/2026", "21:00", "22:00", "Spectacle", "Wet Enough?!", "Wet Enough (Laszlo)", "Amplifiée", "Groove / Funk", 5, ""],
    "SAMEDI 29 AOÛT 2026",
    ["29/08/2026", "10:00", "12:00", "Atelier", "Masterclass publique chant lyrique", "Romane Cabaret + Nils Van Keulen", "Acoustique", "Action culturelle", 2, ""],
    ["29/08/2026", "10:00", "12:00", "Atelier", "Atelier Fabulo + ateliers culturels", "Fabulo (Lorraine + Antoine Vermot)", "Amplifiée", "Action culturelle", 5, "Répétitions (4 payées)"],
    ["29/08/2026", "10:00", "12:00", "Atelier", "Atelier éloquence", "Lewis Janier-Dubry / Cie du 6ème mur", "—", "Action culturelle", 1, "Famille — atelier"],
    ["29/08/2026", "10:00", "12:00", "Atelier", "Atelier écriture", "Sylvain Septours", "—", "Action culturelle", 1, ""],
    ["29/08/2026", "10:00", "12:00", "Atelier", "Atelier DJ / sound design", "Antoine Vermot + Léopold Janier-Dubry", "—", "Action culturelle", 2, "Léopold = famille"],
    ["29/08/2026", "12:30", "13:30", "Spectacle", "Wambo", "Wambo (Manu)", "Amplifiée", "Chanson française métissée", 1, ""],
    ["29/08/2026", "14:00", "15:00", "Spectacle", "Patatra", "Cie Nenni ma foi (Claire Delafolie)", "Déambulatoire", "Cirque & Musique", 4, ""],
    ["29/08/2026", "15:15", "16:30", "Spectacle", "À travers la fenêtre des heures", "Lisa Willems & co", "Acoustique", "Musique de chambre", 2, ""],
    ["29/08/2026", "17:00", "18:00", "Spectacle", "Vladimir Torres Trio", "Constantia Prod (Martin Shiffman)", "Amplifiée", "Jazz", 3, ""],
    ["29/08/2026", "18:15", "19:15", "Spectacle", "Tu ne pouvais pas être bien tu étais dan...", "Cie Vous faites un feu? (Hannah Mimoun)", "Acoustique", "Théâtre", 2, ""],
    ["29/08/2026", "19:45", "21:00", "Spectacle", "Europe Cellists meet Augustin", "Europe Cellists + Augustin Lemonnier", "Acoustique", "Musique de chambre", 5, "Lorraine = famille"],
    ["29/08/2026", "21:00", "22:00", "Spectacle", "Osmosis", "Osmosis (Maël)", "Amplifiée", "Breakbeat / Drum'n'Bass", 3, ""],
    "DIMANCHE 30 AOÛT 2026",
    ["30/08/2026", "10:00", "12:00", "Atelier", "Masterclass & ateliers culturels (générale)", "Fabulo (Lorraine + Antoine Vermot)", "Acoustique", "Action culturelle", 5, ""],
    ["30/08/2026", "10:00", "12:00", "Atelier", "Atelier éloquence", "Lewis Janier-Dubry", "—", "Action culturelle", 1, "Famille"],
    ["30/08/2026", "10:00", "12:00", "Atelier", "Atelier écriture", "Sylvain Septours", "—", "Action culturelle", 1, ""],
    ["30/08/2026", "10:00", "12:00", "Atelier", "Atelier DJ", "Antoine Vermot + Léopold", "—", "Action culturelle", 2, "Léopold = famille"],
    ["30/08/2026", "12:45", "14:45", "Spectacle", "Suites de Bach", "Europe Cellists (5) + Volodia + Lorraine", "Acoustique", "Musique de chambre", 6, "Lorraine = famille"],
    ["30/08/2026", "14:30", "15:30", "Spectacle", "Sithar & Violoncelle", "Volodia Van Keulen", "Acoustique", "Performance", 1, ""],
    ["30/08/2026", "15:45", "17:00", "Spectacle", "Trafic", "Cie Derrière le mur (Matthieu Creuze)", "Acoustique", "Théâtre (-14 ans)", 2, ""],
    ["30/08/2026", "17:15", "18:00", "Spectacle", "Nos voix", "Fabulo (Romane + Lorraine)", "Amplifiée", "Chanson", 2, "Lorraine = famille"],
    ["30/08/2026", "18:15", "19:15", "Spectacle", "Kaïnza", "Kaïnza (Carina Zeiger)", "Amplifiée", "Grooves oriental", 3, ""],
    ["30/08/2026", "19:30", "20:00", "Spectacle", "V.I.E.N.S (Vers Ici Entre Nos rienS)", "Cie L'Entre (Juliette Jeanmougin)", "Acoustique", "Théâtre", 3, ""],
    ["30/08/2026", "20:30", "21:00", "Spectacle", "Création Monique", "Fabulo (Lorraine et participants)", "Acoustique", "Restitution ateliers", 5, "Lorraine = famille"],
]


def write_programmation(wb):
    ws = wb.create_sheet("Programmation")
    ws["A1"] = "PROGRAMMATION — Monique Festival 2026 (28-30 août 2026)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")

    row = 3
    for c, h in enumerate(PROGRAMMATION_HEADERS, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(PROGRAMMATION_HEADERS))
    row += 1

    spectacles = ateliers = artistes_total = 0

    for item in PROGRAMMATION:
        if isinstance(item, str):
            cell = ws.cell(row=row, column=1, value=item)
            cell.font = DAY_FONT
            cell.fill = DAY_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            row += 1
            continue
        for c, v in enumerate(item, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = BORDER
        if item[3] == "Spectacle":
            spectacles += 1
        elif item[3] == "Atelier":
            ateliers += 1
        artistes_total += item[8]
        row += 1

    # Totaux
    ws.cell(row=row, column=1, value="TOTAUX")
    ws.cell(row=row, column=4, value=f"{spectacles} spectacles + {ateliers} ateliers")
    ws.cell(row=row, column=9, value=artistes_total)
    for c in range(1, 11):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER

    autosize(ws, max_width=45)
    ws.row_dimensions[1].height = 22


# ============================================
# ONGLET 4 : LOGISTIQUE
# ============================================
LOGISTIQUE_VEN = [
    ("09:15", "BALANCES WET ENOUGH?!", ""),
    ("10:15", "", "INSTALLATION TREIZE CLIQUE"),
    ("10:30", "BALANCES LUTHER & LORETTA", ""),
    ("11:15", "BALANCES FANÉLIE", ""),
    ("12:45", "", "BALANCES FABULO"),
    ("13:00", "ACCESSIBILITÉ FANÉLIE", ""),
    ("13:45", "", "ACCESSIBILITÉ FABULO"),
    ("14:00", "BALANCES DUO CABARET", ""),
    ("14:15", "", "ACCESSIBILITÉ TREIZE CLIQUE"),
    ("14:45", "OUVERTURE SITE + BUVETTE / RESTAURATION", ""),
    ("15:00", "DUO CABARET", ""),
    ("15:45", "CHANGEMENT PLATEAU", "FABULO"),
    ("16:30", "", "CHANGEMENT PLATEAU"),
    ("16:45", "", "TREIZE CLIQUE"),
    ("17:30", "", "RANGEMENT SCÈNE"),
    ("17:45", "ACCESSIBILITÉ LUTHER & LORETTA", ""),
    ("18:15", "ACCESSIBILITÉ WET ENOUGH?!", ""),
    ("18:30", "FANÉLIE", ""),
    ("19:30", "CHANGEMENT PLATEAU", ""),
    ("19:45", "LUTHER & LORETTA", ""),
    ("20:45", "CHANGEMENT PLATEAU", ""),
    ("21:00", "WET ENOUGH ?!", ""),
    ("22:00", "RANGEMENT SCÈNE", ""),
    ("22:30", "FERMETURE BUVETTE / RESTAURATION", ""),
    ("22:45", "FERMETURE SITE", ""),
]
LOGISTIQUE_SAM = [
    ("09:00", "BALANCES WAMBO", "BALANCES À TRAVERS LA FENÊTRE DES HEURES"),
    ("09:30", "", "BALANCES MASTERCLASS"),
    ("09:45", "OUVERTURE SITE + BUVETTE / RESTAURATION", ""),
    ("10:00", "INSTALLATION PATATRA", "MASTERCLASS PUBLIQUE"),
    ("12:00", "ACCESSIBILITÉ WAMBO", "BALANCES EUROPE CELLIST MEET AUGUSTIN"),
    ("12:15", "ACCESSIBILITÉ PATATRA", ""),
    ("12:30", "WAMBO", "TEMPS TECHNIQUE V.I.E.N.S"),
    ("13:30", "BALANCES OSMOSIS", ""),
    ("14:00", "PATATRA", ""),
    ("14:30", "", "ACCESSIBILITÉ VLADIMIR TORRES TRIO"),
    ("15:00", "DESINSTALLATION PATATRA", "ACCESSIBILITÉ ENTRE LA FENÊTRE DES HEURES"),
    ("15:15", "", "À TRAVERS LA FENÊTRE DES HEURES"),
    ("16:30", "BALANCES VLADIMIR TORRES TRIO", "CHANGEMENT PLATEAU"),
    ("16:45", "", "ACCESSIBILITÉ VOUS FAITES UN FEU?"),
    ("17:00", "VLADIMIR TORRES TRIO", ""),
    ("17:15", "", "INSTALLATION V.I.E.N.S"),
    ("18:00", "CHANGEMENT PLATEAU", "ACCESSIBILITÉ EUROPE CELLIST MEET AUGUSTIN"),
    ("18:15", "", "VOUS FAITES UN FEU ?"),
    ("19:15", "", "CHANGEMENT PLATEAU"),
    ("19:30", "ACCESSIBILITÉ OSMOSIS", ""),
    ("19:45", "", "EUROPE CELLIST MEET AUGUSTIN"),
    ("21:00", "OSMOSIS", "RANGEMENT SCÈNE"),
    ("22:00", "RANGEMENT SCÈNE", ""),
    ("22:30", "FERMETURE BUVETTE / RESTAURATION", ""),
    ("22:45", "FERMETURE SITE", ""),
]
LOGISTIQUE_DIM = [
    ("09:00", "", "BALANCES TRAFIC"),
    ("09:15", "BALANCES KAÏNZA", ""),
    ("09:30", "", "BALANCES VOUS FAITES UN FEU ?"),
    ("10:00", "", "MASTERCLASS & ATELIERS"),
    ("10:15", "ACCESSIBILITÉ KAÏNZA", ""),
    ("10:45", "ACCESSIBILITÉ NOS VOIX", ""),
    ("11:15", "BALANCES CRÉATION MONIQUE", "ACCESSIBILITÉ EUROPE CELLIST + VOLODIA"),
    ("11:45", "", "ACCESSIBILITÉ TRAFIC"),
    ("12:00", "OUVERTURE SITE + BUVETTE / RESTAURATION", ""),
    ("12:15", "", "BALANCES EUROPE CELLIST + VOLODIA"),
    ("12:45", "", "EUROPE CELLIST : SUITES DE BACH"),
    ("14:45", "", "VOLODIA VAN KEULEN"),
    ("15:30", "", "CHANGEMENT PLATEAU"),
    ("15:45", "", "CIE DERRIÈRE LE MUR — TRAFIC"),
    ("17:00", "", "CHANGEMENT PLATEAU"),
    ("17:15", "NOS VOIX", ""),
    ("18:00", "CHANGEMENT PLATEAU", "ACCESSIBILITÉ V.I.E.N.S."),
    ("18:15", "KAÏNZA", ""),
    ("19:15", "RANGEMENT SCÈNE", ""),
    ("19:30", "", "V.I.E.N.S"),
    ("20:15", "", "CHANGEMENT PLATEAU"),
    ("20:30", "", "CRÉATION MONIQUE"),
    ("21:00", "FERMETURE BUVETTE / RESTAURATION", ""),
    ("21:15", "FERMETURE SITE", ""),
    ("21:45", "", "RANGEMENT SCÈNE"),
]


def write_logistique(wb):
    ws = wb.create_sheet("Logistique")
    ws["A1"] = "LOGISTIQUE — Planning des scènes (balances / accessibilité / changements)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    ws["A2"] = "Source : Google Sheet « Planning des scènes » au 28/04/2026"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:C2")

    row = 4
    headers = ["Heure", "Scène Amplifiée", "Scène Acoustique"]

    for jour, donnees in [("VENDREDI 28/08/2026", LOGISTIQUE_VEN),
                           ("SAMEDI 29/08/2026", LOGISTIQUE_SAM),
                           ("DIMANCHE 30/08/2026", LOGISTIQUE_DIM)]:
        cell = ws.cell(row=row, column=1, value=jour)
        cell.font = DAY_FONT
        cell.fill = DAY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, len(headers))
        row += 1
        for h, ampli, acou in donnees:
            ws.cell(row=row, column=1, value=h)
            ws.cell(row=row, column=2, value=ampli)
            ws.cell(row=row, column=3, value=acou)
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = BORDER
            row += 1
        row += 1

    autosize(ws, max_width=50)
    ws.row_dimensions[1].height = 22


# ============================================
# MAIN
# ============================================
def main():
    wb = Workbook()
    wb.remove(wb.active)
    write_artistes(wb)
    write_compagnies(wb)
    write_programmation(wb)
    write_logistique(wb)
    wb.save(OUT)
    print(f"OK : {OUT.relative_to(REPO)} ({OUT.stat().st_size // 1024} KB)")
    print(f"Onglets : {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  - {sn}: {ws.max_row} lignes x {ws.max_column} cols")


if __name__ == "__main__":
    main()
