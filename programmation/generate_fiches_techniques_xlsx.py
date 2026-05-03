"""
Génère le fichier Excel Fiches_techniques_Monique_Festival.xlsx
1 ligne par spectacle/atelier, 1 colonne par champ de la fiche technique.
Première colonne = statut de complétude.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Fiches techniques"

HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PARTIEL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
MANQUANT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ATELIER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
DAY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

headers = [
    "Statut fiche",             # A
    "Données individuelles",    # B
    "Données administratives",  # C
    "Données techniques",       # D
    "Jour",                     # D
    "Horaire",                  # E
    "Type",                     # F - Spectacle/Atelier
    "Titre",                    # G
    "Nom de scène",             # H
    "Artistes",                 # I
    "Compagnie / structure",    # J
    "Nb artistes plateau",      # K
    "Nb techniciens",           # L
    "Durée",                    # M
    "Public cible",             # N
    "Scène",                    # O
    "Instruments",              # P
    "Besoins son",              # Q
    "Besoins lumière",          # R
    "Temps montage",            # S
    "Temps démontage",          # T
    "Plan de scène / rider",    # U
    "Mode de rémunération",     # V
    "N° licence spectacle",     # W
    "N° SIRET / RNA",           # X
    "Attestation RC pro",       # Y
    "RIB reçu",                 # Z
    "Contrat signé",            # AA
    "Fiche renseignements",     # AB
    "Notes",                    # AC
]

NC = len(headers)

# Title row
ws.cell(row=1, column=1, value="FICHES TECHNIQUES — Monique Festival 2026")
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="2F5496")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)

# Headers
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = THIN_BORDER

# Helper
OK = "OK"
TODO = "À COMPLÉTER"
VERIF = "À VÉRIFIER"
FOURNIR = "À FOURNIR"
NA = "N/A"
APRES = "Après sélection"

# (statut, jour, horaire, type, titre, nom_scene, artistes, compagnie,
#  nb_artistes, nb_tech, durée, public, scène, instruments,
#  son, lumière, montage, démontage, rider, remun,
#  licence, siret, rc, rib, contrat, fiche_rens, notes)

# statut_fiche, admin, tech, jour, horaire, type, titre, nom_scene, artistes, compagnie,
# nb_artistes, nb_tech, durée, public, scène, instruments,
# son, lumière, montage, démontage, rider, remun,
# licence, siret, rc, rib, contrat, fiche_rens, notes

INCOMP = "Incomplet"
COMP = "Complet"

fiches = [
    # VENDREDI
    ("PARTIEL", INCOMP, INCOMP, INCOMP, "Vendredi", "16h-16h30", "Spectacle", "Duo Cabaret — Reprise de variétés", "Duo Cabaret",
     "Romane Cabaret + George Cabaret", "", 2, 0, "30 min", "Tous publics", "Amplifiée/restauration",
     "Piano ou guitare + voix", "Sonorisation requise", TODO, "10 min", "10 min",
     FOURNIR, "George=cachet, Romane=facture Fabulo", VERIF, TODO, VERIF, TODO, TODO, TODO, ""),

    ("PARTIEL", INCOMP, INCOMP, COMP, "Vendredi", "17h-17h45", "Spectacle", "Concert Fabulo + introduction", "Fabulo",
     "Lorraine, Marius, Maya, Raphaëlle, Romane", "Collectif Fabulo", 5, 0, "45 min", "Tous publics",
     "Acoustique (cour)", "Violoncelle, voix, instruments divers", "Non", "Jour", "10 min", "10 min",
     FOURNIR, "Facture collectif", VERIF, TODO, VERIF, TODO, TODO, TODO, "Lorraine = famille"),

    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Vendredi", "18h15-19h", "Spectacle", "Comment on en est arrivés là ?", "Cie Treize Clique",
     "Marius Ponnelle + 2 comédiens", "Cie Treize Clique", 3, TODO, "45 min", TODO, "Acoustique",
     TODO, TODO, TODO, TODO, TODO, FOURNIR, "Facture collectif", VERIF, TODO, VERIF, TODO, TODO, TODO, ""),

    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Vendredi", "19h30-20h15", "Spectacle", "À définir", "? (non attribué)",
     "À définir", "", 1, "", "", "", "Amplifiée", "", "", "", "", "",
     "", "", "", "", "", "", "", "", "Créneau non attribué"),

    ("PARTIEL", INCOMP, INCOMP, INCOMP, "Vendredi", "20h45-21h45", "Spectacle", "Luther & Loretta", "Luther et Loretta",
     "Luther + Loretta", "", 2, 0, "1h", "Tous publics", "Amplifiée",
     "Piano, micro, carte son", "Sonorisation requise", TODO, "10-15 min", "10-15 min",
     FOURNIR, "Luther=famille, Loretta=cachet", NA, NA, NA, TODO, TODO, TODO, ""),

    ("PARTIEL", INCOMP, INCOMP, COMP, "Vendredi", "22h15-23h15", "Spectacle", "Concert Wet Enough!?", "Wet Enough",
     "Laszlo, Marius, Maël, Baptiste, Matthieu", "Collectif", 5, 0, "1h", "Tous publics", "Amplifiée",
     "Basse, batterie, guitare, chant, clavier, trombone", "Sonorisation requise", TODO,
     "10-15 min", "10-15 min", "CT fournie (PRESS KIT)", "Facture collectif", VERIF, TODO, VERIF, TODO, TODO, TODO, ""),

    # SAMEDI — ATELIERS
    ("OK", COMP, COMP, COMP, "Samedi", "10h-12h", "Atelier", "Atelier Fabulo — musique et théâtre", "Fabulo",
     "Lorraine, Marius, Maya, Raphaëlle, Romane", "Collectif Fabulo", 5, 0, "2h", "30 participants, 50€/p",
     "Devant scène amplifiée", "", "", "", "", "",
     NA, "Répétition — facture Fabulo", NA, NA, NA, NA, NA, NA, "Répétition Création Monique"),

    ("OK", COMP, COMP, COMP, "Samedi", "10h-12h15", "Atelier", "Masterclass chant lyrique", "Fabulo",
     "Romane + pianiste + élèves", "Collectif Fabulo", 3, 0, "2h15", "4 élèves, 60€/p",
     "Acoustique", "Voix, piano", "", "", "", "",
     NA, "Incl. dans facture Fabulo", NA, NA, NA, NA, NA, NA, "Piano droit ok"),

    ("OK", COMP, COMP, COMP, "Samedi", "10h-12h", "Atelier", "Atelier éloquence", "Lewis Janier-Dubry",
     "Lewis", "", 1, 0, "2h", "Ouvert", "Devant scène amplifiée", "", "", "", "", "",
     NA, "Famille (0€)", NA, NA, NA, NA, NA, NA, ""),

    ("OK", INCOMP, INCOMP, COMP, "Samedi", "10h-12h", "Atelier", "Atelier écriture", "Sylvain Septours",
     "Sylvain", "", 1, 0, "2h", "15 participants, 50€/p", "Devant scène amplifiée", "", "", "", "", "",
     NA, "Cachet répétition", NA, NA, NA, NA, NA, TODO, ""),

    ("PARTIEL", INCOMP, INCOMP, INCOMP, "Samedi", "10h-12h", "Atelier", "Atelier DJ / sound design", "Jeff The Fool",
     "Jeff + Léopold", "", 2, 0, "2h", TODO, TODO, "Platines, enceintes, matériel DJ",
     "", "", "", "", NA, "Jeff=cachet rép, Léopold=famille", NA, NA, NA, NA, NA, TODO, ""),

    # SAMEDI — SPECTACLES
    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Samedi", "12h45-13h30", "Spectacle", "Concert Wambo", "Wambo",
     "Manu", "Wambo Productions", 1, 0, "45 min", "Tous publics", "Amplifiée",
     TODO, "Sonorisation requise", "Jour", TODO, TODO, FOURNIR, "Cachet individuel", VERIF, TODO, VERIF, TODO, TODO, TODO, ""),

    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Samedi", "14h-15h", "Spectacle", "Cie de théâtre (appel à projet)", "Cie théâtre 1",
     APRES, APRES, 4, TODO, "1h", TODO, "Acoustique",
     TODO, TODO, TODO, TODO, TODO, APRES, "Cachet individuel", APRES, APRES, APRES, APRES, APRES, APRES, "Appel à projet"),

    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Samedi", "15h30-16h40", "Spectacle", "À travers les fenêtres des heures", "À travers les fenêtres des heures",
     TODO, TODO, 2, TODO, "1h10", TODO, "Acoustique",
     TODO, TODO, TODO, TODO, TODO, "Dossier FdH disponible", "Cachet individuel", VERIF, TODO, VERIF, TODO, TODO, TODO, "Nouveau"),

    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Samedi", "17h10-18h10", "Spectacle", "Concert Fannelie", "Fannelie",
     "Fannelie", TODO, 1, TODO, "1h", "Tous publics", "Amplifiée",
     TODO, TODO, TODO, TODO, TODO, "Bio disponible", "Cachet individuel", VERIF, TODO, VERIF, TODO, TODO, TODO, "Nouveau"),

    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Samedi", "18h40-19h30", "Spectacle", "Cie de théâtre (appel à projet)", "Cie théâtre 2",
     APRES, APRES, 4, TODO, "50 min", TODO, "Acoustique",
     TODO, TODO, TODO, TODO, TODO, APRES, "Cachet individuel", APRES, APRES, APRES, APRES, APRES, APRES, "Appel à projet"),

    ("PARTIEL", INCOMP, INCOMP, COMP, "Samedi", "20h-21h10", "Spectacle", "Europe Cellists meet Augustin", "Europe Cellists",
     "Marc Trembovelski, Jakub Wyciślik, Sophie Ehling, Violoncelliste 4, Augustin Lemonnier",
     "", 5, 0, "1h10", "Tous publics", "Acoustique",
     "4 violoncelles + piano", "Non", TODO, "5 min / 20 min si piano", "5 min",
     "Description concert dispo", "Cachet individuel", NA, NA, NA, TODO, TODO, TODO, ""),

    ("PARTIEL", INCOMP, INCOMP, COMP, "Samedi", "21h35-22h35", "Spectacle", "Concert Osmosis", "Osmosis",
     "Maël, Paul, Baptiste", "Collectif", 3, 0, "1h", "Tous publics", "Amplifiée",
     "Basse, batterie, clavier", "Sonorisation requise", TODO, "10-15 min", "10-15 min",
     "Press kit disponible", "Facture collectif", VERIF, TODO, VERIF, TODO, TODO, TODO, ""),

    # DIMANCHE — ATELIERS
    ("OK", COMP, COMP, COMP, "Dimanche", "10h-12h", "Atelier", "Ateliers (tous) + générale 11h30", "Fabulo + intervenants",
     "Lorraine, Marius, Maya, Raphaëlle, Romane, Lewis, Sylvain, Jeff, Léopold", "Collectif Fabulo",
     9, 0, "2h", "Participants inscrits", "Amplifiée + acoustique", "", "", "", "", "",
     NA, "Voir ateliers samedi", NA, NA, NA, NA, NA, NA, "Générale à 11h30 pour Création Monique"),

    # DIMANCHE — SPECTACLES
    ("OK", INCOMP, INCOMP, COMP, "Dimanche", "12h15-14h", "Spectacle", "Suites de Bach", "Europe Cellists",
     "Marc T., Jakub W., Sophie E., Violoncelliste 4, Lorraine, Volodia",
     "", 6, 0, "~1h45", "Tous publics", "Acoustique / restauration",
     "Violoncelle solo", "Non", "Non (jour)", "5 min", "5 min",
     OK, "Cachet indiv (sauf Lorraine=fam)", NA, NA, NA, TODO, TODO, TODO,
     "Suites 1-6: 15/20/25/30/30/40 min"),

    ("OK", INCOMP, INCOMP, COMP, "Dimanche", "14h30-15h", "Spectacle", "Volodia Van Keulen — Sitariste", "Volodia Van Keulen",
     "Volodia", "", 1, 0, "30 min", "Tous publics", "Acoustique",
     "Sitar", "Non", "Non", "5 min", "5 min",
     OK, "Cachet individuel", NA, NA, NA, TODO, TODO, TODO, ""),

    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Dimanche", "15h30-16h30", "Spectacle", "Cie de théâtre (appel à projet)", "Cie théâtre 3",
     APRES, APRES, 4, TODO, "1h", TODO, TODO,
     TODO, TODO, TODO, TODO, TODO, APRES, "Cachet individuel", APRES, APRES, APRES, APRES, APRES, APRES, "Appel à projet"),

    ("OK", COMP, COMP, COMP, "Dimanche", "17h-17h45", "Spectacle", "Nos voix", "Nos Voix",
     "Romane Cabaret + Lorraine", "", 2, 0, "45 min", "Tous publics", "Acoustique",
     "Violoncelle, voix", "Non", "Guirlandes lumineuses", "10-15 min", "10-15 min",
     OK, "Romane=facture Fabulo, Lorraine=famille", NA, NA, NA, NA, NA, NA,
     "2 chaises, 2 pupitres, électricité guirlandes"),

    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Dimanche", "18h15-19h", "Spectacle", "Concert Karina Ziegler", "Karina Ziegler",
     "Carina Zeiger + 2 musiciens", "Ensemble", 3, TODO, "45 min", "Tous publics", "Amplifiée",
     TODO, TODO, TODO, TODO, TODO, FOURNIR, "Facture collectif", VERIF, TODO, VERIF, TODO, TODO, TODO, ""),

    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Dimanche", "19h30-20h15", "Spectacle", "Cie de théâtre (appel à projet)", "Cie théâtre 4",
     APRES, APRES, 4, TODO, "45 min", TODO, "Acoustique",
     TODO, TODO, TODO, TODO, TODO, APRES, "Cachet individuel", APRES, APRES, APRES, APRES, APRES, APRES, "Appel à projet"),

    ("MANQUANT", INCOMP, INCOMP, INCOMP, "Dimanche", "20h30-21h", "Spectacle", "Création Monique — Restitution", "Fabulo",
     "Lorraine, Marius, Maya, Raphaëlle, Romane + participants", "Fabulo / JD Production",
     "5+", TODO, "30 min", "Tous publics", "Amplifiée",
     TODO, TODO, TODO, TODO, TODO, TODO, "Facture collectif Fabulo", VERIF, TODO, VERIF, NA, TODO, TODO,
     "Production cat.2 — convention de création nécessaire"),
]

r = 3
current_day = ""
for fiche in fiches:
    statut = fiche[0]
    jour = fiche[1]

    # Séparateur de jour
    if jour != current_day:
        current_day = jour
        ws.cell(row=r, column=1, value=jour.upper())
        ws.cell(row=r, column=1).font = Font(bold=True, size=12, color="2F5496")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        for c in range(1, NC + 1):
            ws.cell(row=r, column=c).fill = DAY_FILL
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    for c, val in enumerate(fiche, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Recalculer le statut global : OK seulement si indiv + admin + technique sont tous Complet
    indiv_val = fiche[1]
    admin_val = fiche[2]
    tech_val = fiche[3]
    all_comp = [indiv_val, admin_val, tech_val]
    if all(v == COMP for v in all_comp):
        statut = "OK"
    elif all(v == INCOMP for v in all_comp):
        statut = "MANQUANT"
    else:
        statut = "PARTIEL"

    # Écrire le statut recalculé
    ws.cell(row=r, column=1, value=statut)

    # Coloration statut (col A)
    statut_cell = ws.cell(row=r, column=1)
    if statut == "OK":
        statut_cell.fill = OK_FILL
    elif statut == "PARTIEL":
        statut_cell.fill = PARTIEL_FILL
    elif statut == "MANQUANT":
        statut_cell.fill = MANQUANT_FILL

    # Coloration indiv (col B), admin (col C), technique (col D)
    for col_idx in [2, 3, 4]:
        cell = ws.cell(row=r, column=col_idx)
        if cell.value == COMP:
            cell.fill = OK_FILL
        elif cell.value == INCOMP:
            cell.fill = MANQUANT_FILL

    # Coloration atelier
    if fiche[6] == "Atelier":  # col G = type (index 6 dans le tuple)
        for c in range(5, NC + 1):
            ws.cell(row=r, column=c).fill = ATELIER_FILL

    # Coloration des cellules À COMPLÉTER / À FOURNIR / À VÉRIFIER
    for c in range(5, NC + 1):
        val = ws.cell(row=r, column=c).value
        if val in (TODO, FOURNIR, VERIF, APRES):
            ws.cell(row=r, column=c).font = Font(color="FF0000", italic=True, size=10)

    r += 1

# Légende
r += 2
ws.cell(row=r, column=1, value="LÉGENDE STATUT").font = Font(bold=True)
r += 1
ws.cell(row=r, column=1, value="OK").fill = OK_FILL
ws.cell(row=r, column=2, value="Fiche technique complète ou quasi-complète")
r += 1
ws.cell(row=r, column=1, value="PARTIEL").fill = PARTIEL_FILL
ws.cell(row=r, column=2, value="Fiche partiellement remplie — quelques champs manquants")
r += 1
ws.cell(row=r, column=1, value="MANQUANT").fill = MANQUANT_FILL
ws.cell(row=r, column=2, value="Fiche largement incomplète — action requise")

# Statistiques
r += 2
ws.cell(row=r, column=1, value="STATISTIQUES").font = Font(bold=True, size=12)
r += 1
nb_ok = sum(1 for f in fiches if f[0] == "OK")
nb_partiel = sum(1 for f in fiches if f[0] == "PARTIEL")
nb_manquant = sum(1 for f in fiches if f[0] == "MANQUANT")
ws.cell(row=r, column=1, value=f"Fiches OK : {nb_ok}")
ws.cell(row=r, column=1).fill = OK_FILL
r += 1
ws.cell(row=r, column=1, value=f"Fiches partielles : {nb_partiel}")
ws.cell(row=r, column=1).fill = PARTIEL_FILL
r += 1
ws.cell(row=r, column=1, value=f"Fiches manquantes : {nb_manquant}")
ws.cell(row=r, column=1).fill = MANQUANT_FILL
r += 1
ws.cell(row=r, column=1, value=f"Total : {len(fiches)} prestations")

# Auto-width
for col_cells in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col_cells[0].column)
    for cell in col_cells:
        if cell.value and not isinstance(cell.value, int):
            max_len = max(max_len, min(len(str(cell.value)), 30))
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Figer les volets
ws.freeze_panes = "B3"

output = "c:/Users/charl/Monique-Festival/programmation/Fiches_techniques_Monique_Festival.xlsx"
wb.save(output)
print(f"Fichier créé : {output}")
print(f"Stats : {nb_ok} OK, {nb_partiel} partielles, {nb_manquant} manquantes / {len(fiches)} total")
